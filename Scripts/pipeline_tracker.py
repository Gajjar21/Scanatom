# Scripts/pipeline_tracker.py
# Standalone pipeline processing-time tracker.
# Tracks ACTIVE processing time only — idle/wait time is excluded.
#
# Stages tracked:
#   1. HOT_START  → HOT_END    : AWB hotfolder (OCR + match)
#   2. EDM_START  → EDM_END    : EDM duplicate check
#   3. BATCH_ADDED             : File added to a print batch
#
# All paths come from config.py — no hardcoded values here.

import os
import sys
import time
import atexit
from datetime import datetime
from pathlib import Path

# Allow running this file directly from the Scripts/ subfolder
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
import config
from openpyxl import load_workbook, Workbook

TRACKER_PATH = config.TRACKER_PATH

HEADERS = [
    "AWB",                    # A
    "OriginalFilename",       # B
    "ProcessedFilename",      # C
    "HotFolder_Start",        # D
    "HotFolder_End",          # E
    "HotFolder_Secs",         # F
    "MatchMethod",            # G
    "EDM_Start",              # H
    "EDM_End",                # I
    "EDM_Secs",               # J
    "EDM_Result",             # K
    "Batch_Added_Time",       # L
    "Batch_Number",           # M
    "Total_Processing_Secs",  # N
    "Total_Processing_HMS",   # O
    "Status",                 # P
    "Notes",                  # Q
]

_COL = {h: i + 1 for i, h in enumerate(HEADERS)}
_WB_CACHE = {"wb": None, "ws": None, "mtime": None}


# ── Helpers ───────────────────────────────────────────────────────────────────

def _now_str():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def _secs_between(start_str, end_str):
    fmt = "%Y-%m-%d %H:%M:%S"
    try:
        s = datetime.strptime(start_str, fmt)
        e = datetime.strptime(end_str, fmt)
        return max(0, round((e - s).total_seconds(), 1))
    except Exception:
        return None


def _secs_to_hms(secs):
    if secs is None:
        return None
    secs = int(secs)
    h = secs // 3600
    m = (secs % 3600) // 60
    s = secs % 60
    return f"{h:02d}:{m:02d}:{s:02d}"


def _load_or_create():
    TRACKER_PATH.parent.mkdir(parents=True, exist_ok=True)
    current_mtime = TRACKER_PATH.stat().st_mtime if TRACKER_PATH.exists() else None

    # Reuse cached workbook only when on-disk tracker mtime is unchanged.
    if _WB_CACHE["wb"] is not None and _WB_CACHE["mtime"] == current_mtime:
        return _WB_CACHE["wb"], _WB_CACHE["ws"]

    if _WB_CACHE["wb"] is not None:
        try:
            _WB_CACHE["wb"].close()
        except Exception:
            pass
        _WB_CACHE["wb"] = None
        _WB_CACHE["ws"] = None
        _WB_CACHE["mtime"] = None

    if TRACKER_PATH.exists():
        wb = load_workbook(TRACKER_PATH)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Pipeline Tracker"
        ws.append(HEADERS)
        ws.freeze_panes = "A2"
        col_widths = {
            "A": 15, "B": 30, "C": 20, "D": 22, "E": 22,
            "F": 14, "G": 20, "H": 22, "I": 22, "J": 14,
            "K": 18, "L": 22, "M": 14, "N": 22, "O": 20,
            "P": 16, "Q": 40,
        }
        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = width
        wb.save(TRACKER_PATH)
        current_mtime = TRACKER_PATH.stat().st_mtime

    _WB_CACHE["wb"] = wb
    _WB_CACHE["ws"] = ws
    _WB_CACHE["mtime"] = current_mtime
    return wb, ws


def _find_row(ws, awb=None, original_filename=None, processed_filename=None):
    best_row = None
    for row in range(2, ws.max_row + 1):
        status = ws.cell(row, _COL["Status"]).value
        if status not in ("IN-PROGRESS", None):
            continue
        if awb and ws.cell(row, _COL["AWB"]).value == awb:
            best_row = row
        elif original_filename and ws.cell(row, _COL["OriginalFilename"]).value == original_filename:
            best_row = row
        elif processed_filename and ws.cell(row, _COL["ProcessedFilename"]).value == processed_filename:
            best_row = row
    return best_row


def _set(ws, row, col_name, value):
    ws.cell(row, _COL[col_name]).value = value


def _get(ws, row, col_name):
    return ws.cell(row, _COL[col_name]).value


def _recalc_total(ws, row):
    hot_secs = _get(ws, row, "HotFolder_Secs") or 0
    edm_secs = _get(ws, row, "EDM_Secs") or 0
    total = hot_secs + edm_secs
    _set(ws, row, "Total_Processing_Secs", total)
    _set(ws, row, "Total_Processing_HMS", _secs_to_hms(total))


def _retry_save(wb, retries=5):
    for attempt in range(retries):
        try:
            wb.save(TRACKER_PATH)
            try:
                _WB_CACHE["mtime"] = TRACKER_PATH.stat().st_mtime
            except Exception:
                _WB_CACHE["mtime"] = None
            return True
        except PermissionError:
            time.sleep(0.4 * (attempt + 1))
        except Exception:
            return False
    return False


def _close_cached_wb():
    wb = _WB_CACHE.get("wb")
    if wb is not None:
        try:
            wb.close()
        except Exception:
            pass


atexit.register(_close_cached_wb)


def _with_retry(fn):
    for attempt in range(5):
        try:
            fn()
            return
        except PermissionError:
            time.sleep(0.4 * (attempt + 1))
        except Exception:
            return


# ── Public API ────────────────────────────────────────────────────────────────

def record_hotfolder_start(original_filename):
    def _write():
        wb, ws = _load_or_create()
        ws.append([
            None, original_filename, None,
            _now_str(), None, None, None,
            None, None, None, None,
            None, None, None, None,
            "IN-PROGRESS", None,
        ])
        _retry_save(wb)
    _with_retry(_write)


def record_hotfolder_end(original_filename, awb, processed_filename,
                         match_method, notes=None):
    def _write():
        wb, ws = _load_or_create()
        row = _find_row(ws, original_filename=original_filename)

        if row is None:
            ws.append([
                awb, original_filename, processed_filename,
                None, _now_str(), None, match_method,
                None, None, None, None,
                None, None, None, None,
                "IN-PROGRESS", notes,
            ])
            _retry_save(wb)
            return

        end_str = _now_str()
        start_str = _get(ws, row, "HotFolder_Start")
        secs = _secs_between(start_str, end_str) if start_str else None

        _set(ws, row, "AWB", awb)
        _set(ws, row, "ProcessedFilename", processed_filename)
        _set(ws, row, "HotFolder_End", end_str)
        _set(ws, row, "HotFolder_Secs", secs)
        _set(ws, row, "MatchMethod", match_method)
        if notes:
            _set(ws, row, "Notes", notes)
        _recalc_total(ws, row)
        _retry_save(wb)
    _with_retry(_write)


def record_hotfolder_needs_review(original_filename, reason):
    def _write():
        wb, ws = _load_or_create()
        row = _find_row(ws, original_filename=original_filename)
        end_str = _now_str()

        if row is None:
            ws.append([
                None, original_filename, None,
                None, end_str, None, "No Match",
                None, None, None, None,
                None, None, None, None,
                "NEEDS-REVIEW", reason,
            ])
            _retry_save(wb)
            return

        start_str = _get(ws, row, "HotFolder_Start")
        secs = _secs_between(start_str, end_str) if start_str else None

        _set(ws, row, "HotFolder_End", end_str)
        _set(ws, row, "HotFolder_Secs", secs)
        _set(ws, row, "MatchMethod", "No Match")
        _set(ws, row, "Status", "NEEDS-REVIEW")
        _set(ws, row, "Notes", reason)
        _recalc_total(ws, row)
        _retry_save(wb)
    _with_retry(_write)


def record_edm_start(processed_filename):
    def _write():
        wb, ws = _load_or_create()
        row = _find_row(ws, processed_filename=processed_filename)

        if row is None:
            awb = os.path.splitext(processed_filename)[0]
            ws.append([
                awb, None, processed_filename,
                None, None, None, None,
                _now_str(), None, None, None,
                None, None, None, None,
                "IN-PROGRESS", "No hotfolder record found",
            ])
        else:
            _set(ws, row, "EDM_Start", _now_str())
        _retry_save(wb)
    _with_retry(_write)


def record_edm_end(processed_filename, edm_result, final_folder, notes=None):
    def _write():
        wb, ws = _load_or_create()
        row = _find_row(ws, processed_filename=processed_filename)
        end_str = _now_str()

        status_map = {
            "CLEAN":           "IN-PROGRESS",
            "CLEAN-UNCHECKED": "IN-PROGRESS",
            "PARTIAL-CLEAN":   "IN-PROGRESS",
            "REJECTED":        "REJECTED",
            "NEEDS-REVIEW":    "NEEDS-REVIEW",
        }
        status = status_map.get(edm_result, "IN-PROGRESS")

        if row is None:
            awb = os.path.splitext(processed_filename)[0]
            ws.append([
                awb, None, processed_filename,
                None, None, None, None,
                None, end_str, None, edm_result,
                None, None, None, None,
                status, notes,
            ])
            _retry_save(wb)
            return

        start_str = _get(ws, row, "EDM_Start")
        secs = _secs_between(start_str, end_str) if start_str else None

        _set(ws, row, "EDM_End", end_str)
        _set(ws, row, "EDM_Secs", secs)
        _set(ws, row, "EDM_Result", edm_result)
        _set(ws, row, "Status", status)
        if notes:
            existing = _get(ws, row, "Notes") or ""
            _set(ws, row, "Notes", (existing + " | " + notes).strip(" | "))
        _recalc_total(ws, row)
        _retry_save(wb)
    _with_retry(_write)


def record_batch_added(awb, batch_number):
    def _write():
        wb, ws = _load_or_create()
        row = _find_row(ws, awb=awb)

        if row is None:
            ws.append([
                awb, None, f"{awb}.pdf",
                None, None, None, None,
                None, None, None, None,
                _now_str(), batch_number,
                None, None,
                "COMPLETE", None,
            ])
            _retry_save(wb)
            return

        _set(ws, row, "Batch_Added_Time", _now_str())
        _set(ws, row, "Batch_Number", batch_number)
        _set(ws, row, "Status", "COMPLETE")
        _recalc_total(ws, row)
        _retry_save(wb)
    _with_retry(_write)


# ── Reporting ─────────────────────────────────────────────────────────────────

def get_summary():
    counts = {
        "TOTAL": 0, "COMPLETE": 0, "IN-PROGRESS": 0,
        "REJECTED": 0, "NEEDS-REVIEW": 0,
    }
    total_secs_list = []
    try:
        wb, ws = _load_or_create()
        for row in range(2, ws.max_row + 1):
            status = _get(ws, row, "Status")
            if not status:
                continue
            counts["TOTAL"] += 1
            if status in counts:
                counts[status] += 1
            secs = _get(ws, row, "Total_Processing_Secs")
            if secs and status == "COMPLETE":
                total_secs_list.append(secs)
    except Exception:
        pass
    counts["Avg_Processing_HMS"] = (
        _secs_to_hms(sum(total_secs_list) / len(total_secs_list))
        if total_secs_list else "N/A"
    )
    return counts


def get_stale_records(stale_threshold_seconds=600):
    stale = []
    cutoff = time.time() - stale_threshold_seconds
    try:
        wb, ws = _load_or_create()
        for row in range(2, ws.max_row + 1):
            status = _get(ws, row, "Status")
            hot_start = _get(ws, row, "HotFolder_Start")
            filename = _get(ws, row, "OriginalFilename")
            awb = _get(ws, row, "AWB")
            if status != "IN-PROGRESS":
                continue
            if hot_start:
                try:
                    ts = datetime.strptime(hot_start, "%Y-%m-%d %H:%M:%S").timestamp()
                    if ts < cutoff:
                        stale.append({"filename": filename, "awb": awb, "since": hot_start})
                except Exception:
                    pass
    except Exception:
        pass
    return stale


# ── Self-test ─────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("Testing pipeline_tracker.py...")
    record_hotfolder_start("test_invoice.pdf")
    time.sleep(1)
    record_hotfolder_end("test_invoice.pdf", "123456789012", "123456789012.pdf", "OCR-Main")
    time.sleep(1)
    record_edm_start("123456789012.pdf")
    time.sleep(1)
    record_edm_end("123456789012.pdf", "CLEAN", "CLEAN")
    time.sleep(1)
    record_batch_added("123456789012", batch_number=1)
    print("Summary:", get_summary())
    print(f"Tracker saved to: {TRACKER_PATH}")
