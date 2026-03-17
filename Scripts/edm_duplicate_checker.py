# Scripts/edm_duplicate_checker.py
# EDM Duplicate Checker
#
# Watches PROCESSED folder for new PDFs.
# For each file:
#   1. Queries FedEx EDM API using the AWB number (from filename).
#   2. Downloads existing documents.
#   3. Compares pages (exact hash -> perceptual hash -> text similarity).
#   4. Routes to CLEAN (new), REJECTED (all-duplicate), or splits partial matches.
#
# All paths and API settings come from config.py / .env.

import os
import re
import sys
import time
import csv
import atexit
import hashlib
import logging
import requests
import zipfile
import uuid
import io
import shutil
import threading
from collections import Counter
import fitz
from pathlib import Path
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
from datetime import datetime
from openpyxl import load_workbook, Workbook

# Allow running from Scripts/ subfolder
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
import config
from Scripts.pipeline_tracker import record_edm_start, record_edm_end
from Scripts.audit_logger import audit_event

# ── Paths from config ─────────────────────────────────────────────────────────
PROCESSED_FOLDER    = config.PROCESSED_DIR
CLEAN_FOLDER        = config.CLEAN_DIR
REJECTED_FOLDER     = config.REJECTED_DIR
NEEDS_REVIEW_FOLDER = config.NEEDS_REVIEW_DIR
AWB_LOGS_PATH       = config.AWB_LOGS_PATH
CSV_PATH            = config.CSV_PATH
TESSERACT_PATH      = str(config.TESSERACT_PATH)
STAGE_CACHE_CSV     = config.STAGE_CACHE_CSV
PIPELINE_SUMMARY_CSV = config.PIPELINE_SUMMARY_CSV

# ── EDM API settings from config ──────────────────────────────────────────────
OPERATING_COMPANY = config.EDM_OPERATING_COMPANY
METADATA_URL      = config.EDM_METADATA_URL
DOWNLOAD_URL      = config.EDM_DOWNLOAD_URL

# ── Tuning from config ────────────────────────────────────────────────────────
FILE_SETTLE_SECONDS         = config.FILE_SETTLE_SECONDS
TEXT_SIMILARITY_THRESHOLD   = config.TEXT_SIMILARITY_THRESHOLD
OCR_TOP_PERCENT             = config.OCR_TOP_PERCENT
PHASH_THRESHOLD             = config.PHASH_THRESHOLD
PAGE_OCR_LIMIT              = config.PAGE_OCR_LIMIT
MIN_EMBEDDED_TEXT_LENGTH    = config.MIN_EMBEDDED_TEXT_LENGTH
EARLY_FOCUS_MATCH_THRESHOLD = config.EARLY_FOCUS_MATCH_THRESHOLD
OCR_COMPARE_LIMIT           = 10
REJECT_IF_DUP_PAGES_OVER    = 5
REJECT_IF_DUP_RATIO         = 0.70

# ── Logging ───────────────────────────────────────────────────────────────────
config.LOG_DIR.mkdir(parents=True, exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[
        logging.FileHandler(config.EDM_LOG),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger("EDMChecker")

# ── AWB session cache (current AWB only) ─────────────────────────────────────
AWB_SESSION_CACHE = {
    "awb": None,
    "doc_ids": None,
    "edm_pdf_list": None,
    "edm_fingerprints": None,
}

_SUMMARY_HEADERS = [
    "Timestamp",
    "InputFileName",
    "AWB_Detected",
    "AWB_Detection_Type",
    "EDM_Check_Status",
    "Duplicate_Detection_Type",
    "Detection_Type_Match_Score",
    "AWB_Extraction_Seconds",
    "EDM_Check_Minutes",
    "TOTAL_Minutes_AWB_plus_EDM",
    "Total_Pages",
    "Duplicate_Pages",
    "Pages_To_Clean",
    "Decision_Trace",
]
_SUMMARY_QUEUE = []
_SUMMARY_LOCK = threading.Lock()
_SUMMARY_LAST_FLUSH = time.time()
_SUMMARY_FLUSH_INTERVAL_SECONDS = 5
_SUMMARY_FLUSH_BATCH_SIZE = 20
_STAGE_CACHE_INDEX = {"mtime": None, "rows": {}}


def _flush_summary_queue(force=False):
    global _SUMMARY_LAST_FLUSH
    with _SUMMARY_LOCK:
        if not _SUMMARY_QUEUE:
            return
        if not force:
            if len(_SUMMARY_QUEUE) < _SUMMARY_FLUSH_BATCH_SIZE and (time.time() - _SUMMARY_LAST_FLUSH) < _SUMMARY_FLUSH_INTERVAL_SECONDS:
                return
        rows = list(_SUMMARY_QUEUE)
        _SUMMARY_QUEUE.clear()
        _SUMMARY_LAST_FLUSH = time.time()

    try:
        PIPELINE_SUMMARY_CSV.parent.mkdir(parents=True, exist_ok=True)
        new_file = not PIPELINE_SUMMARY_CSV.exists()
        with open(PIPELINE_SUMMARY_CSV, "a", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=_SUMMARY_HEADERS)
            if new_file:
                w.writeheader()
            w.writerows(rows)
    except Exception as e:
        log.warning(f"[SUMMARY] Could not flush summary log: {e}")
        with _SUMMARY_LOCK:
            _SUMMARY_QUEUE[:0] = rows


def _queue_summary_row(row):
    with _SUMMARY_LOCK:
        _SUMMARY_QUEUE.append(row)
    _flush_summary_queue(force=False)


atexit.register(lambda: _flush_summary_queue(force=True))


def _clear_awb_cache(reason=""):
    prev = AWB_SESSION_CACHE.get("awb")
    if prev:
        if reason:
            log.info(f"[CACHE] Clearing AWB cache for {prev}: {reason}")
        else:
            log.info(f"[CACHE] Clearing AWB cache for {prev}")
    AWB_SESSION_CACHE["awb"] = None
    AWB_SESSION_CACHE["doc_ids"] = None
    AWB_SESSION_CACHE["edm_pdf_list"] = None
    AWB_SESSION_CACHE["edm_fingerprints"] = None


def _get_stage_cache_row(processed_filename):
    """Fetch hotfolder stage metadata by processed filename with mtime-based cache."""
    try:
        if not STAGE_CACHE_CSV.exists():
            return None
        mtime = STAGE_CACHE_CSV.stat().st_mtime
        if _STAGE_CACHE_INDEX["mtime"] != mtime:
            rows = {}
            with open(STAGE_CACHE_CSV, "r", encoding="utf-8", newline="") as f:
                for r in csv.DictReader(f):
                    key = (r.get("ProcessedFileName") or "").strip()
                    if key:
                        rows[key] = r
            _STAGE_CACHE_INDEX["rows"] = rows
            _STAGE_CACHE_INDEX["mtime"] = mtime
        return _STAGE_CACHE_INDEX["rows"].get(processed_filename)
    except Exception as e:
        log.warning(f"[STAGE_CACHE] Could not read stage cache: {e}")
        return None


# =========================
# AWB EXTRACTION FROM FILENAME
# Strips _2, _3 suffix so 123456789012_2.pdf queries AWB 123456789012
# =========================
def _awb_from_processed_filename(filename):
    base = os.path.splitext(filename)[0]
    # Primary pattern: 123456789012.pdf or 123456789012_2.pdf
    m = re.match(r"^(\d{12})(?:_\d+)?$", base)
    if m:
        return m.group(1)
    # Fallback: keep backward compatibility with older names
    m = re.match(r"^(\d{12})", base)
    return m.group(1) if m else None


def _ms(start_ts):
    return round((time.perf_counter() - start_ts) * 1000, 1)


def _log_timing(awb, filename, t):
    log.info(
        "[TIMING] file=%s awb=%s cache=%s metadata_ms=%.1f download_ms=%.1f extract_ms=%.1f "
        "compare_ms=%.1f route_ms=%.1f total_active_ms=%.1f",
        filename, awb, t.get("cache", "MISS"), t["metadata_ms"], t["download_ms"], t["extract_ms"],
        t["compare_ms"], t["route_ms"], t["total_active_ms"],
    )


# =========================
# CSV LOGGER -- CLEAN ONLY (Gap 1)
# Written after EDM pass, not in hotfolder.
# =========================
def append_to_csv(filename):
    awb = _awb_from_processed_filename(filename)
    CSV_PATH.parent.mkdir(parents=True, exist_ok=True)
    new_file = not CSV_PATH.exists()
    try:
        with open(CSV_PATH, "a", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            if new_file:
                w.writerow(["AWB", "SourceFile", "Timestamp"])
            w.writerow([awb, filename, datetime.now().isoformat(timespec="seconds")])
    except Exception as e:
        log.warning(f"[CSV] Could not write to awb_list.csv: {e}")


# =========================
# REJECTED SHEET LOGGER (Gap 2)
# =========================
def append_to_rejected_sheet(filename, reason, match_stats):
    awb = _awb_from_processed_filename(filename)
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    headers = ["AWB", "SourceFile", "Timestamp", "Reason", "MatchStats"]
    row = [awb, filename, ts, reason, match_stats]

    for attempt in range(5):
        try:
            if AWB_LOGS_PATH.exists():
                wb = load_workbook(AWB_LOGS_PATH)
            else:
                AWB_LOGS_PATH.parent.mkdir(parents=True, exist_ok=True)
                wb = Workbook()
                ws_main = wb.active
                ws_main.title = "AWB Logs"

            if "Rejected" not in wb.sheetnames:
                ws = wb.create_sheet("Rejected")
                ws.append(headers)
            else:
                ws = wb["Rejected"]

            ws.append(row)
            wb.save(AWB_LOGS_PATH)
            return
        except PermissionError:
            time.sleep(0.4 * (attempt + 1))
        except Exception as e:
            log.warning(f"[AWB_LOGS] Could not write to Rejected sheet: {e}")
            return

    log.warning(f"[AWB_LOGS] File still locked after retries -- skipping rejected log for {awb}.")


# =========================
# DUPLICATE-SAFE FILE MOVE
# =========================
def _file_md5(path):
    h = hashlib.md5()
    try:
        with open(path, "rb") as f:
            for chunk in iter(lambda: f.read(65536), b""):
                h.update(chunk)
    except Exception:
        return None
    return h.hexdigest()


def safe_move(src_path, dest_folder, filename):
    dest_path = Path(dest_folder) / filename
    if dest_path.exists():
        src_md5 = _file_md5(src_path)
        dst_md5 = _file_md5(dest_path)
        if src_md5 and dst_md5 and src_md5 == dst_md5:
            log.warning(f"Identical content already at destination -- removing source: {filename}")
            try:
                os.remove(src_path)
            except Exception:
                pass
            return str(dest_path)
        base, ext = os.path.splitext(filename)
        counter = 2
        while dest_path.exists():
            dest_path = Path(dest_folder) / f"{base}_{counter}{ext}"
            counter += 1
        log.warning(f"Destination exists (different content) -- saving as: {dest_path.name}")

    shutil.move(src_path, dest_path)
    return str(dest_path)


# =========================
# EDM API
# =========================
def _normalize_token(raw):
    if not raw:
        return None
    token = str(raw).strip().strip('"').strip("'")
    if token.lower().startswith("bearer "):
        token = token[7:].strip()
    return token or None


def _read_token_file():
    if not config.TOKEN_FILE.exists():
        return None
    raw = config.TOKEN_FILE.read_text(encoding="utf-8-sig")
    return _normalize_token(raw)


def _get_token_and_source():
    """
    Resolve EDM auth token and identify source.
    Priority: data/token.txt (if present and non-empty), then EDM_TOKEN in .env.
    """
    file_token = _read_token_file()
    if file_token:
        return file_token, "data/token.txt"

    env_token = _normalize_token(config.EDM_TOKEN)
    if env_token and env_token != "paste_your_token_here":
        return env_token, ".env:EDM_TOKEN"

    log.warning("EDM token not found. Set EDM_TOKEN in .env or create data/token.txt. EDM check will be skipped.")
    return None, None


def _get_token():
    token, _ = _get_token_and_source()
    return token


def get_headers():
    token = _get_token()
    return {
        "Authorization": "Bearer " + token,
        "Content-Type": "application/json",
        "Accept": "application/json, text/plain, */*",
        "Origin": "https://shipment-portal-g.prod.cloud.fedex.com",
        "Referer": "https://shipment-portal-g.prod.cloud.fedex.com/",
    }


def get_document_ids(awb):
    payload = {
        "documentClass": "SHIPMENT",
        "group": [{"operatingCompany": OPERATING_COMPANY, "trackingNumber": [awb]}],
        "responseTypes": ["metadata"],
    }
    params = {"pageSize": 25, "continuationToken": "", "archiveSelection": "false"}
    try:
        r = requests.post(METADATA_URL, headers=get_headers(), params=params, json=payload, timeout=30)
        if r.status_code == 401:
            log.error("TOKEN EXPIRED")
            return None
        if r.status_code == 404:
            return []
        if r.status_code != 200:
            log.warning(f"Unexpected status {r.status_code} for AWB {awb}")
            return []
        doc_ids = []
        for group in r.json().get("groups", []):
            for doc in group.get("documents", []):
                doc_id = doc.get("documentId") or doc.get("id")
                if doc_id:
                    doc_ids.append(doc_id)
        return doc_ids
    except requests.exceptions.Timeout:
        log.warning(f"Timeout querying AWB {awb} - treating as new")
        return []
    except Exception as e:
        log.warning(f"Error querying AWB {awb}: {e}")
        return []


def download_edm_zip(doc_ids):
    if not doc_ids:
        return None
    params = {"documentClass": "SHIPMENT", "archiveSelection": "false"}
    body = {"requestId": str(uuid.uuid4()), "smallerSizeDocumentId": ",".join(doc_ids)}
    try:
        headers = get_headers()
        headers["Accept"] = "application/zip, */*"
        r = requests.post(DOWNLOAD_URL, headers=headers, params=params, json=body, timeout=60)
        if r.status_code != 200:
            log.warning(f"EDM download failed - status {r.status_code}")
            return None
        ct = r.headers.get("Content-Type", "").lower()
        if "zip" in ct:
            if zip_has_supported_docs(r.content):
                return r.content
            log.warning("ZIP was empty - retrying individually")
            return download_edm_individually(doc_ids)
        if "pdf" in ct:
            return wrap_pdf_in_zip(r.content)
        log.warning(f"Unexpected content type: {ct}")
        return None
    except Exception as e:
        log.warning(f"Error downloading EDM ZIP: {e}")
        return None


def download_edm_individually(doc_ids):
    zip_buffer = io.BytesIO()
    found = 0
    with zipfile.ZipFile(zip_buffer, "w") as z:
        for doc_id in doc_ids:
            params = {"documentClass": "SHIPMENT", "archiveSelection": "false"}
            body = {"requestId": str(uuid.uuid4()), "smallerSizeDocumentId": doc_id}
            try:
                headers = get_headers()
                headers["Accept"] = "application/zip, */*"
                r = requests.post(DOWNLOAD_URL, headers=headers, params=params, json=body, timeout=60)
                if r.status_code == 200:
                    ct = r.headers.get("Content-Type", "").lower()
                    if "pdf" in ct:
                        z.writestr(doc_id + ".pdf", r.content)
                        found += 1
                    elif "zip" in ct:
                        for j, pdf in enumerate(extract_pdfs_from_zip(r.content)):
                            z.writestr(f"{doc_id}_{j}.pdf", pdf)
                            found += 1
            except Exception as e:
                log.warning(f"Error downloading doc {doc_id}: {e}")
    if found == 0:
        return None
    zip_buffer.seek(0)
    return zip_buffer.read()


def wrap_pdf_in_zip(pdf_bytes):
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as z:
        z.writestr("document.pdf", pdf_bytes)
    buf.seek(0)
    return buf.read()


def zip_has_supported_docs(zip_bytes):
    """Fast ZIP precheck to avoid full TIFF->PDF conversion during download step."""
    try:
        with zipfile.ZipFile(io.BytesIO(zip_bytes)) as z:
            for name in z.namelist():
                lower = name.lower()
                if lower.endswith(".pdf") or lower.endswith((".tiff", ".tif")):
                    return True
    except Exception as e:
        log.warning(f"Error inspecting ZIP: {e}")
    return False


def extract_pdfs_from_zip(zip_bytes):
    pdfs = []
    try:
        z = zipfile.ZipFile(io.BytesIO(zip_bytes))
        for name in z.namelist():
            lower = name.lower()
            if lower.endswith(".pdf"):
                pdfs.append(z.read(name))
                continue
            if lower.endswith((".tiff", ".tif")):
                try:
                    from PIL import Image as PILImage
                    tiff_bytes = z.read(name)
                    tiff_img = PILImage.open(io.BytesIO(tiff_bytes))
                    frames = []
                    try:
                        while True:
                            frames.append(tiff_img.copy().convert("RGB"))
                            tiff_img.seek(tiff_img.tell() + 1)
                    except EOFError:
                        pass
                    if not frames:
                        continue
                    pdf_doc = fitz.open()
                    for frame in frames:
                        frame_buf = io.BytesIO()
                        frame.save(frame_buf, format="PNG")
                        frame_buf.seek(0)
                        img_doc = fitz.open("png", frame_buf.read())
                        pdfbytes = img_doc.convert_to_pdf()
                        img_doc.close()
                        page_doc = fitz.open("pdf", pdfbytes)
                        pdf_doc.insert_pdf(page_doc)
                        page_doc.close()
                    pdf_buf = io.BytesIO()
                    pdf_doc.save(pdf_buf)
                    pdf_doc.close()
                    pdfs.append(pdf_buf.getvalue())
                    log.info(f"Converted TIFF->PDF: {name} ({len(frames)} frame(s))")
                except Exception as e:
                    log.warning(f"Failed to convert TIFF {name} to PDF: {e}")
    except Exception as e:
        log.warning(f"Error extracting ZIP: {e}")
    return pdfs


# =========================
# PAGE ANALYSIS HELPERS
# =========================
def hash_page(page):
    pix = page.get_pixmap(dpi=100)
    return hashlib.md5(pix.tobytes()).hexdigest()


def perceptual_hash_page(page):
    try:
        import imagehash
        from PIL import Image, ImageOps
        pix = page.get_pixmap(dpi=150)
        img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
        img = ImageOps.grayscale(img)
        img = ImageOps.autocontrast(img)
        return imagehash.phash(img)
    except Exception as e:
        log.warning(f"Error computing perceptual hash: {e}")
        return None


def extract_embedded_text(page, top_percent=100, page_index=0):
    text = ""
    try:
        rect = page.rect
        clip = fitz.Rect(rect.x0, rect.y0, rect.x1, rect.y0 + rect.height * top_percent / 100)
        text = page.get_text("text", clip=clip).strip().lower()
    except Exception as e:
        log.warning(f"Error extracting embedded text: {e}")

    if len(text) < MIN_EMBEDDED_TEXT_LENGTH:
        if page_index < PAGE_OCR_LIMIT:
            log.info(f"    Page {page_index+1}: embedded text too short ({len(text)} chars) -- using OCR fallback")
            ocr_text = extract_ocr_text(page, top_percent)
            if ocr_text:
                return ocr_text
        else:
            log.info(f"    Page {page_index+1}: embedded text too short but beyond PAGE_OCR_LIMIT -- skipping OCR")
    return text


def extract_embedded_text_only(page, top_percent=100):
    """Embedded text only (no OCR fallback)."""
    try:
        rect = page.rect
        clip = fitz.Rect(rect.x0, rect.y0, rect.x1, rect.y0 + rect.height * top_percent / 100)
        return page.get_text("text", clip=clip).strip().lower()
    except Exception as e:
        log.warning(f"Error extracting embedded text: {e}")
        return ""


def preprocess_image_for_ocr(img):
    import cv2
    import numpy as np
    from PIL import Image as PILImage
    arr = np.array(img)
    gray = cv2.cvtColor(arr, cv2.COLOR_RGB2GRAY)
    denoised = cv2.fastNlMeansDenoising(gray, h=10)
    clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
    enhanced = clahe.apply(denoised)
    thresh = cv2.threshold(enhanced, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)[1]
    coords = np.column_stack(np.where(thresh > 0))
    if len(coords) > 0:
        angle = cv2.minAreaRect(coords)[-1]
        if angle < -45:
            angle = 90 + angle
        if abs(angle) > 0.5:
            h2, w2 = thresh.shape
            center = (w2 // 2, h2 // 2)
            M = cv2.getRotationMatrix2D(center, angle, 1.0)
            thresh = cv2.warpAffine(thresh, M, (w2, h2), flags=cv2.INTER_CUBIC, borderMode=cv2.BORDER_REPLICATE)
    return PILImage.fromarray(thresh)


def extract_ocr_text(page, top_percent=50):
    try:
        import pytesseract
        from PIL import Image
        pytesseract.pytesseract.tesseract_cmd = TESSERACT_PATH
        pix = page.get_pixmap(dpi=200)
        img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
        crop_height = int(img.height * top_percent / 100)
        cropped = img.crop((0, 0, img.width, crop_height))
        processed = preprocess_image_for_ocr(cropped)
        return pytesseract.image_to_string(processed, config="--psm 6").strip().lower()
    except Exception as e:
        log.warning(f"Error during OCR: {e}")
        return ""


def text_similarity(text1, text2):
    try:
        from rapidfuzz import fuzz
        scores = [fuzz.ratio(text1, text2), fuzz.partial_ratio(text1, text2),
                  fuzz.token_sort_ratio(text1, text2), fuzz.token_set_ratio(text1, text2)]
        best = max(scores)
        log.info("    Similarity: ratio=%.1f partial=%.1f token_sort=%.1f token_set=%.1f best=%.1f"
                 % tuple(scores + [best]))
        return best
    except Exception as e:
        log.warning(f"Error comparing text: {e}")
        return 0


def page_is_cargo_control_document(page):
    try:
        text = page.get_text("text") or ""
        if not text.strip():
            try:
                import pytesseract
                from PIL import Image
                pytesseract.pytesseract.tesseract_cmd = TESSERACT_PATH
                pix = page.get_pixmap(dpi=200)
                img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                text = pytesseract.image_to_string(img, config="--psm 6")
            except Exception as e:
                log.warning(f"OCR fallback failed in CCD check: {e}")

        text_upper = text.upper()
        has_ccd = ("CARGO CONTROL DOCUMENT" in text_upper or
                   "FEUILLE DE RECAPITULATION" in text_upper)
        has_400 = bool(re.search(r"400[\s\-]?\d{10,12}", text))

        if has_ccd and has_400:
            return True
        if has_ccd or has_400:
            log.info(f"    CCD partial match -- has_ccd={has_ccd} has_400={has_400} (both required to exempt)")
        return False
    except Exception as e:
        log.warning(f"Error checking CCD status on page: {e}")
        return False


# =========================
# AWB LOGS WRITER
# =========================
def append_edm_result_to_awb_logs(awb, filename, result, reason, match_stats):
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    row = [awb, filename, ts, "EDM-Check", result, reason, match_stats]
    headers = ["AWB", "SourceFile", "Timestamp", "MatchMethod", "Status", "Reason", "MatchStats"]

    for attempt in range(5):
        try:
            if AWB_LOGS_PATH.exists():
                wb = load_workbook(AWB_LOGS_PATH)
                ws = wb.active
                existing_headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
                for col, h in enumerate(headers, start=1):
                    if h not in existing_headers:
                        ws.cell(1, col).value = h
            else:
                AWB_LOGS_PATH.parent.mkdir(parents=True, exist_ok=True)
                wb = Workbook()
                ws = wb.active
                ws.title = "AWB Logs"
                ws.append(headers)
            ws.append(row)
            wb.save(AWB_LOGS_PATH)
            return
        except PermissionError:
            time.sleep(0.4 * (attempt + 1))
        except Exception as e:
            log.warning(f"[AWB_LOGS] Could not write EDM result: {e}")
            return

    log.warning(f"[AWB_LOGS] File still locked after retries -- skipping log for {awb}.")


def build_edm_fingerprints(edm_pdf_list):
    """Precompute non-OCR fingerprints for EDM docs so same-AWB files can reuse them."""
    fingerprints = []
    for edm_bytes in edm_pdf_list:
        fp = {
            "valid": False,
            "page_count": 0,
            "hash_map": {},
            "phashes": [],
            "texts": [],
            "numeric_top_tokens": set(),
        }
        try:
            edm_doc = fitz.open(stream=edm_bytes, filetype="pdf")
            fp["valid"] = True
            fp["page_count"] = len(edm_doc)

            hash_map = {}
            phashes = []
            texts = []
            num_counter = Counter()

            for ei in range(len(edm_doc)):
                ep = edm_doc[ei]
                eh = hash_page(ep)
                if eh not in hash_map:
                    hash_map[eh] = ei
                phashes.append(perceptual_hash_page(ep))
                txt = extract_embedded_text_only(ep, top_percent=100)
                texts.append(txt)
                for tok in re.findall(r"\b\d{6,16}\b", txt or ""):
                    num_counter[tok] += 1

            fp["hash_map"] = hash_map
            fp["phashes"] = phashes
            fp["texts"] = texts
            fp["numeric_top_tokens"] = set(t for t, _ in num_counter.most_common(12))
            edm_doc.close()
        except Exception as e:
            log.warning(f"    Could not fingerprint EDM doc: {e}")
        fingerprints.append(fp)
    return fingerprints


def _rejection_confidence(dup_meta):
    counts = dup_meta.get("method_counts", {}) if dup_meta else {}
    h = counts.get("HASH", 0)
    p = counts.get("PHASH", 0)
    t = counts.get("TEXT", 0)
    o = counts.get("OCR", 0)
    if h >= 1:
        return "HIGH"
    if p >= 2 and (t + o) >= 1:
        return "HIGH"
    if p >= 3:
        return "MEDIUM"
    if (t + o) >= 3 and p >= 1:
        return "MEDIUM"
    return "LOW"


# =========================
# DUPLICATE PAGE DETECTION
# =========================
def find_duplicate_pages(incoming_path, edm_pdf_list, edm_fingerprints=None):
    duplicate_pages = set()
    duplicate_page_details = {}
    focused_edm_idx = None
    kept_edm_indices = None
    prefilter_kept_human = []
    prefilter_skipped_human = []
    ocr_gate_hit = False
    force_ocr_fallback = False
    prefilter_token_signal_found = False
    ocr_max_pages = OCR_COMPARE_LIMIT

    try:
        incoming_doc = fitz.open(incoming_path)
        if len(incoming_doc) == 0:
            return duplicate_pages, {
                "methods": [],
                "score_summary": "",
                "page_details": {},
                "method_counts": {},
                "decision_trace": "",
            }

        total_incoming = len(incoming_doc)
        log.info(f"    Checking against {len(edm_pdf_list)} EDM doc(s)")

        if edm_fingerprints is None:
            edm_fingerprints = build_edm_fingerprints(edm_pdf_list)
            log.info("    Built EDM fingerprints (cache miss)")
        else:
            log.info("    Reusing EDM fingerprints from AWB cache")

        edm_docs = [None] * len(edm_pdf_list)  # lazily opened only when OCR needs page access
        edm_doc_valid = [bool(fp.get("valid")) for fp in edm_fingerprints]
        edm_page_counts = [int(fp.get("page_count", 0)) for fp in edm_fingerprints]
        edm_hash_maps = [fp.get("hash_map", {}) for fp in edm_fingerprints]
        edm_phash_lists = [fp.get("phashes", []) for fp in edm_fingerprints]
        edm_text_lists = [fp.get("texts", []) for fp in edm_fingerprints]
        edm_numeric_top_tokens = [fp.get("numeric_top_tokens", set()) for fp in edm_fingerprints]

        inc_pages = [incoming_doc[p] for p in range(total_incoming)]
        edm_match_counts = [0] * len(edm_docs)
        inc_hashes = {}
        inc_phashes = {}
        inc_texts = {}
        inc_ocr_texts = {}
        edm_ocr_texts = [{} for _ in edm_docs]
        inc_is_ccd = {}

        # Conservative prefilter tuning (skip only clearly cold docs)
        PREFILTER_LOW_PHASH_HITS = 1
        PREFILTER_NEAR_ZERO_PHASH = 0
        PREFILTER_LOW_TOKEN_OVERLAP = 0.2
        PREFILTER_VERY_LOW_TOKEN_OVERLAP = 0.05
        PREFILTER_MIN_TOPN = 3

        def should_check_edm(i):
            in_prefilter = kept_edm_indices is None or i in kept_edm_indices
            if not in_prefilter:
                return False
            return focused_edm_idx is None or i == focused_edm_idx

        def ensure_edm_doc_open(i):
            if not edm_doc_valid[i]:
                return None
            if edm_docs[i] is None:
                try:
                    edm_docs[i] = fitz.open(stream=edm_pdf_list[i], filetype="pdf")
                except Exception as e:
                    log.warning(f"    Could not open EDM doc for OCR (index {i+1}): {e}")
                    edm_doc_valid[i] = False
                    edm_docs[i] = None
            return edm_docs[i]

        def page_is_ccd_cached(ii, page):
            if ii not in inc_is_ccd:
                inc_is_ccd[ii] = page_is_cargo_control_document(page)
            return inc_is_ccd[ii]

        def update_focus(i):
            nonlocal focused_edm_idx
            if focused_edm_idx is None and edm_match_counts[i] >= EARLY_FOCUS_MATCH_THRESHOLD:
                focused_edm_idx = i
                log.info(f"    EDM {i+1}: {edm_match_counts[i]} pages matched -- focusing remaining checks on this doc")

        def get_inc_embedded_text(ii, ip):
            if ii not in inc_texts:
                inc_texts[ii] = extract_embedded_text_only(ip, top_percent=100)
            return inc_texts[ii]

        def mark_duplicate(ii, i, method, score_repr=""):
            if ii in duplicate_pages:
                return
            duplicate_pages.add(ii)
            duplicate_page_details[ii + 1] = {
                "method": method,
                "score": score_repr or "exact",
            }
            edm_match_counts[i] += 1
            update_focus(i)

        def prefilter_edm_candidates():
            nonlocal kept_edm_indices, prefilter_token_signal_found
            if len(edm_docs) <= 3:
                kept_edm_indices = set(i for i, d in enumerate(edm_docs) if d is not None)
                prefilter_token_signal_found = False
                return

            inc_hash_set = set()
            inc_phash_values = []
            inc_num_counter = Counter()
            for ii, ip in enumerate(inc_pages):
                if page_is_ccd_cached(ii, ip):
                    continue
                ih = hash_page(ip)
                inc_hashes[ii] = ih
                inc_hash_set.add(ih)

                iph = perceptual_hash_page(ip)
                if iph is not None:
                    inc_phash_values.append(iph)
                    inc_phashes[ii] = iph

                txt = get_inc_embedded_text(ii, ip)
                for tok in re.findall(r"\b\d{6,16}\b", txt or ""):
                    inc_num_counter[tok] += 1

            inc_top_tokens = set(t for t, _ in inc_num_counter.most_common(12))
            page_count_ref = max(total_incoming, 1)

            scored = []
            for i, edm_doc in enumerate(edm_docs):
                if not edm_doc_valid[i]:
                    continue

                hash_overlap = sum(1 for h in inc_hash_set if h in edm_hash_maps[i])

                phash_hits = 0
                edm_phashes = [p for p in edm_phash_lists[i] if p is not None]
                if inc_phash_values and edm_phashes:
                    for iph in inc_phash_values:
                        matched = False
                        for eph in edm_phashes:
                            if (iph - eph) <= PHASH_THRESHOLD:
                                phash_hits += 1
                                matched = True
                                break
                        if matched:
                            continue

                edm_tokens = edm_numeric_top_tokens[i]
                token_intersection = len(inc_top_tokens & edm_tokens)
                token_overlap = (token_intersection / max(1, len(inc_top_tokens)))
                if token_overlap >= PREFILTER_LOW_TOKEN_OVERLAP:
                    prefilter_token_signal_found = True

                page_diff = abs(edm_page_counts[i] - total_incoming)
                page_proximity = 1.0 - min(1.0, page_diff / page_count_ref)

                combined = (
                    (2.5 * hash_overlap)
                    + (1.5 * phash_hits)
                    + (2.0 * token_overlap)
                    + (0.5 * page_proximity)
                )

                scored.append({
                    "idx": i,
                    "hash_overlap": hash_overlap,
                    "phash_hits": phash_hits,
                    "token_overlap": token_overlap,
                    "page_proximity": page_proximity,
                    "combined": combined,
                })

            if not scored:
                kept_edm_indices = set(i for i, d in enumerate(edm_docs) if d is not None)
                return

            scored.sort(key=lambda x: x["combined"], reverse=True)
            top_n = min(max(PREFILTER_MIN_TOPN, 1), len(scored))
            top_n_indices = set(x["idx"] for x in scored[:top_n])

            keep = set()
            for s in scored:
                idx = s["idx"]
                keep_by_signal = (
                    s["hash_overlap"] > 0
                    or s["phash_hits"] >= PREFILTER_LOW_PHASH_HITS
                    or s["token_overlap"] >= PREFILTER_LOW_TOKEN_OVERLAP
                    or idx in top_n_indices
                )
                is_cold = (
                    s["hash_overlap"] == 0
                    and s["phash_hits"] <= PREFILTER_NEAR_ZERO_PHASH
                    and s["token_overlap"] <= PREFILTER_VERY_LOW_TOKEN_OVERLAP
                    and idx not in top_n_indices
                )
                if keep_by_signal and not is_cold:
                    keep.add(idx)

                log.info(
                    "    PREFILTER EDM %s: hash_overlap=%s phash_hits=%s token_overlap=%.3f "
                    "page_proximity=%.3f combined=%.3f keep=%s cold=%s",
                    idx + 1,
                    s["hash_overlap"],
                    s["phash_hits"],
                    s["token_overlap"],
                    s["page_proximity"],
                    s["combined"],
                    keep_by_signal,
                    is_cold,
                )

            if not keep:
                keep = set(top_n_indices)

            kept_edm_indices = keep
            nonlocal prefilter_kept_human, prefilter_skipped_human
            prefilter_kept_human = [i + 1 for i in sorted(kept_edm_indices)]
            prefilter_skipped_human = [i + 1 for i, ok in enumerate(edm_doc_valid) if ok and i not in kept_edm_indices]
            log.info("    PREFILTER kept EDM docs: %s | skipped cold docs: %s", prefilter_kept_human, prefilter_skipped_human)

        prefilter_edm_candidates()

        def get_inc_ocr_text(ii):
            if ii in inc_ocr_texts:
                return inc_ocr_texts[ii]
            if ocr_max_pages is not None and ii >= ocr_max_pages:
                inc_ocr_texts[ii] = ""
                return ""
            inc_ocr_texts[ii] = extract_ocr_text(inc_pages[ii], top_percent=100)
            return inc_ocr_texts[ii]

        def get_edm_ocr_text(i, ei):
            cache = edm_ocr_texts[i]
            if ei in cache:
                return cache[ei]
            edm_doc = ensure_edm_doc_open(i)
            if edm_doc is None or ei >= len(edm_doc) or (ocr_max_pages is not None and ei >= ocr_max_pages):
                cache[ei] = ""
                return ""
            cache[ei] = extract_ocr_text(edm_doc[ei], top_percent=100)
            return cache[ei]

        def edm_ocr_limit(i):
            if ocr_max_pages is None:
                return len(edm_text_lists[i])
            return min(len(edm_text_lists[i]), ocr_max_pages)

        def ocr_quick_indication():
            anchor_pages = []
            max_anchor_pages = min(3, total_incoming)
            for ii in range(max_anchor_pages):
                if ii in duplicate_pages or page_is_ccd_cached(ii, inc_pages[ii]):
                    continue
                anchor_pages.append(ii)

            if not anchor_pages:
                log.info("    OCR gate: no eligible incoming anchor pages (1-3) -- skipping full OCR fallback")
                return False

            log.info("    OCR gate: quick indication pass (incoming p1-3 vs EDM p1-5)")

            # Priority: incoming page 1 against page 1 of every EDM doc.
            if 0 in anchor_pages:
                inc_p1_ocr = get_inc_ocr_text(0)
                if inc_p1_ocr:
                    for i, edm_doc in enumerate(edm_docs):
                        if not edm_doc_valid[i] or not edm_text_lists[i]:
                            continue
                        edm_p1_text = edm_text_lists[i][0] or get_edm_ocr_text(i, 0)
                        if not edm_p1_text:
                            continue
                        score = text_similarity(inc_p1_ocr, edm_p1_text)
                        log.info(f"    OCR gate: text similarity={score} (incoming p1 vs EDM {i+1} p1)")
                        if score >= TEXT_SIMILARITY_THRESHOLD:
                            log.info(f"    OCR gate: indication found (incoming p1 vs EDM {i+1} p1)")
                            return True

            # Then: incoming pages 1-3 against EDM pages 1-5.
            for ii in anchor_pages:
                inc_ocr = get_inc_ocr_text(ii)
                if not inc_ocr:
                    continue
                for i, edm_doc in enumerate(edm_docs):
                    if not edm_doc_valid[i]:
                        continue
                    lim = min(5, len(edm_text_lists[i]))
                    for ei in range(lim):
                        if ii == 0 and ei == 0:
                            continue
                        edm_text = edm_text_lists[i][ei] or get_edm_ocr_text(i, ei)
                        if not edm_text:
                            continue
                        score = text_similarity(inc_ocr, edm_text)
                        log.info(f"    OCR gate: text similarity={score} (incoming p{ii+1} vs EDM {i+1} p{ei+1})")
                        if score >= TEXT_SIMILARITY_THRESHOLD:
                            log.info(f"    OCR gate: indication found (incoming p{ii+1} vs EDM {i+1} p{ei+1})")
                            return True

            log.info("    OCR gate: no duplicate indication in quick window (p1-3 vs EDM p1-5)")
            return False

        # 1) Exact hash -- no page limit
        for ii, ip in enumerate(inc_pages):
            if ii in duplicate_pages:
                continue
            if page_is_ccd_cached(ii, ip):
                log.info(f"    Page {ii+1}: CCD detected -- exempt from all checks")
                continue
            if ii not in inc_hashes:
                inc_hashes[ii] = hash_page(ip)
            ih = inc_hashes[ii]
            for i, edm_doc in enumerate(edm_docs):
                if not edm_doc_valid[i] or not should_check_edm(i):
                    continue
                ei = edm_hash_maps[i].get(ih)
                if ei is not None:
                    log.info(f"    EDM {i+1}: DUPLICATE (exact hash) incoming p{ii+1} vs EDM p{ei+1}")
                    mark_duplicate(ii, i, "HASH", "exact")
                    break

        # 2) Perceptual hash -- within PAGE_OCR_LIMIT
        for ii, ip in enumerate(inc_pages):
            if ii in duplicate_pages or page_is_ccd_cached(ii, ip):
                continue
            if ii not in inc_phashes:
                inc_phashes[ii] = perceptual_hash_page(ip)
            iph = inc_phashes[ii]
            if iph is None:
                continue
            for i, edm_doc in enumerate(edm_docs):
                if not edm_doc_valid[i] or not should_check_edm(i):
                    continue
                for ei, eph in enumerate(edm_phash_lists[i]):
                    if eph is None:
                        continue
                    diff = iph - eph
                    log.info(f"    EDM {i+1}: phash diff={diff} (incoming p{ii+1} vs EDM p{ei+1})")
                    if diff <= PHASH_THRESHOLD:
                        log.info(f"    EDM {i+1}: DUPLICATE (visual match) incoming p{ii+1}")
                        mark_duplicate(ii, i, "PHASH", f"diff={diff}")
                        break

        # 3) Text similarity -- embedded text first (OCR deferred behind quick gate)
        ocr_needed = False
        for ii, ip in enumerate(inc_pages):
            if ii in duplicate_pages or page_is_ccd_cached(ii, ip):
                continue
            inc_text = get_inc_embedded_text(ii, ip)
            for i, edm_doc in enumerate(edm_docs):
                if not edm_doc_valid[i] or not should_check_edm(i):
                    continue
                for ei, edm_text in enumerate(edm_text_lists[i]):
                    if not inc_text or not edm_text:
                        ocr_needed = True
                        continue
                    score = text_similarity(inc_text, edm_text)
                    log.info(f"    EDM {i+1}: text similarity={score} (incoming p{ii+1} vs EDM p{ei+1})")
                    if score >= TEXT_SIMILARITY_THRESHOLD:
                        log.info(f"    EDM {i+1}: DUPLICATE (text match) incoming p{ii+1}")
                        mark_duplicate(ii, i, "TEXT", f"{score:.1f}")
                        break

        # 4) OCR text fallback -- gated by OCR quick window, with conservative safety fallback.
        non_ocr_methods = {v.get("method") for v in duplicate_page_details.values()}
        no_non_ocr_match_signals = (
            "HASH" not in non_ocr_methods
            and "PHASH" not in non_ocr_methods
            and "TEXT" not in non_ocr_methods
            and not prefilter_token_signal_found
        )

        should_run_quick_ocr = ocr_needed or no_non_ocr_match_signals
        force_full_ocr_all_pages = no_non_ocr_match_signals

        if should_run_quick_ocr:
            ocr_gate_hit = ocr_quick_indication()
            # Safety net: if non-OCR already found signals concentrated on one EDM doc, do not skip OCR.
            force_ocr_fallback = (focused_edm_idx is not None and len(duplicate_pages) > 0)

            if ocr_gate_hit or force_ocr_fallback:
                if ocr_gate_hit:
                    log.info("    OCR gate: indication found -- running full OCR text fallback")
                else:
                    log.info("    OCR gate: no window hit, but focused duplicate signals exist -- running safety OCR fallback")

                if force_full_ocr_all_pages:
                    ocr_max_pages = None
                    log.info("    OCR mode: no hash/phash/text/token matches -- escalating to full OCR on all pages")
                else:
                    ocr_max_pages = OCR_COMPARE_LIMIT

                for ii, ip in enumerate(inc_pages):
                    if ii in duplicate_pages or page_is_ccd_cached(ii, ip):
                        continue
                    if ocr_max_pages is not None and ii >= ocr_max_pages:
                        continue

                    inc_text = get_inc_embedded_text(ii, ip)
                    if not inc_text:
                        inc_text = get_inc_ocr_text(ii)

                    for i, edm_doc in enumerate(edm_docs):
                        if not edm_doc_valid[i] or not should_check_edm(i):
                            continue
                        ocr_lim = edm_ocr_limit(i)
                        for ei in range(ocr_lim):
                            edm_text = edm_text_lists[i][ei]
                            if not edm_text:
                                edm_text = get_edm_ocr_text(i, ei)
                            if not inc_text or not edm_text:
                                log.info(f"    EDM {i+1}: insufficient text (incoming p{ii+1} vs EDM p{ei+1}) -- treating as new")
                                continue
                            score = text_similarity(inc_text, edm_text)
                            log.info(f"    EDM {i+1}: OCR text similarity={score} (incoming p{ii+1} vs EDM p{ei+1})")
                            if score >= TEXT_SIMILARITY_THRESHOLD:
                                log.info(f"    EDM {i+1}: DUPLICATE (OCR text match) incoming p{ii+1}")
                                mark_duplicate(ii, i, "OCR", f"{score:.1f}")
                                break
            else:
                log.info("    OCR gate: no indication -- skipping full OCR text fallback for fast clean pass")

        for edm_doc in edm_docs:
            if edm_doc is not None:
                try:
                    edm_doc.close()
                except Exception:
                    pass
        incoming_doc.close()

    except Exception as e:
        log.warning(f"Error during duplicate check: {e}")

    methods = []
    seen = set()
    method_counts = {}
    for page_no in sorted(duplicate_page_details):
        method = duplicate_page_details[page_no]["method"]
        method_counts[method] = method_counts.get(method, 0) + 1
        if method not in seen:
            methods.append(method)
            seen.add(method)

    primary_score = ""
    if duplicate_page_details:
        first_page = sorted(duplicate_page_details)[0]
        d = duplicate_page_details[first_page]
        primary_score = f"{d['method']}:{d['score']}"

    decision_trace = (
        f"prefilter_kept={prefilter_kept_human};"
        f"prefilter_skipped={prefilter_skipped_human};"
        f"ocr_gate_hit={ocr_gate_hit};"
        f"ocr_force_fallback={force_ocr_fallback};"
        f"methods={methods};"
        f"method_counts={method_counts}"
    )

    return duplicate_pages, {
        "methods": methods,
        "score_summary": primary_score,
        "page_details": duplicate_page_details,
        "method_counts": method_counts,
        "decision_trace": decision_trace,
    }


# =========================
# MAIN FILE PROCESSOR
# =========================
def process_file(filepath):
    total_start = time.perf_counter()
    t = {
        "cache": "MISS",
        "metadata_ms": 0.0,
        "download_ms": 0.0,
        "extract_ms": 0.0,
        "compare_ms": 0.0,
        "route_ms": 0.0,
        "total_active_ms": 0.0,
    }
    filename = os.path.basename(filepath)
    awb = _awb_from_processed_filename(filename)
    stage_row = _get_stage_cache_row(filename) or {}

    stage_awb = (stage_row.get("AWB_Detected") or awb or "").strip()
    stage_detection_type = (stage_row.get("AWB_Detection_Type") or "").strip()
    try:
        stage_awb_secs = float(stage_row.get("AWB_Extraction_Seconds") or 0)
    except Exception:
        stage_awb_secs = 0.0

    def finalize_audit(status, route, reason, match_stats="N/A"):
        t["total_active_ms"] = _ms(total_start)
        _log_timing(awb, filename, t)
        audit_event(
            "EDM_CHECK",
            file=filename,
            awb=awb,
            status=status,
            route=route,
            reason=reason,
            match_stats=match_stats,
            timings_ms={
                "metadata": t["metadata_ms"],
                "download": t["download_ms"],
                "extract": t["extract_ms"],
                "compare": t["compare_ms"],
                "route": t["route_ms"],
                "total_active": t["total_active_ms"],
            },
            cache=t.get("cache", "MISS"),
        )

    def emit_pipeline_summary(edm_status, total_pages=0, duplicate_pages=None, clean_pages=None, dup_meta=None):
        duplicate_pages = duplicate_pages or []
        clean_pages = clean_pages or []
        dup_meta = dup_meta or {}
        methods = dup_meta.get("methods") or []
        score_summary = dup_meta.get("score_summary") or ""
        decision_trace = dup_meta.get("decision_trace") or ""

        edm_minutes = round((t.get("total_active_ms", 0.0) / 1000.0) / 60.0, 4)
        total_minutes = round((stage_awb_secs / 60.0) + edm_minutes, 4)

        row = {
            "Timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "InputFileName": filename,
            "AWB_Detected": stage_awb or awb or "",
            "AWB_Detection_Type": stage_detection_type or "UNKNOWN",
            "EDM_Check_Status": edm_status,
            "Duplicate_Detection_Type": "|".join(methods) if methods else "",
            "Detection_Type_Match_Score": score_summary,
            "AWB_Extraction_Seconds": round(stage_awb_secs, 3),
            "EDM_Check_Minutes": edm_minutes,
            "TOTAL_Minutes_AWB_plus_EDM": total_minutes,
            "Total_Pages": total_pages,
            "Duplicate_Pages": str(duplicate_pages),
            "Pages_To_Clean": str(clean_pages),
            "Decision_Trace": decision_trace,
        }
        _queue_summary_row(row)

    log.info("=" * 55)
    log.info(f"File:  {filename}")
    log.info(f"AWB:   {awb}")

    if not awb:
        log.warning(f"Invalid filename format for AWB extraction: {filename} -- passing through CLEAN-UNCHECKED")
        safe_move(filepath, CLEAN_FOLDER, filename)
        append_to_csv(filename)
        append_edm_result_to_awb_logs(
            awb or "UNKNOWN",
            filename,
            result="CLEAN-UNCHECKED",
            reason="Invalid filename format for AWB extraction",
            match_stats="N/A",
        )
        record_edm_end(
            filename,
            edm_result="CLEAN-UNCHECKED",
            final_folder="CLEAN",
            notes="Invalid filename format for AWB extraction",
        )
        finalize_audit("CLEAN-UNCHECKED", "CLEAN", "Invalid filename format for AWB extraction")
        emit_pipeline_summary("CLEAN-UNCHECKED")
        return

    # Keep cache only for the active AWB, clear before moving to next AWB.
    if AWB_SESSION_CACHE["awb"] and AWB_SESSION_CACHE["awb"] != awb:
        _clear_awb_cache("moving to next AWB")

    record_edm_start(filename)

    cache_ready = (
        AWB_SESSION_CACHE["awb"] == awb
        and AWB_SESSION_CACHE["doc_ids"] is not None
        and AWB_SESSION_CACHE["edm_pdf_list"] is not None
        and AWB_SESSION_CACHE["edm_fingerprints"] is not None
    )

    if cache_ready:
        t["cache"] = "HIT"
        doc_ids = AWB_SESSION_CACHE["doc_ids"]
        edm_pdf_list = AWB_SESSION_CACHE["edm_pdf_list"]
        edm_fingerprints = AWB_SESSION_CACHE["edm_fingerprints"]
        log.info(f"[CACHE] AWB cache hit for {awb} -- reusing EDM snapshot")
    else:
        t["cache"] = "MISS"
        log.info("Querying EDM...")
        meta_start = time.perf_counter()
        doc_ids = get_document_ids(awb)
        t["metadata_ms"] = _ms(meta_start)
        log.info(f"[TIMING] metadata call completed in {t['metadata_ms']} ms")

        if doc_ids is None:
            finalize_audit("STOPPED", "STOP", "TOKEN EXPIRED")
            emit_pipeline_summary("STOPPED")
            log.error("Stopping -- token expired. Refresh token in data/token.txt or .env and restart.")
            sys.exit(1)

        if not doc_ids:
            AWB_SESSION_CACHE["awb"] = awb
            AWB_SESSION_CACHE["doc_ids"] = []
            AWB_SESSION_CACHE["edm_pdf_list"] = []
            AWB_SESSION_CACHE["edm_fingerprints"] = []
            edm_pdf_list = []
            edm_fingerprints = []
        else:
            log.info(f"Found {len(doc_ids)} existing doc(s) in EDM")
            log.info("Downloading from EDM...")
            dl_start = time.perf_counter()
            zip_bytes = download_edm_zip(doc_ids)
            t["download_ms"] = _ms(dl_start)
            log.info(f"[TIMING] EDM download completed in {t['download_ms']} ms")

            if not zip_bytes:
                # Do not cache failures.
                route_start = time.perf_counter()
                log.warning("Could not download from EDM -- passing through unchecked")
                safe_move(filepath, CLEAN_FOLDER, filename)
                append_to_csv(filename)
                append_edm_result_to_awb_logs(awb, filename, result="CLEAN-UNCHECKED", reason="EDM download failed", match_stats="N/A")
                record_edm_end(filename, edm_result="CLEAN-UNCHECKED", final_folder="CLEAN", notes="EDM download failed")
                t["route_ms"] = _ms(route_start)
                finalize_audit("CLEAN-UNCHECKED", "CLEAN", "EDM download failed")
                emit_pipeline_summary("CLEAN-UNCHECKED")
                return

            extract_start = time.perf_counter()
            edm_pdf_list = extract_pdfs_from_zip(zip_bytes)
            t["extract_ms"] = _ms(extract_start)
            log.info(f"[TIMING] EDM ZIP extraction completed in {t['extract_ms']} ms")
            if not edm_pdf_list:
                # Do not cache failures.
                route_start = time.perf_counter()
                log.warning("No PDFs in EDM ZIP -- passing through unchecked")
                safe_move(filepath, CLEAN_FOLDER, filename)
                append_to_csv(filename)
                append_edm_result_to_awb_logs(awb, filename, result="CLEAN-UNCHECKED", reason="EDM ZIP empty or unreadable", match_stats="N/A")
                record_edm_end(filename, edm_result="CLEAN-UNCHECKED", final_folder="CLEAN", notes="EDM ZIP empty or unreadable")
                t["route_ms"] = _ms(route_start)
                finalize_audit("CLEAN-UNCHECKED", "CLEAN", "EDM ZIP empty or unreadable")
                emit_pipeline_summary("CLEAN-UNCHECKED")
                return

            AWB_SESSION_CACHE["awb"] = awb
            AWB_SESSION_CACHE["doc_ids"] = list(doc_ids)
            AWB_SESSION_CACHE["edm_pdf_list"] = list(edm_pdf_list)
            AWB_SESSION_CACHE["edm_fingerprints"] = build_edm_fingerprints(edm_pdf_list)
            edm_fingerprints = AWB_SESSION_CACHE["edm_fingerprints"]
            log.info(f"[CACHE] Cached EDM snapshot for {awb} ({len(doc_ids)} doc id(s), {len(edm_pdf_list)} PDF(s))")

    if not doc_ids:
        route_start = time.perf_counter()
        log.info("AWB not in EDM -- passing through as new")
        safe_move(filepath, CLEAN_FOLDER, filename)
        log.info("RESULT: NEW -> CLEAN")
        append_to_csv(filename)
        append_edm_result_to_awb_logs(awb, filename, result="CLEAN", reason="AWB not found in EDM", match_stats="N/A")
        record_edm_end(filename, edm_result="CLEAN", final_folder="CLEAN")
        t["route_ms"] = _ms(route_start)
        finalize_audit("CLEAN", "CLEAN", "AWB not found in EDM")
        emit_pipeline_summary("CLEAN")
        return

    log.info(f"Extracted {len(edm_pdf_list)} PDF(s) from EDM ZIP")
    log.info("Comparing pages...")
    compare_start = time.perf_counter()
    duplicate_pages, dup_meta = find_duplicate_pages(filepath, edm_pdf_list, edm_fingerprints=edm_fingerprints)
    t["compare_ms"] = _ms(compare_start)
    log.info(f"[TIMING] page comparison completed in {t['compare_ms']} ms")

    incoming_doc = fitz.open(filepath)
    total_pages = len(incoming_doc)
    incoming_doc.close()

    match_stats = (f"dup_pages={sorted([p+1 for p in duplicate_pages])} "
                   f"total_pages={total_pages} edm_docs={len(edm_pdf_list)}")
    dup_count = len(duplicate_pages)
    dup_ratio = (dup_count / total_pages) if total_pages else 0.0
    reject_conf = _rejection_confidence(dup_meta)
    dup_meta["decision_trace"] = (
        f"{dup_meta.get('decision_trace', '')};dup_count={dup_count};"
        f"dup_ratio={dup_ratio:.2f};reject_conf={reject_conf}"
    ).strip(";")
    match_stats = (
        f"{match_stats} dup_count={dup_count} dup_ratio={dup_ratio:.2f} "
        f"reject_confidence={reject_conf}"
    )

    # Case 1: No duplicates
    if not duplicate_pages:
        route_start = time.perf_counter()
        log.info("RESULT: NO duplicates -> CLEAN")
        safe_move(filepath, CLEAN_FOLDER, filename)
        append_to_csv(filename)
        append_edm_result_to_awb_logs(awb, filename, result="CLEAN", reason="No matching pages found in EDM", match_stats=match_stats)
        record_edm_end(filename, edm_result="CLEAN", final_folder="CLEAN")
        t["route_ms"] = _ms(route_start)
        finalize_audit("CLEAN", "CLEAN", "No matching pages found in EDM", match_stats=match_stats)
        emit_pipeline_summary("CLEAN", total_pages=total_pages, duplicate_pages=[], clean_pages=list(range(1, total_pages + 1)), dup_meta=dup_meta)
        return

    # Case 2: All pages duplicate
    if dup_count == total_pages:
        if reject_conf == "LOW":
            route_start = time.perf_counter()
            reason = "All pages matched but evidence confidence is LOW -- passing through CLEAN-UNCHECKED"
            log.warning(f"RESULT: {reason}")
            safe_move(filepath, CLEAN_FOLDER, filename)
            append_to_csv(filename)
            append_edm_result_to_awb_logs(awb, filename, result="CLEAN-UNCHECKED", reason=reason, match_stats=match_stats)
            record_edm_end(filename, edm_result="CLEAN-UNCHECKED", final_folder="CLEAN", notes=reason)
            t["route_ms"] = _ms(route_start)
            finalize_audit("CLEAN-UNCHECKED", "CLEAN", reason, match_stats=match_stats)
            emit_pipeline_summary(
                "CLEAN-UNCHECKED",
                total_pages=total_pages,
                duplicate_pages=[p + 1 for p in sorted(duplicate_pages)],
                clean_pages=list(range(1, total_pages + 1)),
                dup_meta=dup_meta,
            )
            return

        route_start = time.perf_counter()
        log.info(f"RESULT: ALL {total_pages} page(s) are duplicates -> REJECTED")
        safe_move(filepath, REJECTED_FOLDER, filename)
        append_to_rejected_sheet(filename, reason=f"All {total_pages} page(s) matched EDM", match_stats=match_stats)
        append_edm_result_to_awb_logs(awb, filename, result="REJECTED", reason=f"All {total_pages} page(s) matched EDM", match_stats=match_stats)
        record_edm_end(filename, edm_result="REJECTED", final_folder="REJECTED")
        t["route_ms"] = _ms(route_start)
        finalize_audit("REJECTED", "REJECTED", f"All {total_pages} page(s) matched EDM", match_stats=match_stats)
        emit_pipeline_summary("REJECTED", total_pages=total_pages, duplicate_pages=[p + 1 for p in sorted(duplicate_pages)], clean_pages=[], dup_meta=dup_meta)
        return

    # Case 3: More than threshold duplicate pages + high ratio + adequate confidence -> REJECTED
    if dup_count > REJECT_IF_DUP_PAGES_OVER and dup_ratio >= REJECT_IF_DUP_RATIO and reject_conf != "LOW":
        route_start = time.perf_counter()
        reason = (
            f"Duplicate threshold exceeded ({dup_count} pages, ratio={dup_ratio:.2f}, confidence={reject_conf}) "
            f"-- routing to REJECTED"
        )
        log.info(f"RESULT: {reason} -> REJECTED")
        safe_move(filepath, REJECTED_FOLDER, filename)
        append_to_rejected_sheet(filename, reason=reason, match_stats=match_stats)
        append_edm_result_to_awb_logs(awb, filename, result="REJECTED", reason=reason, match_stats=match_stats)
        record_edm_end(filename, edm_result="REJECTED", final_folder="REJECTED", notes=reason)
        t["route_ms"] = _ms(route_start)
        finalize_audit("REJECTED", "REJECTED", reason, match_stats=match_stats)
        emit_pipeline_summary(
            "REJECTED",
            total_pages=total_pages,
            duplicate_pages=[p + 1 for p in sorted(duplicate_pages)],
            clean_pages=[],
            dup_meta=dup_meta,
        )
        return

    # Case 4: Mixed -- strip duplicates
    clean_pages = [i for i in range(total_pages) if i not in duplicate_pages]
    log.info(f"RESULT: {len(duplicate_pages)} duplicate page(s) out of {total_pages} total.")
    log.info(f"  Duplicate page(s): {[p + 1 for p in sorted(duplicate_pages)]}")
    log.info(f"  Keeping page(s):   {[p + 1 for p in clean_pages]}")

    try:
        route_start = time.perf_counter()
        src_doc = fitz.open(filepath)

        stripped_doc = fitz.open()
        for p in clean_pages:
            stripped_doc.insert_pdf(src_doc, from_page=p, to_page=p)

        rejected_doc = fitz.open()
        for p in sorted(duplicate_pages):
            rejected_doc.insert_pdf(src_doc, from_page=p, to_page=p)

        src_doc.close()

        tmp_clean = filepath + "_clean.pdf"
        tmp_rejected = filepath + "_rejected.pdf"
        stripped_doc.save(tmp_clean)
        rejected_doc.save(tmp_rejected)
        stripped_doc.close()
        rejected_doc.close()

        os.remove(filepath)

        safe_move(tmp_clean, CLEAN_FOLDER, filename)
        safe_move(tmp_rejected, REJECTED_FOLDER, filename)

        log.info(f"Stripped PDF ({len(clean_pages)} page(s)) -> CLEAN")
        log.info(f"Duplicate pages ({len(duplicate_pages)} page(s)) -> REJECTED")

        append_to_csv(filename)
        append_edm_result_to_awb_logs(awb, filename, result="PARTIAL-CLEAN",
            reason=f"Pages {[p+1 for p in sorted(duplicate_pages)]} matched EDM -- stripped remainder to CLEAN",
            match_stats=match_stats)
        record_edm_end(filename, edm_result="PARTIAL-CLEAN", final_folder="CLEAN",
                       notes=f"Pages {[p+1 for p in sorted(duplicate_pages)]} removed")
        t["route_ms"] = _ms(route_start)
        finalize_audit("PARTIAL-CLEAN", "CLEAN+REJECTED", "Partial duplicates stripped", match_stats=match_stats)
        emit_pipeline_summary(
            "PARTIAL-CLEAN",
            total_pages=total_pages,
            duplicate_pages=[p + 1 for p in sorted(duplicate_pages)],
            clean_pages=[p + 1 for p in clean_pages],
            dup_meta=dup_meta,
        )

    except Exception as e:
        route_start = time.perf_counter()
        log.warning(f"Error stripping pages: {e} -- passing through original as CLEAN-UNCHECKED")
        safe_move(filepath, CLEAN_FOLDER, filename)
        append_to_csv(filename)
        append_edm_result_to_awb_logs(awb, filename, result="CLEAN-UNCHECKED",
            reason=f"Page stripping failed: {e}", match_stats=match_stats)
        record_edm_end(filename, edm_result="CLEAN-UNCHECKED", final_folder="CLEAN",
                       notes=f"Page stripping failed: {e}")
        t["route_ms"] = _ms(route_start)
        finalize_audit("CLEAN-UNCHECKED", "CLEAN", f"Page stripping failed: {e}", match_stats=match_stats)
        emit_pipeline_summary(
            "CLEAN-UNCHECKED",
            total_pages=total_pages,
            duplicate_pages=[p + 1 for p in sorted(duplicate_pages)],
            clean_pages=list(range(1, total_pages + 1)),
            dup_meta=dup_meta,
        )


# =========================
# WATCHDOG
# =========================
class PDFHandler(FileSystemEventHandler):
    def on_created(self, event):
        if event.is_directory or not event.src_path.lower().endswith(".pdf"):
            return
        filepath = event.src_path
        filename = os.path.basename(filepath)
        log.info(f"New file detected: {filename}")
        time.sleep(FILE_SETTLE_SECONDS)
        if not os.path.exists(filepath):
            log.warning(f"File gone before processing: {filename}")
            return
        try:
            process_file(filepath)
        except Exception as e:
            log.error(f"Unexpected error on {filename}: {e}")


if __name__ == "__main__":
    config.ensure_dirs()

    # Check if token is available
    token, token_source = _get_token_and_source()
    if token is None:
        log.warning("No EDM token available. Exiting EDM checker.")
        sys.exit(0)

    log.info("EDM Duplicate Checker -- started")
    log.info(f"Watching:  {PROCESSED_FOLDER}")
    log.info(f"Clean:     {CLEAN_FOLDER}")
    log.info(f"Rejected:  {REJECTED_FOLDER}")
    log.info(f"Similarity threshold: {TEXT_SIMILARITY_THRESHOLD}")
    log.info(f"CSV written to: {CSV_PATH} (CLEAN outcomes only)")
    log.info(f"Token source: {token_source}")

    existing = [f for f in PROCESSED_FOLDER.iterdir() if f.suffix.lower() == ".pdf"]
    if existing:
        log.info(f"Found {len(existing)} existing file(s) -- processing now")
        for fp in existing:
            try:
                process_file(str(fp))
            except Exception as e:
                log.error(f"Error on {fp.name}: {e}")

    observer = Observer()
    observer.schedule(PDFHandler(), str(PROCESSED_FOLDER), recursive=False)
    observer.start()
    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        log.info("Shutting down...")
        observer.stop()
    observer.join()
    log.info("EDM Duplicate Checker -- stopped")
