# Scripts/awb_hotfolder.py
# FAST Local Hot Folder Pipeline (AWB Excel list matching) - PAGE 1 ONLY
#
# All paths and tuning values come from config.py / .env.
# No hardcoded paths in this file.
#
# FLOW:
#   filename -> 400-pattern -> text-layer DB -> OCR(+400) -> strong OCR(+400) -> rotation(+400)
# SAFETY:
#   - NEVER output unless AWB exists in Excel list
#   - 1-digit tolerance ONLY when candidate looks like true AWB context
#   - Tolerance disabled when many 12-digit numbers are detected (noise guard)
#   - 400- rule: 12 digits after "400-" accepted immediately, no DB check needed

import os
import re
import sys
import time
import csv
import shutil
from queue import Queue, Empty
from datetime import datetime
from pathlib import Path

# Allow running from Scripts/ subfolder
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
import config

from Scripts.pipeline_tracker import (
    record_hotfolder_start,
    record_hotfolder_end,
    record_hotfolder_needs_review,
)
from Scripts.audit_logger import audit_event

import fitz  # PyMuPDF
from PIL import Image, ImageOps
import pytesseract
from openpyxl import load_workbook, Workbook
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler

pytesseract.pytesseract.tesseract_cmd = str(config.TESSERACT_PATH)

# ── Convenience aliases from config ──────────────────────────────────────────
INBOX_DIR        = config.INBOX_DIR
PROCESSED_DIR    = config.PROCESSED_DIR
NEEDS_REVIEW_DIR = config.NEEDS_REVIEW_DIR
AWB_EXCEL_PATH   = config.AWB_EXCEL_PATH
AWB_LOGS_PATH    = config.AWB_LOGS_PATH
LOG_DIR          = config.LOG_DIR
CSV_PATH         = config.CSV_PATH
STAGE_CACHE_CSV  = config.STAGE_CACHE_CSV

DPI_MAIN          = config.OCR_DPI_MAIN
DPI_STRONG        = config.OCR_DPI_STRONG
DPI_ROT_PROBE     = config.OCR_DPI_ROT_PROBE
POLL_SECONDS      = config.POLL_SECONDS
HEARTBEAT_SECONDS = config.HEARTBEAT_SECONDS
EXCEL_REFRESH_SECONDS = config.EXCEL_REFRESH_SECONDS
PAGES_TO_SCAN     = config.PAGES_TO_SCAN

AWB_LEN                     = config.AWB_LEN
ALLOW_1_DIGIT_TOLERANCE     = config.ALLOW_1_DIGIT_TOLERANCE
STRICT_AMBIGUOUS            = config.STRICT_AMBIGUOUS
STOP_EARLY_IF_MANY_12DIGITS = config.STOP_EARLY_IF_MANY_12DIGITS
MANY_12DIGITS_THRESHOLD     = config.MANY_12DIGITS_THRESHOLD
DISABLE_TOLERANCE_IF_MANY   = config.DISABLE_TOLERANCE_IF_MANY
MANY_12DIGITS_TOL_CUTOFF    = config.MANY_12DIGITS_TOL_CUTOFF
AWB_CONTEXT_KEYWORDS        = config.AWB_CONTEXT_KEYWORDS
CONTEXT_WINDOW_CHARS        = config.CONTEXT_WINDOW_CHARS
ENABLE_ROTATION_LAST_RESORT = config.ENABLE_ROTATION_LAST_RESORT


# =========================
# FILENAME AWB PRE-CHECK (SKIP OCR)
# =========================
_FILENAME_AWB_12DIGITS    = re.compile(r"(?<!\d)(\d{12})(?!\d)")
_FILENAME_AWB_4SPACE4SPACE4 = re.compile(r"(?<!\d)(\d{4}\s\d{4}\s\d{4})(?!\d)")


def extract_awb_from_filename_strict(filename):
    base = os.path.basename(filename or "")
    m = _FILENAME_AWB_12DIGITS.search(base)
    if m:
        return m.group(1)
    m = _FILENAME_AWB_4SPACE4SPACE4.search(base)
    if m:
        return m.group(1).replace(" ", "")
    return None


# =========================
# 400- PATTERN AWB EXTRACTION (NO DB CHECK)
# =========================
_400_AWB_PATTERN = re.compile(r"400[\s\-.:]{0,6}(\d[\d\s\-]{10,16})", re.IGNORECASE)


def extract_awb_from_400_pattern(text):
    if not text:
        return None
    for m in _400_AWB_PATTERN.finditer(text):
        digits = re.sub(r"\D", "", m.group(1))
        if len(digits) >= 12:
            return digits[:12]
    for m in re.finditer(r"(?<!\d)400(\d{12,14})(?!\d)", text):
        return m.group(1)[:12]
    return None


# =========================
# UTILS
# =========================
def log(msg):
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{ts}] {msg}"
    print(line)
    try:
        LOG_DIR.mkdir(parents=True, exist_ok=True)
        with open(config.PIPELINE_LOG, "a", encoding="utf-8") as f:
            f.write(line + "\n")
    except Exception:
        pass


def require_tesseract():
    if not config.TESSERACT_PATH.exists():
        raise FileNotFoundError(f"Tesseract not found at: {config.TESSERACT_PATH}")


def file_is_stable(path, checks=4, delay=0.5):
    last = -1
    for _ in range(checks):
        try:
            size = os.path.getsize(path)
        except OSError:
            return False
        if size == last and size > 0:
            return True
        last = size
        time.sleep(delay)
    return False


def safe_move(src, dst_dir):
    name = os.path.basename(src)
    dst = Path(dst_dir) / name
    if dst.exists():
        base, ext = os.path.splitext(name)
        dst = Path(dst_dir) / f"{base}_{int(time.time())}{ext}"
    shutil.move(src, dst)


def move_to_processed_renamed(src, awb):
    import hashlib

    def file_md5(path):
        h = hashlib.md5()
        with open(path, "rb") as f:
            for chunk in iter(lambda: f.read(65536), b""):
                h.update(chunk)
        return h.hexdigest()

    dst = PROCESSED_DIR / f"{awb}.pdf"

    if dst.exists():
        try:
            if file_md5(src) == file_md5(dst):
                log(f"DUPLICATE CONTENT for {awb} -- removing source, skipping move.")
                try:
                    os.remove(src)
                except Exception:
                    pass
                return str(dst)
        except Exception:
            pass
        k = 2
        while True:
            dst = PROCESSED_DIR / f"{awb}_{k}.pdf"
            if not dst.exists():
                break
            k += 1

    shutil.move(src, dst)
    return str(dst)


# =========================
# AWB LOGS EXCEL LOGGER
# =========================
_AWB_LOGS_HEADERS = ["AWB", "SourceFile", "Timestamp", "MatchMethod", "Status"]


def append_to_awb_logs_excel(awb, source_file, match_method, status="MATCHED"):
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    row = [awb, os.path.basename(source_file), ts, match_method, status]

    for attempt in range(5):
        try:
            if AWB_LOGS_PATH.exists():
                wb = load_workbook(AWB_LOGS_PATH)
                ws = wb.active
                if ws.max_row == 0 or ws.cell(1, 1).value != "AWB":
                    ws.insert_rows(1)
                    for col, h in enumerate(_AWB_LOGS_HEADERS, start=1):
                        ws.cell(1, col).value = h
            else:
                AWB_LOGS_PATH.parent.mkdir(parents=True, exist_ok=True)
                wb = Workbook()
                ws = wb.active
                ws.title = "AWB Logs"
                ws.append(_AWB_LOGS_HEADERS)

            ws.append(row)
            wb.save(AWB_LOGS_PATH)
            return
        except PermissionError:
            time.sleep(0.4 * (attempt + 1))
        except Exception as e:
            log(f"[AWB_LOGS] Warning: could not write AWB_Logs.xlsx: {e}")
            return

    log(f"[AWB_LOGS] AWB_Logs.xlsx still locked after retries -- skipping log for {awb}.")


def append_stage_cache_row(input_file, processed_file, awb, detection_type, awb_extraction_secs):
    """Lightweight stage cache for EDM summary logging (CSV append)."""
    headers = [
        "Timestamp",
        "InputFileName",
        "ProcessedFileName",
        "AWB_Detected",
        "AWB_Detection_Type",
        "AWB_Extraction_Seconds",
    ]
    row = [
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        os.path.basename(input_file),
        os.path.basename(processed_file),
        awb,
        detection_type,
        awb_extraction_secs,
    ]
    try:
        STAGE_CACHE_CSV.parent.mkdir(parents=True, exist_ok=True)
        new_file = not STAGE_CACHE_CSV.exists()
        with open(STAGE_CACHE_CSV, "a", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            if new_file:
                w.writerow(headers)
            w.writerow(row)
    except Exception as e:
        log(f"[STAGE_CACHE] Warning: could not write stage cache row: {e}")


# =========================
# LOAD AWBs FROM EXCEL
# =========================
def extract_12_digit_numbers_from_any_text(s):
    if s is None:
        return []
    s = str(s)
    out = set()
    for m in re.finditer(r"\b\d{12}\b", s):
        out.add(m.group(0))
    for m in re.finditer(r"(\d[\d\-\s]{10,30}\d)", s):
        d = re.sub(r"\D", "", m.group(0))
        if len(d) == AWB_LEN:
            out.add(d)
    return list(out)


def load_awb_set_from_excel(xlsx_path):
    if not Path(xlsx_path).exists():
        raise FileNotFoundError(f"Excel file not found: {xlsx_path}")
    wb = load_workbook(xlsx_path, data_only=True)
    awbs = set()
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                for n in extract_12_digit_numbers_from_any_text(cell):
                    if len(n) == AWB_LEN and n.isdigit():
                        awbs.add(n)
    return awbs


def build_buckets(awb_set):
    by_prefix = {}
    by_suffix = {}
    for a in awb_set:
        by_prefix.setdefault(a[:4], []).append(a)
        by_suffix.setdefault(a[-4:], []).append(a)
    return by_prefix, by_suffix


# =========================
# CANDIDATE EXTRACTION
# =========================
def extract_candidates_from_text(s):
    s = s or ""
    out = set()
    for m in re.finditer(r"(?<!\d)(\d{12})(?!\d)", s):
        out.add(m.group(1))
    for m in re.finditer(r"(?<!\d)(\d{4}[\s\-]\d{4}[\s\-]\d{4})(?!\d)", s):
        d = re.sub(r"\D", "", m.group(1))
        if len(d) == AWB_LEN:
            out.add(d)
    for m in re.finditer(r"(?<!\d)400[\s\-:]{0,6}([0-9][0-9\-\s]{10,20})(?!\d)", s, re.IGNORECASE):
        d = re.sub(r"\D", "", m.group(1))
        if len(d) >= 12:
            out.add(d[:12])
    for m in re.finditer(r"(?<!\d)ACI[\s\-:]{0,6}([0-9][0-9\-\s]{10,20})(?!\d)", s, re.IGNORECASE):
        d = re.sub(r"\D", "", m.group(1))
        if len(d) >= 12:
            out.add(d[:12])
    return out


def extract_candidates_near_keywords(s):
    s = s or ""
    su = s.upper()
    out = set()
    for m in re.finditer(r"(\d[\d\-\s]{10,30}\d)", s):
        raw = m.group(0)
        d = re.sub(r"\D", "", raw)
        if len(d) != AWB_LEN:
            continue
        start = max(0, m.start() - CONTEXT_WINDOW_CHARS)
        end = min(len(su), m.end() + CONTEXT_WINDOW_CHARS)
        window = su[start:end]
        if any(k in window for k in AWB_CONTEXT_KEYWORDS):
            out.add(d)
    for m in re.finditer(r"\b\d{12}\b", s):
        d = m.group(0)
        start = max(0, m.start() - CONTEXT_WINDOW_CHARS)
        end = min(len(su), m.end() + CONTEXT_WINDOW_CHARS)
        window = su[start:end]
        if any(k in window for k in AWB_CONTEXT_KEYWORDS):
            out.add(d)
    return out


# =========================
# TEXT LAYER (CHEAP) - PAGE 1 ONLY
# =========================
def match_from_pdf_textlayer(pdf_path, awb_set):
    try:
        doc = fitz.open(pdf_path)
    except Exception:
        return None, [], ""
    matches = set()
    txt = ""
    try:
        if doc.page_count > 0:
            txt = doc.load_page(0).get_text("text") or ""
            for c in extract_candidates_from_text(txt):
                if c in awb_set:
                    matches.add(c)
    finally:
        doc.close()
    matches = sorted(matches)
    if len(matches) == 1:
        return matches[0], matches, txt
    return None, matches, txt


# =========================
# OCR HELPERS
# =========================
def render_page(pdf_path, dpi_value):
    doc = fitz.open(pdf_path)
    try:
        page = doc.load_page(0)
        zoom = dpi_value / 72.0
        pix = page.get_pixmap(matrix=fitz.Matrix(zoom, zoom), alpha=False)
        img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
        return img
    finally:
        doc.close()


def preprocess(img, thr=175, invert=False):
    img = img.convert("L")
    img = ImageOps.autocontrast(img)
    if invert:
        img = ImageOps.invert(img)
    img = img.point(lambda p: 255 if p > thr else 0)
    return img


def ocr_digits_only(img, psm=6):
    cfg = (
        f"--oem 3 --psm {psm} "
        "-c tessedit_char_whitelist=0123456789 "
        "-c preserve_interword_spaces=1 "
    )
    return pytesseract.image_to_string(img, config=cfg)


def digit_score(s):
    return sum(1 for ch in (s or "") if ch.isdigit())


# =========================
# MATCHING
# =========================
def hamming(a, b):
    return sum(1 for x, y in zip(a, b) if x != y)


def pick_unique_close_match(candidate, awb_set, by_prefix, by_suffix, max_distance=2):
    pool = set()
    pool.update(by_prefix.get(candidate[:4], []))
    pool.update(by_suffix.get(candidate[-4:], []))
    if not pool:
        pool = awb_set
    scored = [(a, hamming(candidate, a)) for a in pool if hamming(candidate, a) <= max_distance]
    if not scored:
        return None
    scored.sort(key=lambda x: x[1])
    best_awb, best_d = scored[0]
    if len([a for a, d in scored if d == best_d]) != 1:
        return None
    return best_awb


def pick_tolerance_match_from_candidates(candidates, awb_set, by_prefix, by_suffix, max_distance=2):
    corrected = set()
    for c in candidates:
        if len(c) == AWB_LEN and c.isdigit():
            cm = pick_unique_close_match(c, awb_set, by_prefix, by_suffix, max_distance)
            if cm:
                corrected.add(cm)
    if len(corrected) == 1:
        return list(corrected)[0]
    return None


def decide_from_candidates(candidates, awb_set, by_prefix, by_suffix, allow_tolerance):
    exact = sorted(list(candidates & awb_set))
    if len(exact) == 1:
        return exact[0], exact
    if len(exact) > 1:
        return None, exact
    if not allow_tolerance or not ALLOW_1_DIGIT_TOLERANCE:
        return None, []
    close_matches = set()
    for c in candidates:
        if len(c) == AWB_LEN and c.isdigit():
            cm = pick_unique_close_match(c, awb_set, by_prefix, by_suffix)
            if cm:
                close_matches.add(cm)
    close_matches = sorted(close_matches)
    if len(close_matches) == 1:
        return close_matches[0], close_matches
    return None, close_matches


# =========================
# OCR PIPELINE
# =========================
def ocr_pass(img, invert=False, psm_list=(6, 11), thr=175):
    p = preprocess(img, thr=thr, invert=invert)
    return "\n".join(ocr_digits_only(p, psm=psm) for psm in psm_list)


def match_from_ocr_fullpage(img, awb_set, by_prefix, by_suffix):
    txt = ocr_pass(img, invert=False, psm_list=(6, 11), thr=175)
    cands_all = extract_candidates_from_text(txt)
    awb, matches = decide_from_candidates(cands_all, awb_set, by_prefix, by_suffix, allow_tolerance=False)
    return awb, matches, cands_all, set(), txt


def match_from_ocr_fullpage_strong(img, awb_set, by_prefix, by_suffix):
    txt1 = ocr_pass(img, invert=False, psm_list=(6, 11), thr=170)
    txt2 = ocr_pass(img, invert=True,  psm_list=(6, 11), thr=200)
    txt = txt1 + "\n" + txt2
    cands_all = extract_candidates_from_text(txt)
    awb, matches = decide_from_candidates(cands_all, awb_set, by_prefix, by_suffix, allow_tolerance=False)
    return awb, matches, cands_all, set(), txt


# =========================
# PROCESSOR
# =========================
def process_pdf(pdf_path, awb_set, by_prefix, by_suffix):
    start_ts = time.perf_counter()
    timings = {
        "filename_ms": 0.0,
        "text_layer_ms": 0.0,
        "ocr_main_ms": 0.0,
        "ocr_strong_ms": 0.0,
        "rotation_ms": 0.0,
        "total_active_ms": 0.0,
    }

    def finalize(status, route, reason, match_method, awb=None):
        timings["total_active_ms"] = round((time.perf_counter() - start_ts) * 1000, 1)
        audit_event(
            "AWB_HOTFOLDER",
            file=name,
            awb=awb,
            status=status,
            route=route,
            match_method=match_method,
            reason=reason,
            timings_ms=timings,
        )
        log(
            f"[TIMING] file={name} method={match_method} route={route} "
            f"filename_ms={timings['filename_ms']} text_layer_ms={timings['text_layer_ms']} "
            f"ocr_main_ms={timings['ocr_main_ms']} ocr_strong_ms={timings['ocr_strong_ms']} "
            f"rotation_ms={timings['rotation_ms']} total_active_ms={timings['total_active_ms']}"
        )

    def awb_extract_secs():
        return round((time.perf_counter() - start_ts), 3)

    if not file_is_stable(pdf_path):
        name = os.path.basename(pdf_path)
        finalize("SKIPPED", "INBOX", "File was not stable yet", "StabilityCheck")
        return

    name = os.path.basename(pdf_path)
    log(f"Processing: {name}")
    record_hotfolder_start(name)

    # 1) Filename shortcut
    fn_start = time.perf_counter()
    awb_from_name = extract_awb_from_filename_strict(name)
    timings["filename_ms"] = round((time.perf_counter() - fn_start) * 1000, 1)
    if awb_from_name:
        log(f"AWB USED (filename strict, no DB check): {awb_from_name} ({name})")
        append_to_awb_logs_excel(awb_from_name, pdf_path, match_method="Filename")
        processed_path = move_to_processed_renamed(pdf_path, awb_from_name)
        processed_name = os.path.basename(processed_path)
        append_stage_cache_row(name, processed_name, awb_from_name, "Filename", awb_extract_secs())
        record_hotfolder_end(name, awb_from_name, processed_name, "Filename")
        finalize("MATCHED", "PROCESSED", "Matched by strict filename pattern", "Filename", awb=awb_from_name)
        return

    # 2) Text-layer + 400-pattern + DB match
    tl_start = time.perf_counter()
    awb, matches, txt_layer = match_from_pdf_textlayer(pdf_path, awb_set)
    awb_400 = extract_awb_from_400_pattern(txt_layer)
    timings["text_layer_ms"] = round((time.perf_counter() - tl_start) * 1000, 1)

    if awb_400:
        log(f"AWB MATCHED (text-layer 400-pattern): {awb_400} ({name})")
        append_to_awb_logs_excel(awb_400, pdf_path, match_method="TextLayer-400")
        processed_path = move_to_processed_renamed(pdf_path, awb_400)
        processed_name = os.path.basename(processed_path)
        append_stage_cache_row(name, processed_name, awb_400, "TextLayer-400", awb_extract_secs())
        record_hotfolder_end(name, awb_400, processed_name, "TextLayer-400")
        finalize("MATCHED", "PROCESSED", "Matched via text-layer 400 pattern", "TextLayer-400", awb=awb_400)
        return

    if awb:
        log(f"AWB MATCHED (text-layer): {awb} ({name})")
        append_to_awb_logs_excel(awb, pdf_path, match_method="Text-Layer")
        processed_path = move_to_processed_renamed(pdf_path, awb)
        processed_name = os.path.basename(processed_path)
        append_stage_cache_row(name, processed_name, awb, "Text-Layer", awb_extract_secs())
        record_hotfolder_end(name, awb, processed_name, "Text-Layer")
        finalize("MATCHED", "PROCESSED", "Matched exact AWB in text layer", "Text-Layer", awb=awb)
        return

    if matches and STRICT_AMBIGUOUS:
        log(f"AMBIGUOUS (text-layer {len(matches)}) -> Needs review: {name} | {matches[:8]}")
        safe_move(pdf_path, NEEDS_REVIEW_DIR)
        record_hotfolder_needs_review(name, f"Ambiguous text-layer matches: {matches[:8]}")
        finalize("NEEDS-REVIEW", "NEEDS_REVIEW", f"Ambiguous text-layer matches: {matches[:8]}", "Text-Layer")
        return

    # 3) OCR main
    main_start = time.perf_counter()
    img_main = render_page(pdf_path, DPI_MAIN)
    awb2, ocr_matches, cands, _, txt_main = match_from_ocr_fullpage(img_main, awb_set, by_prefix, by_suffix)
    timings["ocr_main_ms"] = round((time.perf_counter() - main_start) * 1000, 1)

    if awb2:
        log(f"AWB MATCHED (OCR main): {awb2} ({name})")
        append_to_awb_logs_excel(awb2, pdf_path, match_method="OCR-Main")
        processed_path = move_to_processed_renamed(pdf_path, awb2)
        processed_name = os.path.basename(processed_path)
        append_stage_cache_row(name, processed_name, awb2, "OCR-Main", awb_extract_secs())
        record_hotfolder_end(name, awb2, processed_name, "OCR-Main")
        finalize("MATCHED", "PROCESSED", "Matched exact AWB in OCR main pass", "OCR-Main", awb=awb2)
        return

    if ocr_matches and STRICT_AMBIGUOUS:
        log(f"AMBIGUOUS (OCR main {len(ocr_matches)}) -> Needs review: {name} | {ocr_matches[:8]}")
        safe_move(pdf_path, NEEDS_REVIEW_DIR)
        record_hotfolder_needs_review(name, f"Ambiguous OCR-Main matches: {ocr_matches[:8]}")
        finalize("NEEDS-REVIEW", "NEEDS_REVIEW", f"Ambiguous OCR-main matches: {ocr_matches[:8]}", "OCR-Main")
        return

    if STOP_EARLY_IF_MANY_12DIGITS and len(cands) >= MANY_12DIGITS_THRESHOLD:
        log(f"NO MATCH (many 12-digit cands={len(cands)}) -> Needs review: {name}")
        safe_move(pdf_path, NEEDS_REVIEW_DIR)
        record_hotfolder_needs_review(name, f"Too many OCR-main candidates: {len(cands)}")
        finalize("NEEDS-REVIEW", "NEEDS_REVIEW", f"Too many OCR-main candidates: {len(cands)}", "OCR-Main")
        return

    if 0 < len(cands) <= 2:
        tol_awb = pick_tolerance_match_from_candidates(cands, awb_set, by_prefix, by_suffix, max_distance=2)
        if tol_awb:
            log(f"AWB MATCHED (OCR main tolerance<=2): {tol_awb} ({name})")
            append_to_awb_logs_excel(tol_awb, pdf_path, match_method="OCR-Main-Tolerance2")
            processed_path = move_to_processed_renamed(pdf_path, tol_awb)
            processed_name = os.path.basename(processed_path)
            append_stage_cache_row(name, processed_name, tol_awb, "OCR-Main-Tolerance2", awb_extract_secs())
            record_hotfolder_end(name, tol_awb, processed_name, "OCR-Main-Tolerance2")
            finalize("MATCHED", "PROCESSED", "Matched by OCR-main tolerance <=2", "OCR-Main-Tolerance2", awb=tol_awb)
            return

    # 4) OCR strong
    strong_start = time.perf_counter()
    img_strong = render_page(pdf_path, DPI_STRONG)
    awb3, strong_matches, cands2, _, txt_strong = match_from_ocr_fullpage_strong(img_strong, awb_set, by_prefix, by_suffix)
    timings["ocr_strong_ms"] = round((time.perf_counter() - strong_start) * 1000, 1)

    if awb3:
        log(f"AWB MATCHED (OCR strong 0deg): {awb3} ({name})")
        append_to_awb_logs_excel(awb3, pdf_path, match_method="OCR-Strong-0deg")
        processed_path = move_to_processed_renamed(pdf_path, awb3)
        processed_name = os.path.basename(processed_path)
        append_stage_cache_row(name, processed_name, awb3, "OCR-Strong-0deg", awb_extract_secs())
        record_hotfolder_end(name, awb3, processed_name, "OCR-Strong-0deg")
        finalize("MATCHED", "PROCESSED", "Matched exact AWB in OCR-strong 0deg pass", "OCR-Strong-0deg", awb=awb3)
        return

    if strong_matches and STRICT_AMBIGUOUS:
        log(f"AMBIGUOUS (OCR strong {len(strong_matches)}) -> Needs review: {name} | {strong_matches[:8]}")
        safe_move(pdf_path, NEEDS_REVIEW_DIR)
        record_hotfolder_needs_review(name, f"Ambiguous OCR-Strong matches: {strong_matches[:8]}")
        finalize("NEEDS-REVIEW", "NEEDS_REVIEW", f"Ambiguous OCR-strong matches: {strong_matches[:8]}", "OCR-Strong")
        return

    if STOP_EARLY_IF_MANY_12DIGITS and len(cands2) >= MANY_12DIGITS_THRESHOLD:
        log(f"NO MATCH (many 12-digit cands={len(cands2)}) -> Needs review: {name}")
        safe_move(pdf_path, NEEDS_REVIEW_DIR)
        record_hotfolder_needs_review(name, f"Too many OCR-strong candidates: {len(cands2)}")
        finalize("NEEDS-REVIEW", "NEEDS_REVIEW", f"Too many OCR-strong candidates: {len(cands2)}", "OCR-Strong")
        return

    if 0 < len(cands2) <= 2:
        tol_awb2 = pick_tolerance_match_from_candidates(cands2, awb_set, by_prefix, by_suffix, max_distance=2)
        if tol_awb2:
            log(f"AWB MATCHED (OCR strong tolerance<=2): {tol_awb2} ({name})")
            append_to_awb_logs_excel(tol_awb2, pdf_path, match_method="OCR-Strong-Tolerance2")
            processed_path = move_to_processed_renamed(pdf_path, tol_awb2)
            processed_name = os.path.basename(processed_path)
            append_stage_cache_row(name, processed_name, tol_awb2, "OCR-Strong-Tolerance2", awb_extract_secs())
            record_hotfolder_end(name, tol_awb2, processed_name, "OCR-Strong-Tolerance2")
            finalize("MATCHED", "PROCESSED", "Matched by OCR-strong tolerance <=2", "OCR-Strong-Tolerance2", awb=tol_awb2)
            return

    # 5) Rotation last resort
    if ENABLE_ROTATION_LAST_RESORT:
        rot_start = time.perf_counter()
        for rot in [90, 180, 270]:
            rotated = img_strong.rotate(rot, expand=True)
            awb4, rot_matches, _, _, txt_rot = match_from_ocr_fullpage_strong(rotated, awb_set, by_prefix, by_suffix)
            if awb4:
                timings["rotation_ms"] = round((time.perf_counter() - rot_start) * 1000, 1)
                log(f"AWB MATCHED (OCR strong rot={rot}deg): {awb4} ({name})")
                append_to_awb_logs_excel(awb4, pdf_path, match_method=f"OCR-Strong-{rot}deg")
                processed_path = move_to_processed_renamed(pdf_path, awb4)
                processed_name = os.path.basename(processed_path)
                append_stage_cache_row(name, processed_name, awb4, f"OCR-Strong-{rot}deg", awb_extract_secs())
                record_hotfolder_end(name, awb4, processed_name, f"OCR-Strong-{rot}deg")
                finalize("MATCHED", "PROCESSED", f"Matched in OCR-strong rotation {rot}deg", f"OCR-Strong-{rot}deg", awb=awb4)
                return
            if rot_matches and STRICT_AMBIGUOUS:
                timings["rotation_ms"] = round((time.perf_counter() - rot_start) * 1000, 1)
                log(f"AMBIGUOUS (OCR rot={rot}deg {len(rot_matches)}) -> Needs review: {name}")
                safe_move(pdf_path, NEEDS_REVIEW_DIR)
                record_hotfolder_needs_review(name, f"Ambiguous OCR-Strong-{rot}deg: {rot_matches[:8]}")
                finalize("NEEDS-REVIEW", "NEEDS_REVIEW", f"Ambiguous OCR-rotation {rot}deg matches: {rot_matches[:8]}", f"OCR-Strong-{rot}deg")
                return
        timings["rotation_ms"] = round((time.perf_counter() - rot_start) * 1000, 1)

    # 6) Final review
    all_tried = set()
    all_tried.update(cands if "cands" in dir() else [])
    all_tried.update(cands2 if "cands2" in dir() else [])
    log(f"NO MATCH FOUND -> Needs review: {name}")
    log(f"  Candidates tried: {sorted(all_tried)}")
    safe_move(pdf_path, NEEDS_REVIEW_DIR)
    record_hotfolder_needs_review(name, f"No AWB match found after all passes | cands={sorted(all_tried)}")
    finalize("NEEDS-REVIEW", "NEEDS_REVIEW", f"No AWB match after all passes | cands={sorted(all_tried)}", "No-Match")


# =========================
# MAIN
# =========================
class InboxPDFHandler(FileSystemEventHandler):
    def __init__(self, q):
        self.q = q
        self._last_seen = {}

    def _enqueue(self, path):
        p = str(path)
        if not p.lower().endswith(".pdf"):
            return
        now = time.time()
        prev = self._last_seen.get(p, 0)
        if now - prev < 0.8:
            return
        self._last_seen[p] = now
        self.q.put(p)

    def on_created(self, event):
        if not event.is_directory:
            self._enqueue(event.src_path)

    def on_moved(self, event):
        if not event.is_directory:
            self._enqueue(event.dest_path)

    def on_modified(self, event):
        if not event.is_directory:
            self._enqueue(event.src_path)


def main():
    config.ensure_dirs()
    require_tesseract()

    awb_set = set()
    by_prefix = {}
    by_suffix = {}
    last_excel_mtime = 0
    last_excel_load = 0
    last_heartbeat = 0
    last_rescan = 0

    file_queue = Queue()
    handler = InboxPDFHandler(file_queue)
    observer = Observer()
    observer.schedule(handler, str(INBOX_DIR), recursive=False)
    observer.start()

    log("=== AWB Hot Folder Pipeline started ===")
    log(f"INBOX:  {INBOX_DIR}")
    log(f"EXCEL:  {AWB_EXCEL_PATH}")
    log(f"LOGS:   {AWB_LOGS_PATH}")
    log("Flow: filename -> 400-pattern -> text-layer DB -> OCR(+400) -> strong OCR(+400) -> rotation(+400)")
    log("Safety: tolerance only in AWB context; disabled when many candidates exist")
    log("Mode: watchdog event-driven with periodic safety rescan")

    # Startup catch-up: queue existing PDFs.
    try:
        for fn in INBOX_DIR.iterdir():
            if fn.suffix.lower() == ".pdf":
                handler._enqueue(str(fn))
    except Exception as e:
        log(f"Startup scan warning: {e}")

    try:
        while True:
            loop_sleep = POLL_SECONDS
            try:
                now = time.time()

                if now - last_excel_load >= EXCEL_REFRESH_SECONDS:
                    try:
                        mtime = AWB_EXCEL_PATH.stat().st_mtime
                    except Exception:
                        mtime = 0
                    if mtime != last_excel_mtime:
                        awb_set = load_awb_set_from_excel(AWB_EXCEL_PATH)
                        by_prefix, by_suffix = build_buckets(awb_set)
                        last_excel_mtime = mtime
                        log(f"Loaded AWBs: {len(awb_set)} (Excel refreshed)")
                    last_excel_load = now

                if now - last_heartbeat >= HEARTBEAT_SECONDS:
                    try:
                        file_count = len([x for x in INBOX_DIR.iterdir() if x.suffix.lower() == ".pdf"])
                    except Exception:
                        file_count = -1
                    log(f"Watching INBOX | PDF Files: {file_count} | AWBs loaded: {len(awb_set)}")
                    last_heartbeat = now

                # Safety rescan keeps pipeline alive for files that were unstable on first event.
                if now - last_rescan >= max(POLL_SECONDS, 3):
                    try:
                        for fn in INBOX_DIR.iterdir():
                            if fn.suffix.lower() == ".pdf":
                                handler._enqueue(str(fn))
                    except Exception as e:
                        log(f"Rescan warning: {e}")
                    last_rescan = now

                processed_any = False
                while True:
                    try:
                        path = file_queue.get_nowait()
                    except Empty:
                        break
                    if os.path.exists(path) and path.lower().endswith(".pdf"):
                        process_pdf(str(path), awb_set, by_prefix, by_suffix)
                        processed_any = True

                if processed_any:
                    loop_sleep = 0.2

            except Exception as e:
                log(f"LOOP ERROR: {e}")

            time.sleep(loop_sleep)
    except KeyboardInterrupt:
        log("Shutting down hotfolder watcher...")
    finally:
        observer.stop()
        observer.join()


if __name__ == "__main__":
    main()
