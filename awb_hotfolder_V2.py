# Scripts/awb_hotfolder.py
# FAST Local Hot Folder Pipeline (AWB Excel list matching) - PAGE 1 ONLY
#
# All paths and tuning values come from config.py / .env.
# No hardcoded paths in this file.
#
# FLOW:
#   filename -> pre-angle checks -> text-layer(+set_rotation+word-sort) ->
#   OCR main/strong at 0 -> probe (only on fail) -> route A/B/C ->
#   3.5 ROI -> [A/B: 5.5 upscale -> 5 table -> 4 rotation last-resort]
#              [C:   4 rotation probe-ordered -> 5 table -> 5.5 upscale]
#   -> 5.6 airway-label -> 6 EDM -> 7 needs-review
#
# TWO-PASS SCHEDULING:
#   Fast lane: Stages 0-3.5 only. Defer if no match.
#   Long lane: Full pipeline on deferred docs when fast queue is empty.
#
# SAFETY:
#   - NEVER output unless AWB exists in Excel list
#   - Tolerance only after multi-stage evidence, leading-zero relaxed for tolerance only
#   - 400- rule: 12 digits after 400- accepted immediately, no DB check
#   - EDM fallback only for HIGH-confidence persistent candidates

import os
import re
import sys
import time
import csv
import json
import shutil
from collections import Counter
from queue import Queue, Empty
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
import config

from Scripts.pipeline_tracker import (
    record_hotfolder_start,
    record_hotfolder_end,
    record_hotfolder_needs_review,
)
from Scripts.audit_logger import audit_event

try:
    import fitz
except Exception:
    try:
        import pymupdf as fitz
    except Exception as exc:
        raise RuntimeError("PyMuPDF import failed.") from exc

from PIL import Image, ImageOps
import pytesseract
from pytesseract import Output
import requests
from openpyxl import load_workbook, Workbook
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler

pytesseract.pytesseract.tesseract_cmd = str(config.TESSERACT_PATH)

try:
    import cv2
    import numpy as np
    CV2_AVAILABLE = True
except ImportError:
    CV2_AVAILABLE = False

# ── Convenience aliases from config ──────────────────────────────────────────
INBOX_DIR        = config.INBOX_DIR
PROCESSED_DIR    = config.PROCESSED_DIR
NEEDS_REVIEW_DIR = config.NEEDS_REVIEW_DIR
AWB_EXCEL_PATH   = config.AWB_EXCEL_PATH
AWB_LOGS_PATH    = config.AWB_LOGS_PATH
LOG_DIR          = config.LOG_DIR
CSV_PATH         = config.CSV_PATH
STAGE_CACHE_CSV  = config.STAGE_CACHE_CSV
EDM_EXISTS_CACHE_PATH = config.EDM_AWB_EXISTS_CACHE

DPI_MAIN          = config.OCR_DPI_MAIN
DPI_STRONG        = config.OCR_DPI_STRONG
POLL_SECONDS      = config.POLL_SECONDS
HEARTBEAT_SECONDS = config.HEARTBEAT_SECONDS
EXCEL_REFRESH_SECONDS = config.EXCEL_REFRESH_SECONDS

AWB_LEN                     = config.AWB_LEN
ALLOW_1_DIGIT_TOLERANCE     = config.ALLOW_1_DIGIT_TOLERANCE
STRICT_AMBIGUOUS            = config.STRICT_AMBIGUOUS
STOP_EARLY_IF_MANY_12DIGITS = config.STOP_EARLY_IF_MANY_12DIGITS
MANY_12DIGITS_THRESHOLD     = config.MANY_12DIGITS_THRESHOLD
ENABLE_ROTATION_LAST_RESORT = config.ENABLE_ROTATION_LAST_RESORT
EDM_OPERATING_COMPANY       = config.EDM_OPERATING_COMPANY
EDM_METADATA_URL            = config.EDM_METADATA_URL
CONTEXT_WINDOW_CHARS        = int(getattr(config, "CONTEXT_WINDOW_CHARS", 40))

AWB_CONTEXT_KEYWORDS = tuple(
    dict.fromkeys(
        list(getattr(config, "AWB_CONTEXT_KEYWORDS", (
            "AWB", "AIR WAYBILL", "AIRWAY BILL", "AIRWAYBILL",
            "AIRWAY BILL NUMBER", "AIRWAYBILL NUMBER", "WAYBILL",
            "TRACKING", "TRACKING NUMBER", "SHIPMENT", "MASTER",
            "MAWB", "HAWB", "BILL NO", "BOL",
        )))
        + [
            "COMMERCIAL INVOICE", "C/I", "CI NO", "CI NUMBER",
            "AWB NO", "AWB NUMBER", "AWB#", "TRACKING #", "TRACKING NUM",
            "FEDEX TRACKING", "FED EX TRACKING", "AIR WAY BILL",
            "AIR WAY BILL NUMBER", "AIR WAYBILL NUMBER", "WAY BILL",
            "ACI", "ACI NO", "ACI NUMBER", "CARGO CONTROL NUMBER",
            "CARGO CONTROL NO", "CCN", "CONSIGNMENT", "CONSIGNMENT NO",
            "CONSIGNMENT NUMBER", "FDX", "FDE", "FDXE", "FEDEX", "FED-EX",
            "FDX TRACKING", "FDXE TRACKING", "SHIP",
            "TRK", "TRK#", "TRK NO", "TRK NUMBER", "TRACKING NO",
            "B/L", "B/L NO", "B/L NUMBER", "BL NO", "BL NUMBER",
        ]
    )
)

OCR_MAIN_PSMS    = tuple(getattr(config, "OCR_MAIN_PSMS", (6, 11)))
OCR_STRONG_PSMS  = tuple(getattr(config, "OCR_STRONG_PSMS", (6, 11)))
ROTATION_PROBE_DPI = int(getattr(config, "ROTATION_PROBE_DPI", 140))


def _cfg_bool(name, default):
    v = getattr(config, name, default)
    if isinstance(v, bool):
        return v
    if isinstance(v, str):
        return v.strip().lower() in ("1", "true", "yes", "on")
    return bool(v)


HARD_DOC_MODE_ENABLED    = _cfg_bool("HARD_DOC_MODE_ENABLED", False)
LOG_STAGE_SNAPSHOTS      = _cfg_bool("LOG_STAGE_SNAPSHOTS", True)
CANDIDATE_SNAPSHOT_LIMIT = int(getattr(config, "CANDIDATE_SNAPSHOT_LIMIT", 20))

ALLOW_STANDARD_TOLERANCE     = _cfg_bool("ALLOW_STANDARD_TOLERANCE", True)
TOLERANCE_HIGH_MAX_DISTANCE  = int(getattr(config, "TOLERANCE_HIGH_MAX_DISTANCE", 2))
TOLERANCE_STANDARD_MAX_DISTANCE = int(getattr(config, "TOLERANCE_STANDARD_MAX_DISTANCE", 1))
MIN_STAGE_HITS_HIGH_TOL1     = int(getattr(config, "MIN_STAGE_HITS_HIGH_TOL1", 1))
MIN_STAGE_HITS_HIGH_TOL2     = int(getattr(config, "MIN_STAGE_HITS_HIGH_TOL2", 2))
MIN_STAGE_HITS_STANDARD_TOL  = int(getattr(config, "MIN_STAGE_HITS_STANDARD_TOL", 2))
REQUIRE_SINGLE_STANDARD_CANDIDATE_FOR_TOL = _cfg_bool(
    "REQUIRE_SINGLE_STANDARD_CANDIDATE_FOR_TOL", True
)

ENABLE_UPSCALED_RESCUE_PASS  = _cfg_bool("ENABLE_UPSCALED_RESCUE_PASS", True)
ENABLE_AIRWAY_LABEL_RESCUE   = _cfg_bool("ENABLE_AIRWAY_LABEL_RESCUE", True)
FAST_FIRST_SINGLE_DEEP_FALLBACK = _cfg_bool("FAST_FIRST_SINGLE_DEEP_FALLBACK", False)
ENABLE_INBOX_TWO_PASS        = _cfg_bool("ENABLE_INBOX_TWO_PASS", True)

ROTATION_PROBE_MIN_FLIP_MARGIN      = int(getattr(config, "ROTATION_PROBE_MIN_FLIP_MARGIN", 80))
ROTATION_PROBE_DIGIT_CLEAR_MARGIN   = int(getattr(config, "ROTATION_PROBE_DIGIT_CLEAR_MARGIN", 24))
ROTATION_PROBE_CERTAIN_MARGIN       = int(getattr(config, "ROTATION_PROBE_CERTAIN_MARGIN", 300))
ROTATION_PROBE_LIKELY_MARGIN        = int(getattr(config, "ROTATION_PROBE_LIKELY_MARGIN", 120))
LONG_PASS_TIMEOUT_SECONDS           = float(getattr(config, "LONG_PASS_TIMEOUT_SECONDS", 75.0))


class _TimeoutDeferred(Exception):
    """Raised inside process_pdf when the per-file long-pass time budget is exceeded.
    Caught at the top of process_pdf; state is captured and the file is queued
    for the third-pass tier."""


# =============================================================================
# FILENAME PRE-CHECK
# =============================================================================
_FILENAME_AWB_12DIGITS      = re.compile(r"(?<!\d)(\d{12})(?!\d)")
_FILENAME_AWB_4SPACE4SPACE4 = re.compile(r"(?<!\d)(\d{4}\s\d{4}\s\d{4})(?!\d)")


def extract_awb_from_filename_strict(filename):
    """
    Find a 12-digit AWB anywhere in the filename.
    No DB check — filename is trusted as authoritative.
    Handles: bare 12 digits, 4-4-4 grouped, 400-prefix, any other
    surrounding text — the lookbehind/lookahead isolates the number.
    Examples:
      400-399617498819.pdf        → 399617498819
      20260317_399617498819.pdf   → 399617498819
      1234 5678 9012_scan.pdf     → 123456789012
      randomtext399617498819x.pdf → 399617498819
    """
    base = os.path.basename(filename or "")
    m = _FILENAME_AWB_12DIGITS.search(base)
    if m:
        return m.group(1)
    m = _FILENAME_AWB_4SPACE4SPACE4.search(base)
    if m:
        return m.group(1).replace(" ", "")
    return None


# =============================================================================
# 400 TIGHT PATTERN — ONLY EXCEPTION TO DB CHECK
# =============================================================================
# extract_awb_from_400_pattern returns a candidate WITHOUT a DB check.
# This exemption is ONLY for the tight "400-" / "400 " / "400:" prefix where
# the number follows immediately with no label words between them.
# Everything else (400 NUMBER:, ACI:, AWB:, etc.) goes through the normal
# DB-checked tiered extraction pipeline via _candidates_from_label_prefixes
# and prioritize_db_match.
_400_AWB_PATTERN = re.compile(
    r"(?<!\d)400(?:[\s\-]{0,2})(\d(?:[\s\-]?\d){11})(?!\d)", re.IGNORECASE
)


def extract_awb_from_400_pattern(text):
    """
    Returns a 400-prefix AWB without a DB check — tight format only.
    Format: 400-NNNNNNNNNNNN, 400 NNNNNNNNNNNN, 400:NNNNNNNNNNNN, 400NNNNNNNNNNNN.
    Labeled variants (400 NUMBER:, 400 NO:) are intentionally excluded — they
    go through extract_tiered_candidates which always checks the DB first.
    """
    if not text:
        return None
    for m in _400_AWB_PATTERN.finditer(text):
        digits = re.sub(r"\D", "", m.group(1))
        if len(digits) == 12 and not _is_disqualified_candidate(digits):
            return digits
    for m in re.finditer(r"(?<!\d)400(\d{12})(?!\d)", text):
        d = m.group(1)
        if not _is_disqualified_candidate(d):
            return d
    return None


# =============================================================================
# ALL OTHER LABEL PATTERNS — DB CHECK ALWAYS APPLIES
# These feed _candidates_from_label_prefixes → extract_tiered_candidates →
# prioritize_db_match. The DB is always checked before accepting any match.
# =============================================================================
_400_LABELED_PATTERN = re.compile(
    r"(?<!\d)400\s*(?:NO\.?|NUM\.?|NUMBER|#)\s*[:\-]?\s*"
    r"(\d{12}|\d{4}[\s\-]\d{4}[\s\-]\d{4})",
    re.IGNORECASE,
)
_ACI_AWB_PATTERN = re.compile(
    r"(?<!\w)(?:A\s*[CGE6]\s*[I1L])\b[\D]{0,15}(\d(?:[\s\-]?\d){11})(?!\d)",
    re.IGNORECASE,
)
_ACI_LABELED_PATTERN = re.compile(
    r"(?<!\w)(?:A\s*[CGE6]\s*[I1L])\s*(?:NO\.?|NUM\.?|NUMBER|#)"
    r"\s*[:\-]?\s*(\d{12}|\d{4}[\s\-]\d{4}[\s\-]\d{4})",
    re.IGNORECASE,
)
_AWB_LABEL_PATTERN = re.compile(
    r"(?<!\w)(HAWB|MAWB|AWB(?:\s*(?:NO|NUMBER))?)\b[\D]{0,15}(\d[\d\-\s]{10,24})",
    re.IGNORECASE,
)
_TRACK_LABEL_PATTERN = re.compile(
    r"(?<!\w)(?:TRACK(?:ING)?(?:\s*(?:NO|NUMBER|#))?|TRK(?:\s*(?:NO|NUMBER|#))?)"
    r"\b[\D]{0,20}([A-Z0-9][A-Z0-9\-\s:/._]{10,30})",
    re.IGNORECASE,
)
_AIRWAY_BILL_LABEL_PATTERN = re.compile(
    r"(?<!\w)AIR\W*WAY\W*BIL{1,2}(?:\W*(?:NO|NUMBER|#))?\b[\D]{0,30}"
    r"([A-Z0-9][A-Z0-9\-\s:/._]{10,30})",
    re.IGNORECASE,
)
_FEDEX_CARRIER_ROW_PATTERN = re.compile(
    r"(?:FED[\s\-]*EX|FEDEX)[\D]{0,30}(\d{12}|\d{4}[\s\-]\d{4}[\s\-]\d{4})",
    re.IGNORECASE,
)


# =============================================================================
# UTILS
# =============================================================================
def log(msg):
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{ts}] {msg}"
    print(line)
    try:
        LOG_DIR.mkdir(parents=True, exist_ok=True)
        with open(config.PIPELINE_LOG, "a", encoding="utf-8") as f:
            f.write(line + "\n")
    except Exception:
        pass


def require_tesseract():
    if not config.TESSERACT_PATH.exists():
        raise FileNotFoundError(f"Tesseract not found at: {config.TESSERACT_PATH}")


def file_is_stable(path, checks=4, delay=0.5):
    last = -1
    for _ in range(checks):
        try:
            size = os.path.getsize(path)
        except OSError:
            return False
        if size == last and size > 0:
            return True
        last = size
        time.sleep(delay)
    return False


def safe_move(src, dst_dir):
    name = os.path.basename(src)
    dst = Path(dst_dir) / name
    if dst.exists():
        base, ext = os.path.splitext(name)
        dst = Path(dst_dir) / f"{base}_{int(time.time())}{ext}"
    shutil.move(src, dst)


def move_to_processed_renamed(src, awb):
    import hashlib

    def file_md5(path):
        h = hashlib.md5()
        with open(path, "rb") as f:
            for chunk in iter(lambda: f.read(65536), b""):
                h.update(chunk)
        return h.hexdigest()

    dst = PROCESSED_DIR / f"{awb}.pdf"
    if dst.exists():
        try:
            if file_md5(src) == file_md5(dst):
                log(f"DUPLICATE CONTENT for {awb} -- removing source, skipping move.")
                try:
                    os.remove(src)
                except Exception:
                    pass
                return str(dst)
        except Exception:
            pass
        k = 2
        while True:
            dst = PROCESSED_DIR / f"{awb}_{k}.pdf"
            if not dst.exists():
                break
            k += 1
    shutil.move(src, dst)
    return str(dst)


def _normalize_token(raw):
    if not raw:
        return None
    token = str(raw).strip().strip('"').strip("'")
    if token.lower().startswith("bearer "):
        token = token[7:].strip()
    return token or None


def _read_token_file():
    if not config.TOKEN_FILE.exists():
        return None
    try:
        raw = config.TOKEN_FILE.read_text(encoding="utf-8-sig")
    except Exception:
        return None
    return _normalize_token(raw)


def _get_edm_token():
    file_token = _read_token_file()
    if file_token:
        return file_token
    env_token = _normalize_token(config.EDM_TOKEN)
    if env_token and env_token != "paste_your_token_here":
        return env_token
    return None


def _edm_headers(token):
    return {
        "Authorization": "Bearer " + token,
        "Content-Type": "application/json",
        "Accept": "application/json, text/plain, */*",
        "Origin": config.EDM_PORTAL_ORIGIN,
        "Referer": config.EDM_PORTAL_REFERER,
    }


def _read_edm_exists_cache_file():
    try:
        if not EDM_EXISTS_CACHE_PATH.exists():
            return {}
        data = json.loads(EDM_EXISTS_CACHE_PATH.read_text(encoding="utf-8"))
        return data if isinstance(data, dict) else {}
    except Exception:
        return {}


def _write_edm_exists_cache_file(cache):
    try:
        EDM_EXISTS_CACHE_PATH.parent.mkdir(parents=True, exist_ok=True)
        tmp = EDM_EXISTS_CACHE_PATH.with_name(EDM_EXISTS_CACHE_PATH.name + ".tmp")
        tmp.write_text(json.dumps(cache, indent=2, sort_keys=True), encoding="utf-8")
        tmp.replace(EDM_EXISTS_CACHE_PATH)
    except Exception as e:
        log(f"[EDM-AWB-FALLBACK] Warning: could not update cache file: {e}")


def _reset_edm_exists_cache():
    _write_edm_exists_cache_file({})
    log("[EDM-AWB-FALLBACK] Reset shared EDM existence cache for this hotfolder session")


def _get_cached_edm_exists(awb):
    cache = _read_edm_exists_cache_file()
    entry = cache.get(awb)
    if isinstance(entry, dict):
        exists = entry.get("exists")
    else:
        exists = entry
    if isinstance(exists, bool):
        return exists
    return None


def _set_cached_edm_exists(awb, exists):
    cache = _read_edm_exists_cache_file()
    cache[awb] = {
        "exists": bool(exists),
        "checked_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    _write_edm_exists_cache_file(cache)


def edm_awb_exists_fallback(awb):
    cached = _get_cached_edm_exists(awb)
    if cached is not None:
        log(f"[EDM-AWB-FALLBACK] Cache hit for {awb}: exists={cached}")
        return cached
    token = _get_edm_token()
    if not token:
        log(f"[EDM-AWB-FALLBACK] No EDM token available for {awb}; skipping fallback")
        return None
    payload = {
        "documentClass": "SHIPMENT",
        "group": [{"operatingCompany": EDM_OPERATING_COMPANY, "trackingNumber": [awb]}],
        "responseTypes": ["metadata"],
    }
    params = {"pageSize": 25, "continuationToken": "", "archiveSelection": "false"}
    try:
        r = requests.post(
            EDM_METADATA_URL,
            headers=_edm_headers(token),
            params=params,
            json=payload,
            timeout=15,
        )
        if r.status_code == 401:
            log(f"[EDM-AWB-FALLBACK] Token expired while checking {awb}")
            try:
                flag_path = LOG_DIR / "EDM_TOKEN_EXPIRED.flag"
                flag_path.write_text(datetime.now().strftime("%Y-%m-%d %H:%M:%S"), encoding="utf-8")
            except Exception:
                pass
            return None
        if r.status_code == 404:
            _set_cached_edm_exists(awb, False)
            return False
        if r.status_code != 200:
            log(f"[EDM-AWB-FALLBACK] Unexpected status {r.status_code} for {awb}")
            return None
        doc_ids = []
        for group in r.json().get("groups", []):
            for doc in group.get("documents", []):
                doc_id = doc.get("documentId") or doc.get("id")
                if doc_id:
                    doc_ids.append(doc_id)
        exists = bool(doc_ids)
        _set_cached_edm_exists(awb, exists)
        log(f"[EDM-AWB-FALLBACK] EDM existence check for {awb}: exists={exists}")
        return exists
    except requests.exceptions.Timeout:
        log(f"[EDM-AWB-FALLBACK] Timeout checking {awb}")
        return None
    except Exception as e:
        log(f"[EDM-AWB-FALLBACK] Error checking {awb}: {e}")
        return None


# =============================================================================
# AWB LOGS / STAGE CACHE
# =============================================================================
_AWB_LOGS_HEADERS = ["AWB", "SourceFile", "Timestamp", "MatchMethod", "Status"]


def append_to_awb_logs_excel(awb, source_file, match_method, status="MATCHED"):
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    row = [awb, os.path.basename(source_file), ts, match_method, status]
    for attempt in range(5):
        try:
            if AWB_LOGS_PATH.exists():
                wb = load_workbook(AWB_LOGS_PATH)
                ws = wb.active
                if ws.max_row == 0 or ws.cell(1, 1).value != "AWB":
                    ws.insert_rows(1)
                    for col, h in enumerate(_AWB_LOGS_HEADERS, start=1):
                        ws.cell(1, col).value = h
            else:
                AWB_LOGS_PATH.parent.mkdir(parents=True, exist_ok=True)
                wb = Workbook()
                ws = wb.active
                ws.title = "AWB Logs"
                ws.append(_AWB_LOGS_HEADERS)
            ws.append(row)
            wb.save(AWB_LOGS_PATH)
            return
        except PermissionError:
            time.sleep(0.4 * (attempt + 1))
        except Exception as e:
            log(f"[AWB_LOGS] Warning: could not write AWB_Logs.xlsx: {e}")
            return
    log(f"[AWB_LOGS] AWB_Logs.xlsx still locked after retries -- skipping log for {awb}.")


def append_stage_cache_row(input_file, processed_file, awb, detection_type, awb_extraction_secs):
    headers = [
        "Timestamp", "InputFileName", "ProcessedFileName",
        "AWB_Detected", "AWB_Detection_Type", "AWB_Extraction_Seconds",
    ]
    row = [
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        os.path.basename(input_file),
        os.path.basename(processed_file),
        awb, detection_type, awb_extraction_secs,
    ]
    try:
        STAGE_CACHE_CSV.parent.mkdir(parents=True, exist_ok=True)
        new_file = not STAGE_CACHE_CSV.exists()
        with open(STAGE_CACHE_CSV, "a", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            if new_file:
                w.writerow(headers)
            w.writerow(row)
    except Exception as e:
        log(f"[STAGE_CACHE] Warning: could not write stage cache row: {e}")


# =============================================================================
# EXCEL AWB LOADER
# =============================================================================
def extract_12_digit_numbers_from_any_text(s):
    if s is None:
        return []
    s = str(s)
    out = set()
    for m in re.finditer(r"\b\d{12}\b", s):
        out.add(m.group(0))
    for m in re.finditer(r"(\d[\d\-\s]{10,30}\d)", s):
        d = re.sub(r"\D", "", m.group(0))
        if len(d) == AWB_LEN:
            out.add(d)
    return list(out)


def load_awb_set_from_excel(xlsx_path):
    if not Path(xlsx_path).exists():
        raise FileNotFoundError(f"Excel file not found: {xlsx_path}")
    wb = load_workbook(xlsx_path, data_only=True)
    awbs = set()
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                for n in extract_12_digit_numbers_from_any_text(cell):
                    if len(n) == AWB_LEN and n.isdigit():
                        awbs.add(n)
    return awbs


def build_buckets(awb_set):
    by_prefix, by_suffix = {}, {}
    for a in awb_set:
        by_prefix.setdefault(a[:4], []).append(a)
        by_suffix.setdefault(a[-4:], []).append(a)
    return by_prefix, by_suffix


# =============================================================================
# CANDIDATE FILTERS
# =============================================================================
def _strict_awb_from_fragment(text):
    frag = (text or "").strip()
    if re.fullmatch(r"\d{12}", frag):
        return frag
    if re.fullmatch(r"\d{4}[\s\-]\d{4}[\s\-]\d{4}", frag):
        return re.sub(r"\D", "", frag)
    return None


def _is_likely_date_reference(candidate):
    if not candidate or len(candidate) != AWB_LEN or not candidate.isdigit():
        return False
    try:
        year, month = int(candidate[:4]), int(candidate[4:6])
        return 2015 <= year <= 2035 and 1 <= month <= 12
    except Exception:
        return False


def _is_disqualified_candidate(candidate, for_tolerance=False):
    """
    Hard disqualifier.
    for_tolerance=True relaxes the leading-zero rule so OCR near-misses
    like 099617498819 (true AWB 399617498819) can reach tolerance matching.
    Exact matching, clean gate and EDM always pass for_tolerance=False.
    """
    if not candidate or len(candidate) != AWB_LEN or not candidate.isdigit():
        return True
    if not for_tolerance and candidate.startswith("0"):
        return True
    return False


# =============================================================================
# CANDIDATE EXTRACTION
# =============================================================================
_OCR_DIGIT_CHAR_MAP = {
    "O": "0", "Q": "0", "D": "0",
    "I": "1", "L": "1",
    "Z": "2", "S": "5", "G": "6", "B": "8", "T": "7",
}


def _norm_digits_12(raw):
    if not raw:
        return None
    cleaned = re.sub(r"[\s\-:/._]+", "", str(raw).upper())
    if len(cleaned) != AWB_LEN:
        return None
    raw_digit_count = sum(1 for ch in cleaned if ch.isdigit())
    if raw_digit_count < 8:
        return None
    out = []
    for ch in cleaned:
        if ch.isdigit():
            out.append(ch)
            continue
        mapped = _OCR_DIGIT_CHAR_MAP.get(ch)
        if not mapped:
            return None
        out.append(mapped)
    return "".join(out)


def extract_candidates_from_alnum_ocr(text):
    out = set()
    if not text:
        return out
    for m in re.finditer(
        r"(?<![A-Z0-9])([A-Z0-9][A-Z0-9\-\s:/._]{8,30}[A-Z0-9])(?![A-Z0-9])",
        text.upper(),
    ):
        d = _norm_digits_12(m.group(1))
        if d:
            out.add(d)
    return out


def extract_awb_candidates_from_aci_pattern(text):
    out = set()
    if not text:
        return out
    for m in _ACI_AWB_PATTERN.finditer(text):
        d = _norm_digits_12(m.group(1))
        if d:
            out.add(d)
    for m in _ACI_LABELED_PATTERN.finditer(text):
        d = _norm_digits_12(m.group(1))
        if d:
            out.add(d)
    return out


def extract_awb_from_fedex_carrier_row(text):
    out = set()
    if not text:
        return out
    for m in _FEDEX_CARRIER_ROW_PATTERN.finditer(text):
        d = _norm_digits_12(m.group(1))
        if d:
            out.add(d)
    lines = (text or "").splitlines()
    for i, line in enumerate(lines):
        lu = re.sub(r"\s+", "", (line or "").upper())
        if "FEDEX" not in lu and "FED-EX" not in (line or "").upper():
            continue
        block = " ".join(lines[max(0, i - 1): min(len(lines), i + 3)])
        for m in re.finditer(r"(?<!\d)(\d{12})(?!\d)", block):
            out.add(m.group(1))
        for m in re.finditer(r"(?<!\d)(\d{4}[\s\-]\d{4}[\s\-]\d{4})(?!\d)", block):
            d = re.sub(r"\D", "", m.group(1))
            if len(d) == AWB_LEN:
                out.add(d)
    return out


def extract_awb_from_airway_bill_label(text):
    out = set()
    if not text:
        return out
    for m in _AIRWAY_BILL_LABEL_PATTERN.finditer(text):
        d = _norm_digits_12(m.group(1))
        if d:
            out.add(d)
    lines = (text or "").splitlines()
    for i, line in enumerate(lines):
        lu = (line or "").upper()
        lu_norm = re.sub(r"[^A-Z0-9]+", "", lu)
        if not (("AIRWAY" in lu_norm and "BILL" in lu_norm) or "AWAYBILL" in lu_norm):
            continue
        block = " ".join(lines[max(0, i - 1): min(len(lines), i + 3)])
        for m in re.finditer(r"(?<!\d)(\d{12})(?!\d)", block):
            out.add(m.group(1))
        for m in re.finditer(r"(?<!\d)(\d{4}[\s\-]\d{4}[\s\-]\d{4})(?!\d)", block):
            d = re.sub(r"\D", "", m.group(1))
            if len(d) == AWB_LEN:
                out.add(d)
    return out


def extract_candidates_from_text(s):
    s = s or ""
    out = set()
    out.update(extract_candidates_from_alnum_ocr(s))
    for m in re.finditer(r"(?<!\d)(\d{12})(?!\d)", s):
        out.add(m.group(1))
    for m in re.finditer(r"(?<!\d)(\d{4}[\s\-]\d{4}[\s\-]\d{4})(?!\d)", s):
        d = re.sub(r"\D", "", m.group(1))
        if len(d) == AWB_LEN:
            out.add(d)
    for m in re.finditer(r"(?<!\d)400[\s\-:]{0,6}([0-9][0-9\-\s]{10,20})(?!\d)", s, re.IGNORECASE):
        strict = _strict_awb_from_fragment(m.group(1))
        if strict:
            out.add(strict)
    for m in re.finditer(
        r"(?<!\d)(?:A\s*[CGE6]\s*[I1L])[\D]{0,15}([0-9][0-9\-\s]{10,22})(?!\d)",
        s, re.IGNORECASE,
    ):
        d = _norm_digits_12(m.group(1))
        if d:
            out.add(d)
    out.update(extract_awb_candidates_from_aci_pattern(s))
    return out


def extract_db_backed_candidates_from_text(s, awb_set):
    s = s or ""
    out = set()
    for m in re.finditer(r"(?<!\d)(\d[\d\-\s]{10,40}\d)(?!\d)", s):
        digits = re.sub(r"\D", "", m.group(1))
        if len(digits) < AWB_LEN:
            continue
        if len(digits) == AWB_LEN:
            if digits in awb_set:
                out.add(digits)
            continue
        for i in range(len(digits) - AWB_LEN + 1):
            cand = digits[i:i + AWB_LEN]
            if cand in awb_set:
                out.add(cand)
    for m in re.finditer(
        r"(?<![A-Z0-9])([A-Z0-9][A-Z0-9\-\s:/._]{8,36}[A-Z0-9])(?![A-Z0-9])",
        s.upper(),
    ):
        raw = m.group(1)
        norm_chars, raw_digit_count, invalid = [], 0, False
        for ch in raw:
            if ch in " -:/._\t\r\n":
                continue
            if ch.isdigit():
                raw_digit_count += 1
                norm_chars.append(ch)
                continue
            mapped = _OCR_DIGIT_CHAR_MAP.get(ch)
            if mapped:
                norm_chars.append(mapped)
            else:
                invalid = True
                break
        if invalid:
            continue
        norm = "".join(norm_chars)
        if len(norm) < AWB_LEN or raw_digit_count < 8:
            continue
        if len(norm) == AWB_LEN:
            if norm in awb_set:
                out.add(norm)
            continue
        for i in range(len(norm) - AWB_LEN + 1):
            cand = norm[i:i + AWB_LEN]
            if cand in awb_set:
                out.add(cand)
    return out


def extract_candidates_near_keywords(s, line_lookahead=3, line_lookback=1):
    """
    line_lookahead / line_lookback control how many lines around a keyword
    line are scanned.  Text-layer calls use wider windows (5/2);
    OCR calls use the default (3/1) to avoid noise hits.
    """
    s = s or ""
    su = s.upper()
    out = set()

    pattern = re.compile(r"(?<!\d)(\d{12}|\d{4}[\s\-]\d{4}[\s\-]\d{4})(?!\d)")

    def _norm_kw(text):
        return re.sub(r"[^A-Z0-9]+", "", (text or "").upper())

    keywords_norm = {_norm_kw(k) for k in AWB_CONTEXT_KEYWORDS if k}

    def _has_awb_label(window_text):
        if any(k in window_text for k in AWB_CONTEXT_KEYWORDS):
            return True
        window_norm = _norm_kw(window_text)
        if any(kn and kn in window_norm for kn in keywords_norm):
            return True
        relaxed = (
            r"AIR\w{0,6}WAY\w{0,6}BIL{1,2}",
            r"WAY\w{0,6}BIL{1,2}",
            r"TRACK\w{0,10}(NO|NUM|NUMBER)?",
            r"\bAWB\b",
            r"BIL{1,2}\w{0,8}(NO|NUM|NUMBER)",
        )
        return any(re.search(p, window_text, flags=re.IGNORECASE) for p in relaxed)

    for m in pattern.finditer(s):
        d = _strict_awb_from_fragment(m.group(1))
        start = max(0, m.start() - CONTEXT_WINDOW_CHARS)
        end = min(len(su), m.end() + CONTEXT_WINDOW_CHARS)
        window = su[start:end]
        if _has_awb_label(window) and d and len(d) == AWB_LEN:
            out.add(d)

    lines = s.splitlines()
    for i, line in enumerate(lines):
        line_u = line.upper()
        line_norm = _norm_kw(line_u)
        has_label = _has_awb_label(line_u) or any(kn and kn in line_norm for kn in keywords_norm)
        if not has_label:
            continue
        block = " ".join(lines[max(0, i - line_lookback): min(len(lines), i + line_lookahead + 1)])
        for m in pattern.finditer(block):
            d = _strict_awb_from_fragment(m.group(1))
            if d and len(d) == AWB_LEN:
                out.add(d)
    return out


def extract_candidates_from_ocr_data(img):
    out = set()
    try:
        data = pytesseract.image_to_data(img, output_type=Output.DICT, config="--oem 3 --psm 6")
    except Exception:
        return out
    texts = data.get("text", []) or []
    tops  = data.get("top", []) or []
    lefts = data.get("left", []) or []
    img_w = int(getattr(img, "width", 0) or 0)
    img_h = int(getattr(img, "height", 0) or 0)
    y_same_line = max(40, int(img_h * 0.015)) if img_h else 40
    y_below     = max(120, int(img_h * 0.06)) if img_h else 120
    x_span      = max(1400, int(img_w * 0.98)) if img_w else 1400

    def _norm(txt):
        return re.sub(r"[^A-Z0-9]+", "", (txt or "").upper())

    def _num_norm(txt):
        return re.sub(r"[^0-9]", "", (txt or ""))

    label_idx = []
    for i, raw in enumerate(texts):
        n = _norm(raw)
        if not n:
            continue
        if any(kw in n for kw in ("AWB", "AIRWAY", "AIRWAYBILL", "WAYBILL", "TRACK", "NUMBER", "FEDEX", "SHIP")):
            label_idx.append(i)
    if not label_idx:
        return out
    n_tokens = len(texts)
    for i in label_idx:
        try:
            y0, x0 = int(tops[i]), int(lefts[i])
        except Exception:
            continue
        block_tokens = []
        for j in range(n_tokens):
            try:
                y, x = int(tops[j]), int(lefts[j])
            except Exception:
                continue
            if abs(y - y0) <= y_same_line or (0 <= (y - y0) <= y_below):
                if abs(x - x0) <= x_span:
                    t = (texts[j] or "").strip()
                    if t:
                        block_tokens.append(t)
        block = " ".join(block_tokens)
        out.update(extract_candidates_from_text(block))
        out.update(extract_candidates_near_keywords(block))
        for tok in block_tokens:
            d = _num_norm(tok)
            if len(d) == AWB_LEN:
                out.add(d)
            d2 = _norm_digits_12(tok)
            if d2:
                out.add(d2)
    return out


# =============================================================================
# OCR HELPERS
# =============================================================================
def render_page(pdf_path, dpi_value):
    doc = fitz.open(pdf_path)
    try:
        page = doc.load_page(0)
        zoom = dpi_value / 72.0
        pix = page.get_pixmap(matrix=fitz.Matrix(zoom, zoom), alpha=False)
        return Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
    finally:
        doc.close()


def render_page_from_page(page, dpi_value):
    zoom = dpi_value / 72.0
    pix = page.get_pixmap(matrix=fitz.Matrix(zoom, zoom), alpha=False)
    return Image.frombytes("RGB", [pix.width, pix.height], pix.samples)


def preprocess(img, thr=175, invert=False):
    img = img.convert("L")
    img = ImageOps.autocontrast(img)
    if invert:
        img = ImageOps.invert(img)
    return img.point(lambda p: 255 if p > thr else 0)


def preprocess_for_text(img, invert=False):
    img = img.convert("L")
    img = ImageOps.autocontrast(img)
    if invert:
        img = ImageOps.invert(img)
    return img


def ocr_digits_only(img, psm=6):
    cfg = (
        f"--oem 3 --psm {psm} "
        "-c tessedit_char_whitelist=0123456789 "
        "-c preserve_interword_spaces=1 "
    )
    return pytesseract.image_to_string(img, config=cfg)


def ocr_text_general(img, psm=6):
    return pytesseract.image_to_string(img, config=f"--oem 3 --psm {psm}")


def digit_score(s):
    if not s:
        return 0
    return sum(1 for ch in s if ch.isdigit())


def _upscale(img, factor):
    try:
        rs = Image.Resampling.LANCZOS
    except AttributeError:
        rs = Image.LANCZOS
    return img.resize((img.width * factor, img.height * factor), resample=rs)


# =============================================================================
# ROTATION PROBE
# =============================================================================
_PROBE_KEYWORDS = (
    "AWB", "AWB NO", "AWB NUMBER", "AIRWAY", "WAYBILL", "AIRWAY BILL NUMBER",
    "AIR WAY BILL", "TRACKING", "TRACKING NO", "TRACKING #",
    "FDX", "FDE", "FDXE", "FEDEX", "FED-EX", "FDX TRACKING", "FDXE TRACKING",
    "FEDEX TRACKING", "AIRWAY BILL", "BILL NUMBER", "BILL NO",
    "HAWB", "MAWB", "ACI", "CARGO CONTROL NUMBER", "CCN",
    "COMMERCIAL INVOICE", "SHIPMENT", "SHIPPER", "CONSIGNEE", "TRK", "TRK#",
)
ALLOWED_ROTATION_ANGLES = (0, 90, 180, 270)


def rotation_probe_best(img_lowdpi, return_scores=False, preferred_angles=None):
    """
    Raw-rotate-first probe with keyword scoring.
    Returns (best_rot, scores_dict, probe_texts_dict) when return_scores=True.
    probe_texts_dict: rot -> (digit_text, general_text) for reuse.

    preferred_angles: optional subset of ALLOWED_ROTATION_ANGLES to probe.
    Use only when a strong external hint already narrows the likely angle —
    specifically for image-only documents in long-pass where metadata or
    aspect ratio strongly indicated rotation.  Missing angles are filled with
    score=0 so downstream logic stays consistent.
    """
    angles = tuple(preferred_angles) if preferred_angles else ALLOWED_ROTATION_ANGLES
    digit_scores = {}
    probe_texts  = {}

    for rot in angles:
        rimg = img_lowdpi.rotate(rot, expand=True) if rot else img_lowdpi
        t_digits = ocr_digits_only(preprocess(rimg, thr=175, invert=False), psm=6)
        digit_scores[rot] = digit_score(t_digits)
        probe_texts[rot]  = (t_digits, "")  # general text filled lazily below

    # Fill scores for any angles not probed with 0 so downstream logic stays consistent
    for _fill_rot in ALLOWED_ROTATION_ANGLES:
        if _fill_rot not in digit_scores:
            digit_scores[_fill_rot] = 0
            probe_texts[_fill_rot]  = ("", "")

    ranked = sorted(digit_scores.items(), key=lambda x: x[1], reverse=True)
    best_digit_rot, best_digit_sc = ranked[0]
    second_digit_sc = ranked[1][1] if len(ranked) > 1 else -1

    # Fast path: 0° clearly wins on digits → skip expensive text OCR
    # Only valid when 0° was actually probed
    if (
        0 in angles
        and best_digit_rot == 0
        and (best_digit_sc - second_digit_sc) >= ROTATION_PROBE_DIGIT_CLEAR_MARGIN
    ):
        if return_scores:
            return 0, {k: float(v) for k, v in digit_scores.items()}, probe_texts
        return 0

    scores = {}
    for rot in angles:
        rimg = img_lowdpi.rotate(rot, expand=True) if rot else img_lowdpi
        t_text = ocr_text_general(preprocess_for_text(rimg, invert=False), psm=6)
        tu = (t_text or "").upper()
        probe_texts[rot] = (probe_texts[rot][0], t_text)

        kw_hits = sum(1 for kw in _PROBE_KEYWORDS if kw in tu)
        coherent = sum(1 for w in re.findall(r"[A-Za-z]{4,}", t_text or "") if w.isalpha())
        scores[rot] = digit_scores[rot] + (kw_hits * 120) + (coherent * 2)

    # Fill scores for unprobed angles so downstream angle ordering is not broken
    for _fill_rot in ALLOWED_ROTATION_ANGLES:
        if _fill_rot not in scores:
            scores[_fill_rot] = 0

    best_rot = max(scores, key=lambda r: scores[r])
    if best_rot != 0 and (scores[best_rot] - scores.get(0, 0)) < ROTATION_PROBE_MIN_FLIP_MARGIN:
        best_rot = 0
    if best_rot not in ALLOWED_ROTATION_ANGLES:
        best_rot = 0

    if return_scores:
        return best_rot, {k: float(v) for k, v in scores.items()}, probe_texts
    return best_rot


# =============================================================================
# TIERED CANDIDATE EXTRACTION
# =============================================================================
def _candidates_from_label_prefixes(text):
    high = set()
    if not text:
        return high
    for m in _AWB_LABEL_PATTERN.finditer(text):
        d = _norm_digits_12(m.group(2))
        if d:
            high.add(d)
    # Tight 400 prefix (same pattern as the no-DB-check exemption but here
    # the result enters prioritize_db_match which checks DB before accepting)
    for m in re.finditer(r"(?<!\d)400(?:[\s\-]{0,2})(\d(?:[\s\-]?\d){11})(?!\d)", text, re.IGNORECASE):
        d = _norm_digits_12(m.group(1))
        if d:
            high.add(d)
    # Labeled 400 (400 NUMBER:, 400 NO:) — DB-checked path only
    for m in _400_LABELED_PATTERN.finditer(text):
        d = _norm_digits_12(m.group(1))
        if d:
            high.add(d)
    for m in re.finditer(
        r"(?<!\d)(?:A\s*[CGE6]\s*[I1L])(?:\D{0,15})(\d[\d\-\s]{10,24})(?!\d)", text, re.IGNORECASE
    ):
        d = _norm_digits_12(m.group(1))
        if d:
            high.add(d)
    for m in _TRACK_LABEL_PATTERN.finditer(text):
        d = _norm_digits_12(m.group(1))
        if d:
            high.add(d)
    high.update(extract_awb_from_airway_bill_label(text))
    return high


def _promote_keyword_adjacent(text, candidates):
    s = text or ""
    su = s.upper()
    cands = {c for c in (candidates or set()) if isinstance(c, str) and len(c) == AWB_LEN and c.isdigit()}
    promoted = set()
    if not cands:
        return promoted

    def _norm_kw(t):
        return re.sub(r"[^A-Z0-9]+", "", (t or "").upper())

    keywords_norm = {_norm_kw(k) for k in AWB_CONTEXT_KEYWORDS if k}

    def _has_kw(t):
        if any(k in t for k in AWB_CONTEXT_KEYWORDS):
            return True
        tn = _norm_kw(t)
        return any(kn and kn in tn for kn in keywords_norm)

    for c in cands:
        for m in re.finditer(rf"(?<!\d){re.escape(c)}(?!\d)", s):
            start = max(0, m.start() - CONTEXT_WINDOW_CHARS)
            end = min(len(su), m.end() + CONTEXT_WINDOW_CHARS)
            if _has_kw(su[start:end]):
                promoted.add(c)
                break

    lines = s.splitlines()
    line_cands = []
    for line in lines:
        ln = re.sub(r"\D", " ", line)
        found = {tok for tok in ln.split() if len(tok) == AWB_LEN and tok.isdigit() and tok in cands}
        line_cands.append(found)
    for i, line in enumerate(lines):
        if not _has_kw(line.upper()):
            continue
        for j in [i - 1, i, i + 1]:
            if 0 <= j < len(lines):
                promoted.update(line_cands[j])
    return promoted


def extract_tiered_candidates(text, awb_set):
    s = text or ""
    high, standard = set(), set()
    high.update(_candidates_from_label_prefixes(s))
    high.update(extract_awb_candidates_from_aci_pattern(s))
    high.update(extract_awb_from_fedex_carrier_row(s))
    standard.update(extract_candidates_from_text(s))
    standard.update(extract_db_backed_candidates_from_text(s, awb_set))
    promoted = _promote_keyword_adjacent(s, standard)
    high.update(promoted)
    standard.difference_update(high)
    disq = {c for c in (high | standard) if _is_disqualified_candidate(c)}
    high.difference_update(disq)
    standard.difference_update(disq)
    date_refs = {c for c in standard if _is_likely_date_reference(c)}
    standard.difference_update(date_refs)
    return high, standard


def extract_clean_candidates(text):
    s = text or ""
    out = set()
    for m in re.finditer(r"(?<!\d)(\d{12})(?!\d)", s):
        out.add(m.group(1))
    for m in re.finditer(r"(?<!\d)(\d{4}[\s\-]\d{4}[\s\-]\d{4})(?!\d)", s):
        d = re.sub(r"\D", "", m.group(1))
        if len(d) == AWB_LEN:
            out.add(d)
    return out


# =============================================================================
# MATCHING
# =============================================================================
def hamming(a, b):
    return sum(1 for x, y in zip(a, b) if x != y)


def _unique_awb_candidate_count(candidates):
    return len({c for c in (candidates or set()) if isinstance(c, str) and len(c) == AWB_LEN and c.isdigit()})


def pick_unique_close_match(candidate, awb_set, by_prefix, by_suffix, max_distance=2):
    pool = set()
    pool.update(by_prefix.get(candidate[:4], []))
    pool.update(by_suffix.get(candidate[-4:], []))
    if not pool:
        pool = awb_set
    scored = [(a, hamming(candidate, a)) for a in pool if hamming(candidate, a) <= max_distance]
    if not scored:
        return None
    scored.sort(key=lambda x: x[1])
    best_awb, best_d = scored[0]
    if len([a for a, d in scored if d == best_d]) != 1:
        return None
    return best_awb


def tolerance_match_with_tie_guard(candidates, awb_set, by_prefix, by_suffix, max_distance=2):
    best_distance, best_awbs = None, set()
    for c in candidates:
        if len(c) != AWB_LEN or not c.isdigit():
            continue
        pool = set()
        pool.update(by_prefix.get(c[:4], []))
        pool.update(by_suffix.get(c[-4:], []))
        if not pool:
            pool = awb_set
        for a in pool:
            d = hamming(c, a)
            if d > max_distance:
                continue
            if best_distance is None or d < best_distance:
                best_distance = d
                best_awbs = {a}
            elif d == best_distance:
                best_awbs.add(a)
    if best_distance is None:
        return None, None
    if len(best_awbs) == 1:
        return next(iter(best_awbs)), None
    return None, sorted(best_awbs)


def tolerance_match_with_details(candidates, awb_set, by_prefix, by_suffix, max_distance=2):
    best_distance, best_awbs, evidence = None, set(), {}
    for c in (candidates or set()):
        if len(c) != AWB_LEN or not c.isdigit():
            continue
        pool = set()
        pool.update(by_prefix.get(c[:4], []))
        pool.update(by_suffix.get(c[-4:], []))
        if not pool:
            pool = awb_set
        for a in pool:
            d = hamming(c, a)
            if d > max_distance:
                continue
            if best_distance is None or d < best_distance:
                best_distance = d
                best_awbs = {a}
                evidence = {a: {c}}
            elif d == best_distance:
                best_awbs.add(a)
                evidence.setdefault(a, set()).add(c)
    if best_distance is None:
        return {"status": "none"}
    if len(best_awbs) == 1:
        awb = next(iter(best_awbs))
        return {"status": "matched", "awb": awb, "distance": best_distance,
                "evidence_candidates": evidence.get(awb, set())}
    return {"status": "tie", "distance": best_distance, "ties": sorted(best_awbs)}


def _max_stage_hits_for_evidence(evidence_candidates, candidate_stage_hits):
    if not evidence_candidates or not isinstance(candidate_stage_hits, dict):
        return 0
    return max(len(candidate_stage_hits.get(c, set())) for c in evidence_candidates)


def prioritize_db_match(
    high_set, standard_set, awb_set, by_prefix, by_suffix,
    include_tolerance=True, candidate_stage_hits=None,
):
    exact_high = sorted((high_set or set()) & awb_set)
    if len(exact_high) == 1:
        return {"status": "matched", "awb": exact_high[0], "method": "Exact-High"}
    if len(exact_high) > 1:
        return {"status": "tie", "ties": exact_high, "method": "Exact-High"}

    exact_std = sorted((standard_set or set()) & awb_set)
    if len(exact_std) == 1:
        return {"status": "matched", "awb": exact_std[0], "method": "Exact-Standard"}
    if len(exact_std) > 1:
        return {"status": "tie", "ties": exact_std, "method": "Exact-Standard"}

    if include_tolerance:
        # Build tolerance pools with leading-zero rule relaxed
        tol_high_pool = {
            c for c in (high_set or set())
            if not _is_disqualified_candidate(c, for_tolerance=True)
        }
        tol_high = tolerance_match_with_details(
            tol_high_pool, awb_set, by_prefix, by_suffix,
            max_distance=TOLERANCE_HIGH_MAX_DISTANCE,
        )
        if tol_high["status"] == "matched":
            dist = tol_high.get("distance", 99)
            stage_hits = _max_stage_hits_for_evidence(
                tol_high.get("evidence_candidates", set()), candidate_stage_hits or {}
            )
            required = MIN_STAGE_HITS_HIGH_TOL1 if dist <= 1 else MIN_STAGE_HITS_HIGH_TOL2
            if stage_hits >= required:
                return {"status": "matched", "awb": tol_high["awb"],
                        "method": "Tolerance2-High", "distance": dist, "stage_hits": stage_hits}
        if tol_high["status"] == "tie":
            return {"status": "tie", "ties": tol_high.get("ties", []), "method": "Tolerance2-High"}

        if ALLOW_STANDARD_TOLERANCE:
            tol_std_pool = {
                c for c in (standard_set or set())
                if not _is_disqualified_candidate(c, for_tolerance=True)
            }
            tol_std = tolerance_match_with_details(
                tol_std_pool, awb_set, by_prefix, by_suffix,
                max_distance=TOLERANCE_STANDARD_MAX_DISTANCE,
            )
            if tol_std["status"] == "matched":
                stage_hits = _max_stage_hits_for_evidence(
                    tol_std.get("evidence_candidates", set()), candidate_stage_hits or {}
                )
                std_count_ok = (
                    (len(standard_set or set()) == 1)
                    if REQUIRE_SINGLE_STANDARD_CANDIDATE_FOR_TOL
                    else True
                )
                if stage_hits >= MIN_STAGE_HITS_STANDARD_TOL and std_count_ok:
                    return {"status": "matched", "awb": tol_std["awb"],
                            "method": "Tolerance2-Standard",
                            "distance": tol_std.get("distance", 99),
                            "stage_hits": stage_hits}
            if tol_std["status"] == "tie":
                return {"status": "tie", "ties": tol_std.get("ties", []), "method": "Tolerance2-Standard"}

    return {"status": "none"}


def decide_from_candidates(candidates, awb_set, by_prefix, by_suffix, allow_tolerance):
    exact = sorted(candidates & awb_set)
    if len(exact) == 1:
        return exact[0], exact
    if len(exact) > 1:
        return None, exact
    if not allow_tolerance or not ALLOW_1_DIGIT_TOLERANCE:
        return None, []
    close = set()
    for c in candidates:
        if len(c) == AWB_LEN and c.isdigit():
            cm = pick_unique_close_match(c, awb_set, by_prefix, by_suffix)
            if cm:
                close.add(cm)
    close = sorted(close)
    if len(close) == 1:
        return close[0], close
    return None, close


# =============================================================================
# TABLE LINE REMOVAL
# =============================================================================
def remove_table_lines_image(img):
    if not CV2_AVAILABLE:
        return None
    try:
        gray = cv2.cvtColor(np.array(img.convert("RGB")), cv2.COLOR_RGB2GRAY)
        bw = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)[1]
        hk = cv2.getStructuringElement(cv2.MORPH_RECT, (60, 1))
        vk = cv2.getStructuringElement(cv2.MORPH_RECT, (1, 60))
        lines = cv2.bitwise_or(cv2.morphologyEx(bw, cv2.MORPH_OPEN, hk),
                               cv2.morphologyEx(bw, cv2.MORPH_OPEN, vk))
        cleaned = cv2.bitwise_not(cv2.bitwise_and(bw, cv2.bitwise_not(lines)))
        return Image.fromarray(cleaned).convert("RGB")
    except Exception:
        return None


# =============================================================================
# MAIN PROCESSOR
# =============================================================================
def process_pdf(pdf_path, awb_set, by_prefix, by_suffix,
                allow_long_pass=True,
                timeout_seconds=None,
                resume_state=None,
                _state_out=None):
    """
    allow_long_pass=False  → fast lane (Stages 0-3 only, defer after Stage 3 fail)
    allow_long_pass=True   → full pipeline
    timeout_seconds        → abort and defer to third-pass if exceeded (long-pass only)
    resume_state           → dict from a prior timeout; skips Stages 0-3.1
    Returns "MATCHED", "NEEDS_REVIEW", "DEFERRED", or "TIMEOUT_DEFERRED".
    """
    start_ts = time.perf_counter()
    name = os.path.basename(pdf_path)
    all_tried           = set()
    stage_snapshots     = {}
    quarantine          = {}
    running_high        = set()
    running_standard    = set()
    candidate_stage_hits = {}
    candidate_confidence = {}
    image_cache      = {}
    ocr_cache        = {}
    preprocess_cache = {}
    table_clean_cache = [None]

    # Default values for variables set inside early stages.
    # Overwritten by the normal path (if not _proceed_to_route) or by resume restore.
    _is_image_only   = False
    _rotation_hint   = None
    base_angle       = 0
    probe_scores     = {}
    probe_texts      = {}
    _angle_certainty = "UNCERTAIN"

    # Resume from a prior timeout: restore all accumulated state and skip early stages.
    _proceed_to_route = resume_state is not None
    if _proceed_to_route:
        rs = resume_state
        probe_scores     = rs.get('probe_scores', {})
        probe_texts      = rs.get('probe_texts', {})
        base_angle       = rs.get('base_angle', 0)
        _angle_certainty = rs.get('_angle_certainty', 'UNCERTAIN')
        _rotation_hint   = rs.get('_rotation_hint', None)
        _is_image_only   = rs.get('_is_image_only', False)
        running_high.update(rs.get('running_high', []))
        running_standard.update(rs.get('running_standard', []))
        for k, v in rs.get('candidate_stage_hits', {}).items():
            candidate_stage_hits[k] = set(v)
        candidate_confidence.update(rs.get('candidate_confidence', {}))
        all_tried.update(rs.get('all_tried', []))
        for k, v in rs.get('quarantine', {}).items():
            quarantine[k] = set(v)
        ocr_cache.update(rs.get('ocr_cache', {}))
        timings_saved = rs.get('timings', {})
    else:
        timings_saved = {}

    timings = {
        "filename_ms": 0.0, "text_layer_ms": 0.0,
        "ocr_main_ms": 0.0, "ocr_strong_ms": 0.0,
        "ocr_context_ms": 0.0, "rotation_ms": 0.0,
        "total_active_ms": 0.0,
    }
    timings.update(timings_saved)  # carry forward timings from prior pass on resume

    if not file_is_stable(pdf_path):
        return "NEEDS_REVIEW"

    def finalize(status, route, reason, match_method, awb=None):
        timings["total_active_ms"] = round((time.perf_counter() - start_ts) * 1000, 1)
        audit_event("AWB_HOTFOLDER", file=name, awb=awb, status=status, route=route,
                    match_method=match_method, reason=reason, timings_ms=timings)
        log(f"[TIMING] file={name} method={match_method} route={route} "
            f"filename_ms={timings['filename_ms']} text_layer_ms={timings['text_layer_ms']} "
            f"ocr_main_ms={timings['ocr_main_ms']} ocr_strong_ms={timings['ocr_strong_ms']} "
            f"ocr_context_ms={timings['ocr_context_ms']} rotation_ms={timings['rotation_ms']} "
            f"total_active_ms={timings['total_active_ms']}")

    def awb_extract_secs():
        return round(time.perf_counter() - start_ts, 3)

    def snapshot(stage, candidates):
        if not LOG_STAGE_SNAPSHOTS:
            return
        cset = {c for c in (candidates or set()) if isinstance(c, str) and c}
        stage_snapshots[stage] = {"count": len(cset), "sample": sorted(cset)[:CANDIDATE_SNAPSHOT_LIMIT]}

    def log_snapshots():
        if not LOG_STAGE_SNAPSHOTS or not stage_snapshots:
            return
        for stage in sorted(stage_snapshots):
            snap = stage_snapshots[stage]
            log(f"[SNAPSHOT] {stage}: count={snap['count']} sample={snap['sample']}")

    def merge_stage_candidates(high_set, standard_set, stage_name):
        nonlocal running_high, running_standard
        high_set = {
            c for c in (high_set or set())
            if len(c) == AWB_LEN and c.isdigit() and not _is_disqualified_candidate(c)
        }
        standard_set = {
            c for c in (standard_set or set())
            if len(c) == AWB_LEN and c.isdigit()
            and not _is_disqualified_candidate(c)
            and not _is_likely_date_reference(c)
        }
        # Quarantine: single-hit STANDARD candidates from invert or rotation passes
        # that produced a large number of candidates are treated as noisy.
        # They stay out of running_standard until confirmed by a second stage.
        # Label-backed stages (AirwayLabel, ROI) are exempt — they are already
        # context-specific so their candidates should enter the pool directly.
        _is_noisy_source = any(tag in stage_name for tag in (
            "Invert", "AngFallback", "Rotation-180", "Rotation-270",
        ))
        _is_label_backed = "AirwayLabel" in stage_name or "ROI" in stage_name
        if _is_noisy_source and not _is_label_backed and len(standard_set) > 3:
            for c in standard_set:
                if c not in all_tried:
                    quarantine.setdefault(c, set()).add(stage_name)
            all_tried.update(standard_set)
            # Still add HIGH candidates from these stages unconditionally
            running_high.update(high_set)
            for c in high_set:
                candidate_stage_hits.setdefault(c, set()).add(stage_name)
                candidate_confidence[c] = "HIGH"
            return

        all_tried.update(high_set | standard_set)
        running_high.update(high_set)
        running_standard.update(standard_set)
        running_standard.difference_update(running_high)
        for c in high_set:
            candidate_stage_hits.setdefault(c, set()).add(stage_name)
            candidate_confidence[c] = "HIGH"
        for c in standard_set:
            candidate_stage_hits.setdefault(c, set()).add(stage_name)
            candidate_confidence.setdefault(c, "STANDARD")
        # Promote quarantined candidates that now appear in a second stage
        for c in list(quarantine.keys()):
            if c in running_standard or c in running_high:
                del quarantine[c]  # already in pool
            elif c in (high_set | standard_set):
                running_standard.add(c)
                candidate_stage_hits.setdefault(c, set()).update(quarantine.pop(c))
                candidate_stage_hits[c].add(stage_name)

    def _has_quality_candidates():
        """True when candidates are genuinely promising — not just single-pass OCR noise."""
        if running_high:
            return True
        persistent = {c for c in running_standard if len(candidate_stage_hits.get(c, set())) >= 2}
        return bool(persistent)

    def close_pdf():
        nonlocal page_doc, page
        if page_doc is not None:
            try:
                page_doc.close()
            except Exception:
                pass
            page_doc = None
            page = None

    def complete_match(awb, method, reason):
        log(f"AWB MATCHED ({method}): {awb} ({name})")
        close_pdf()
        append_to_awb_logs_excel(awb, pdf_path, match_method=method)
        processed_path = move_to_processed_renamed(pdf_path, awb)
        processed_name = os.path.basename(processed_path)
        append_stage_cache_row(name, processed_name, awb, method, awb_extract_secs())
        record_hotfolder_end(name, awb, processed_name, method)
        finalize("MATCHED", "PROCESSED", reason, method, awb=awb)

    def send_review(reason, method):
        log(f"NO MATCH FOUND -> Needs review: {name}")
        log(f"  Reason: {reason}")
        log(f"  Candidates tried: {sorted(all_tried)}")
        for c in sorted(all_tried):
            stages = sorted(candidate_stage_hits.get(c, set()))
            conf = candidate_confidence.get(c, "STANDARD")
            log(f"  Candidate {c} | conf={conf} | stages={stages}")
        if quarantine:
            qlist = sorted(quarantine.keys())
            log(f"  Quarantined noisy candidates (excluded from matching): {qlist}")
            for c in qlist:
                log(f"  Quarantined {c} | stages={sorted(quarantine[c])}")
        log_snapshots()
        close_pdf()
        safe_move(pdf_path, NEEDS_REVIEW_DIR)
        record_hotfolder_needs_review(name, f"{reason} | cands={sorted(all_tried)}")
        finalize("NEEDS-REVIEW", "NEEDS_REVIEW", reason, method)

    def run_exact_priority():
        return prioritize_db_match(
            running_high, running_standard, awb_set, by_prefix, by_suffix,
            include_tolerance=False, candidate_stage_hits=candidate_stage_hits,
        )

    def run_full_priority():
        return prioritize_db_match(
            running_high, running_standard, awb_set, by_prefix, by_suffix,
            include_tolerance=True, candidate_stage_hits=candidate_stage_hits,
        )

    page_doc = None
    page     = None

    def get_page():
        nonlocal page_doc, page
        if page is None:
            page_doc = fitz.open(pdf_path)
            page = page_doc.load_page(0)
        return page

    def get_image(dpi, rot=0):
        key = (dpi, rot)
        if key in image_cache:
            return image_cache[key]
        base_key = (dpi, 0)
        if base_key not in image_cache:
            image_cache[base_key] = render_page_from_page(get_page(), dpi)
        if rot == 0:
            return image_cache[base_key]
        image_cache[key] = image_cache[base_key].rotate(rot, expand=True)
        return image_cache[key]

    def get_preprocessed(dpi, rot, thr, inv):
        img = get_image(dpi, rot)
        p_key = ((dpi, rot), thr, inv)
        if p_key in preprocess_cache:
            return preprocess_cache[p_key]
        result = preprocess(img, thr=thr, invert=inv)
        preprocess_cache[p_key] = result
        return result

    def get_ocr_digits(dpi, rot, thr, inv, psm):
        img_key = (dpi, rot)
        c_key = (img_key, f"dig_{thr}_{int(inv)}", psm)
        if c_key in ocr_cache:
            return ocr_cache[c_key]
        pre = get_preprocessed(dpi, rot, thr, inv)
        txt = ocr_digits_only(pre, psm=psm)
        ocr_cache[c_key] = txt
        return txt

    def get_ocr_text(dpi, rot, inv, psm):
        img_key = (dpi, rot)
        c_key = (img_key, f"txt_{int(inv)}", psm)
        if c_key in ocr_cache:
            return ocr_cache[c_key]
        img = get_image(dpi, rot)
        pre = preprocess_for_text(img, invert=inv)
        txt = ocr_text_general(pre, psm=psm)
        ocr_cache[c_key] = txt
        return txt

    def run_clean_priority_gate(text, stage_name):
        clean = {
            c for c in extract_clean_candidates(text)
            if len(c) == AWB_LEN and c.isdigit()
            and not _is_disqualified_candidate(c)
            and not _is_likely_date_reference(c)
        }
        if clean:
            snapshot(f"{stage_name}-Clean", clean)
            clean_db = clean & awb_set
            merge_stage_candidates(clean_db, clean - clean_db, f"{stage_name}-Clean")
            if len(clean_db) == 1:
                return {"status": "matched", "awb": next(iter(clean_db)), "method": "Clean-Exact"}
            if len(clean_db) > 1:
                return {"status": "tie", "ties": sorted(clean_db), "method": "Clean-Exact"}
        return {"status": "none", "method": "Clean-Exact"}

    def _check_timeout():
        """Raise _TimeoutDeferred when the long-pass time budget is exceeded.
        Only called at natural angle boundaries — never mid-subpass."""
        if timeout_seconds and (time.perf_counter() - start_ts) > timeout_seconds:
            log(f"[TIMEOUT] {name} exceeded {timeout_seconds:.0f}s budget — "
                f"deferring to third-pass with {len(running_high)} high / "
                f"{len(running_standard)} std candidates accumulated")
            raise _TimeoutDeferred()

    # ─────────────────────────────────────────────────────────────────────────
    # STAGE 0 — FILENAME
    # ─────────────────────────────────────────────────────────────────────────
    log(f"{'[THIRD-PASS] Resuming' if _proceed_to_route else 'Processing'}: {name}")
    if _proceed_to_route:
        log(f"[THIRD-PASS] Restored: base_angle={base_angle}° "
            f"certainty={_angle_certainty} "
            f"high={len(running_high)} std={len(running_standard)} "
            f"ocr_cache_entries={len(ocr_cache)}")
    record_hotfolder_start(name)

    if not _proceed_to_route:
        fn_start = time.perf_counter()
        awb_from_name = extract_awb_from_filename_strict(name)
        timings["filename_ms"] = round((time.perf_counter() - fn_start) * 1000, 1)
        if awb_from_name:
            complete_match(awb_from_name, "Filename", "Matched by strict filename pattern")
            return "MATCHED"
    
        # ─────────────────────────────────────────────────────────────────────────
        # STAGE 1 — TEXT LAYER (+ set_rotation fallback + spatial word sort)
        # ─────────────────────────────────────────────────────────────────────────
        tl_start = time.perf_counter()
        txt_layer = get_page().get_text("text") or ""
    
        # 1a. set_rotation fallback for rotated vector PDFs (0ms, scanned PDFs unaffected)
        if len(txt_layer.strip()) == 0:
            for _hint in [90, 270, 180]:
                try:
                    get_page().set_rotation(_hint)
                    _t = get_page().get_text("text") or ""
                    if len(_t.strip()) > 20:
                        txt_layer = _t
                        log(f"[TEXT-LAYER] Recovered via set_rotation({_hint})")
                        break
                except Exception:
                    pass
            try:
                get_page().set_rotation(0)
            except Exception:
                pass
    
        # 1b. Spatial word sort for scrambled multi-column stream
        if len(txt_layer.strip()) > 20:
            _words = get_page().get_text("words") or []
            if _words:
                _sorted_txt = " ".join(
                    w[4] for w in sorted(_words, key=lambda w: (round(w[1] / 10) * 10, w[0]))
                )
                _h_raw, _s_raw = extract_tiered_candidates(txt_layer, awb_set)
                _h_srt, _s_srt = extract_tiered_candidates(_sorted_txt, awb_set)
                if len(_h_srt | _s_srt) > len(_h_raw | _s_raw):
                    txt_layer = _sorted_txt
                    log("[TEXT-LAYER] Using spatially sorted word order")
    
        timings["text_layer_ms"] = round((time.perf_counter() - tl_start) * 1000, 1)
    
        # 1c. 400-pattern on text layer (no DB check)
        awb_400 = extract_awb_from_400_pattern(txt_layer)
        if awb_400:
            complete_match(awb_400, "TextLayer-400", "Matched via text-layer 400 pattern")
            return "MATCHED"
    
        # 1d. Clean gate then full tiered extraction on text layer
        clean_res = run_clean_priority_gate(txt_layer, "Text-Layer")
        if clean_res["status"] == "matched":
            complete_match(clean_res["awb"], f"Text-Layer-{clean_res['method']}",
                           "Matched exact DB candidate from text layer")
            return "MATCHED"
        if clean_res["status"] == "tie":
            send_review(f"Ambiguous text-layer clean tie: {clean_res.get('ties', [])[:8]}",
                        f"Text-Layer-{clean_res['method']}")
            return "NEEDS_REVIEW"
    
        high1, std1 = extract_tiered_candidates(txt_layer, awb_set)
        merge_stage_candidates(high1, std1, "Text-Layer")
        snapshot("Text-Layer-HIGH", high1)
        snapshot("Text-Layer-STANDARD", std1)
        # Use wider keyword window for clean text layer
        near_kw = extract_candidates_near_keywords(txt_layer, line_lookahead=5, line_lookback=2)
        near_kw_db = near_kw & awb_set
        if near_kw_db:
            merge_stage_candidates(near_kw_db, set(), "Text-Layer-KW")
        res = run_full_priority()
        if res["status"] == "matched":
            complete_match(res["awb"], f"Text-Layer-{res['method']}", "Matched from text-layer candidates")
            return "MATCHED"
        if res["status"] == "tie":
            send_review(f"Ambiguous text-layer priority tie: {res.get('ties', [])[:8]}",
                        f"Text-Layer-{res['method']}")
            return "NEEDS_REVIEW"
    
        # ─────────────────────────────────────────────────────────────────────────
        # PRE-OCR ANGLE DETECTION (0ms checks before any image render)
        # ─────────────────────────────────────────────────────────────────────────
        _is_image_only = len((txt_layer or "").strip()) == 0
        _rotation_hint = None
    
        # Check 1: PDF metadata rotation
        try:
            _page_meta_rot = get_page().rotation
            if _page_meta_rot in (90, 180, 270):
                _rotation_hint = _page_meta_rot
                log(f"[ANGLE-DETECT] PDF metadata rotation={_page_meta_rot}°")
        except Exception:
            pass
    
        # Check 2: Page aspect ratio
        if _rotation_hint is None:
            try:
                rect = get_page().rect
                if (rect.width / max(rect.height, 1)) > 1.3:
                    _rotation_hint = 90
                    log(f"[ANGLE-DETECT] Landscape page ratio ({rect.width:.0f}x{rect.height:.0f}) — likely 90°")
            except Exception:
                pass
    
        # Check 3: Text character spread (only when text layer exists)
        if _rotation_hint is None and len(txt_layer.strip()) > 20:
            try:
                _words_chk = get_page().get_text("words") or []
                if len(_words_chk) > 5:
                    xs = [w[0] for w in _words_chk]
                    ys = [w[1] for w in _words_chk]
                    if (max(ys) - min(ys)) > (max(xs) - min(xs)) * 1.5:
                        _rotation_hint = 90
                        log("[ANGLE-DETECT] Text y-spread >> x-spread — likely 90°/270°")
            except Exception:
                pass
    
        # Check 4: Pixel row variance (~50ms, only if still unknown and cv2 available)
        if _rotation_hint is None and CV2_AVAILABLE:
            try:
                tiny = get_image(60, 0)
                arr = np.array(tiny.convert("L"))
                row_var = float(np.var(arr, axis=1).mean())
                col_var = float(np.var(arr, axis=0).mean())
                if col_var > row_var * 1.4:
                    _rotation_hint = 90
                    log(f"[ANGLE-DETECT] Pixel variance col={col_var:.1f} >> row={row_var:.1f} — likely rotated")
            except Exception:
                pass
    
        # ─────────────────────────────────────────────────────────────────────────
        # STAGE 2 — OCR MAIN at 0°
        # ─────────────────────────────────────────────────────────────────────────
        main_start = time.perf_counter()
        _ocr_angle = 0  # always 0° for Stages 2-3; probe runs later
    
        for _psm in OCR_MAIN_PSMS:
            txt_m = get_ocr_digits(DPI_MAIN, _ocr_angle, 175, False, _psm)
            awb_400_m = extract_awb_from_400_pattern(txt_m)
            if awb_400_m:
                timings["ocr_main_ms"] = round((time.perf_counter() - main_start) * 1000, 1)
                complete_match(awb_400_m, f"OCR-Main-PSM{_psm}-400", "Matched by OCR-main 400 pattern")
                return "MATCHED"
            cr = run_clean_priority_gate(txt_m, f"OCR-Main-PSM{_psm}")
            if cr["status"] == "matched":
                timings["ocr_main_ms"] = round((time.perf_counter() - main_start) * 1000, 1)
                complete_match(cr["awb"], f"OCR-Main-PSM{_psm}-{cr['method']}",
                               "Matched exact clean in OCR-main")
                return "MATCHED"
            if cr["status"] == "tie":
                timings["ocr_main_ms"] = round((time.perf_counter() - main_start) * 1000, 1)
                send_review(f"Ambiguous OCR-main PSM{_psm} clean tie: {cr.get('ties', [])[:8]}",
                            f"OCR-Main-PSM{_psm}-{cr['method']}")
                return "NEEDS_REVIEW"
            hm, sm = extract_tiered_candidates(txt_m, awb_set)
            merge_stage_candidates(hm, sm, f"OCR-Main-PSM{_psm}")
            snapshot(f"OCR-Main-PSM{_psm}", hm | sm)
            res = run_exact_priority()
            if res["status"] == "matched":
                timings["ocr_main_ms"] = round((time.perf_counter() - main_start) * 1000, 1)
                complete_match(res["awb"], f"OCR-Main-PSM{_psm}-{res['method']}",
                               "Matched exact in OCR-main")
                return "MATCHED"
            if res["status"] == "tie":
                timings["ocr_main_ms"] = round((time.perf_counter() - main_start) * 1000, 1)
                send_review(f"Ambiguous OCR-main PSM{_psm} exact tie: {res.get('ties', [])[:8]}",
                            f"OCR-Main-PSM{_psm}-{res['method']}")
                return "NEEDS_REVIEW"
            # Skip PSM11 if PSM6 found nothing and earlier stages have quality candidates
            if _psm == 6 and _has_quality_candidates() and not (hm | sm):
                log("[FAST] Skipping OCR-Main PSM11 — PSM6 empty, quality candidates already present")
                break
    
        # Soft text pass on OCR-Main
        txt_ms = get_ocr_text(DPI_MAIN, _ocr_angle, False, 11)
        if not _has_quality_candidates() or not (running_high | running_standard):
            awb_400_ms = extract_awb_from_400_pattern(txt_ms)
            if awb_400_ms:
                timings["ocr_main_ms"] = round((time.perf_counter() - main_start) * 1000, 1)
                complete_match(awb_400_ms, "OCR-Main-Soft-400", "Matched by OCR-main soft 400 pattern")
                return "MATCHED"
            hms, sms = extract_tiered_candidates(txt_ms, awb_set)
            merge_stage_candidates(hms, sms, "OCR-Main-Soft")
            snapshot("OCR-Main-Soft", hms | sms)
    
        res = run_full_priority()
        timings["ocr_main_ms"] = round((time.perf_counter() - main_start) * 1000, 1)
        if res["status"] == "matched":
            complete_match(res["awb"], f"OCR-Main-{res['method']}", "Matched after OCR-main sequence")
            return "MATCHED"
        if res["status"] == "tie":
            send_review(f"Ambiguous OCR-main priority tie: {res.get('ties', [])[:8]}",
                        f"OCR-Main-{res['method']}")
            return "NEEDS_REVIEW"
    
        # ─────────────────────────────────────────────────────────────────────────
        # STAGE 3 — OCR STRONG at 0°
        # ─────────────────────────────────────────────────────────────────────────
        strong_start = time.perf_counter()
        strong_subpasses = [
            ("OCR-Strong-PSM6",  170, False, 6),
            ("OCR-Strong-PSM11", 170, False, 11),
        ]
        # Only add invert passes if normal passes yielded nothing useful
        _run_strong_invert = not _has_quality_candidates()
    
        for stage_nm, thr, inv, psm in strong_subpasses:
            txt_s = get_ocr_digits(DPI_STRONG, 0, thr, inv, psm)
            awb_400_s = extract_awb_from_400_pattern(txt_s)
            if awb_400_s:
                timings["ocr_strong_ms"] = round((time.perf_counter() - strong_start) * 1000, 1)
                complete_match(awb_400_s, f"{stage_nm}-400", f"Matched by {stage_nm} 400 pattern")
                return "MATCHED"
            cr = run_clean_priority_gate(txt_s, stage_nm)
            if cr["status"] == "matched":
                timings["ocr_strong_ms"] = round((time.perf_counter() - strong_start) * 1000, 1)
                complete_match(cr["awb"], f"{stage_nm}-{cr['method']}", f"Matched clean exact in {stage_nm}")
                return "MATCHED"
            if cr["status"] == "tie":
                timings["ocr_strong_ms"] = round((time.perf_counter() - strong_start) * 1000, 1)
                send_review(f"Ambiguous {stage_nm} clean tie: {cr.get('ties', [])[:8]}",
                            f"{stage_nm}-{cr['method']}")
                return "NEEDS_REVIEW"
            hs, ss = extract_tiered_candidates(txt_s, awb_set)
            merge_stage_candidates(hs, ss, stage_nm)
            snapshot(stage_nm, hs | ss)
            res = run_exact_priority()
            if res["status"] == "matched":
                timings["ocr_strong_ms"] = round((time.perf_counter() - strong_start) * 1000, 1)
                complete_match(res["awb"], f"{stage_nm}-{res['method']}", f"Matched exact in {stage_nm}")
                return "MATCHED"
            if res["status"] == "tie":
                timings["ocr_strong_ms"] = round((time.perf_counter() - strong_start) * 1000, 1)
                send_review(f"Ambiguous {stage_nm} exact tie: {res.get('ties', [])[:8]}",
                            f"{stage_nm}-{res['method']}")
                return "NEEDS_REVIEW"
            if psm == 6:
                _run_strong_invert = not _has_quality_candidates()
    
        if _run_strong_invert:
            for stage_nm, thr, inv, psm in [
                ("OCR-Strong-Invert-PSM6",  200, True, 6),
                ("OCR-Strong-Invert-PSM11", 200, True, 11),
            ]:
                txt_si = get_ocr_digits(DPI_STRONG, 0, thr, inv, psm)
                awb_400_si = extract_awb_from_400_pattern(txt_si)
                if awb_400_si:
                    timings["ocr_strong_ms"] = round((time.perf_counter() - strong_start) * 1000, 1)
                    complete_match(awb_400_si, f"{stage_nm}-400", f"Matched by {stage_nm} 400 pattern")
                    return "MATCHED"
                cr = run_clean_priority_gate(txt_si, stage_nm)
                if cr["status"] == "matched":
                    timings["ocr_strong_ms"] = round((time.perf_counter() - strong_start) * 1000, 1)
                    complete_match(cr["awb"], f"{stage_nm}-{cr['method']}", f"Matched in {stage_nm}")
                    return "MATCHED"
                if cr["status"] == "tie":
                    timings["ocr_strong_ms"] = round((time.perf_counter() - strong_start) * 1000, 1)
                    send_review(f"Ambiguous {stage_nm} clean tie: {cr.get('ties', [])[:8]}",
                                f"{stage_nm}-{cr['method']}")
                    return "NEEDS_REVIEW"
                hsi, ssi = extract_tiered_candidates(txt_si, awb_set)
                merge_stage_candidates(hsi, ssi, stage_nm)
                snapshot(stage_nm, hsi | ssi)
                res = run_exact_priority()
                if res["status"] == "matched":
                    timings["ocr_strong_ms"] = round((time.perf_counter() - strong_start) * 1000, 1)
                    complete_match(res["awb"], f"{stage_nm}-{res['method']}", f"Matched exact in {stage_nm}")
                    return "MATCHED"
                if res["status"] == "tie":
                    timings["ocr_strong_ms"] = round((time.perf_counter() - strong_start) * 1000, 1)
                    send_review(f"Ambiguous {stage_nm} exact tie: {res.get('ties', [])[:8]}",
                                f"{stage_nm}-{res['method']}")
                    return "NEEDS_REVIEW"
    
        # Strong soft pass
        if not _has_quality_candidates() or len(running_high) < 2:
            txt_3e = get_ocr_text(DPI_STRONG, 0, False, 11)
            h3e, s3e = extract_tiered_candidates(txt_3e, awb_set)
            if not (h3e or s3e):
                txt_3e2 = get_ocr_text(DPI_STRONG, 0, False, 6)
                h3e2, s3e2 = extract_tiered_candidates(txt_3e2, awb_set)
                h3e.update(h3e2); s3e.update(s3e2)
            merge_stage_candidates(h3e, s3e, "OCR-Strong-Soft")
            snapshot("OCR-Strong-Soft", h3e | s3e)
            if not (h3e or s3e):
                box_cands = extract_candidates_from_ocr_data(
                    preprocess_for_text(get_image(DPI_STRONG, 0), invert=False)
                )
                if box_cands:
                    merge_stage_candidates(set(), box_cands, "OCR-Strong-Boxes")
    
        res = run_full_priority()
        timings["ocr_strong_ms"] = round((time.perf_counter() - strong_start) * 1000, 1)
        if res["status"] == "matched":
            complete_match(res["awb"], f"OCR-Strong-{res['method']}", "Matched after OCR-strong sequence")
            return "MATCHED"
        if res["status"] == "tie":
            send_review(f"Ambiguous OCR-strong priority tie: {res.get('ties', [])[:8]}",
                        f"OCR-Strong-{res['method']}")
            return "NEEDS_REVIEW"
    
        # ── FAST-LANE EXIT ────────────────────────────────────────────────────────
        # Defer immediately after Stage 3 fails — probe, ROI, and all rescue stages
        # run only in long-pass so the fast lane drains the inbox as quickly as possible.
        # The deferred file goes through the full pipeline unchanged in long-pass.
        if not allow_long_pass:
            log(f"[FAST-LANE] Deferred after Stage 3 (no match at 0°): {name}")
            close_pdf()
            return "DEFERRED"
    
        # ─────────────────────────────────────────────────────────────────────────
        # STAGE 3.1 — ROTATION PROBE (runs only when Stages 2-3 failed at 0°)
        # ─────────────────────────────────────────────────────────────────────────
        probe_img = get_image(ROTATION_PROBE_DPI, 0)
        # For image-only documents in long-pass where pre-checks strongly indicated
        # rotation, narrow the probe to 0° + hint angle only — saves 2 redundant
        # OCR calls without changing downstream behavior (unprobed angles get score=0).
        # Never applied in fast-lane or for documents with a text layer.
        _probe_angles = ALLOWED_ROTATION_ANGLES
        if (
            allow_long_pass            # only in long-pass, never fast-lane
            and _is_image_only
            and _rotation_hint in (90, 180, 270)
        ):
            _probe_angles = (0, _rotation_hint)
            log(f"[ROTATION-PROBE] Narrowed probe to {_probe_angles} (image-only + hint={_rotation_hint}°)")
        base_angle, probe_scores, probe_texts = rotation_probe_best(
            probe_img, return_scores=True, preferred_angles=_probe_angles
        )
    
        if base_angle not in ALLOWED_ROTATION_ANGLES:
            base_angle = 0
    
        # Override with pre-angle detection hint if probe was uncertain
        if _rotation_hint is not None and base_angle == 0:
            _best_score = probe_scores.get(base_angle, 0)
            _second = max(v for k, v in probe_scores.items() if k != base_angle)
            if (_best_score - _second) < ROTATION_PROBE_MIN_FLIP_MARGIN:
                base_angle = _rotation_hint
                log(f"[ANGLE-DETECT] Pre-check hint overrides uncertain probe → {base_angle}°")
    
        score_view = {k: int(v) for k, v in sorted(probe_scores.items())}
        if base_angle:
            log(f"[ROTATION-PROBE] Base angle {base_angle}° selected | scores={score_view}")
        else:
            log(f"[ROTATION-PROBE] No rotation needed (0deg) | scores={score_view}")
    
        # Probe text early exit (free check using low-DPI OCR already done)
        # general_txt first so label/context words are present for the extractor,
        # then digit_txt to contribute cleaner raw number strings.
        _probe_digit_txt, _probe_general_txt = probe_texts.get(base_angle, ("", ""))
    _probe_combined_txt = "\n".join(
        part for part in (_probe_general_txt, _probe_digit_txt) if part
    )
    if _probe_combined_txt:
        # 400 tight-prefix check — both sources, no DB check (same exemption as elsewhere)
        _probe_awb_400 = (
            extract_awb_from_400_pattern(_probe_digit_txt)
            or extract_awb_from_400_pattern(_probe_general_txt)
        )
        if _probe_awb_400:
            complete_match(_probe_awb_400, "Probe-400", "Matched via probe text 400 pattern")
            return "MATCHED"
        # Exact-high check on combined text — digit OCR may have seen the number
        # more cleanly than general OCR; both sources contribute candidates
        _ph, _ps = extract_tiered_candidates(_probe_combined_txt, awb_set)
        _ph_db = sorted(_ph & awb_set)
        if len(_ph_db) == 1:
            complete_match(_ph_db[0], "Probe-Exact-High", "Matched via probe combined exact high")
            return "MATCHED"

    # Angle certainty tiers
    _margin = probe_scores.get(base_angle, 0) - max(
        (v for k, v in probe_scores.items() if k != base_angle), default=0
    )
    if _margin >= ROTATION_PROBE_CERTAIN_MARGIN:
        _angle_certainty = "CERTAIN"
    elif _margin >= ROTATION_PROBE_LIKELY_MARGIN:
        _angle_certainty = "LIKELY"
    else:
        _angle_certainty = "UNCERTAIN"
    log(f"[ROTATION-PROBE] certainty={_angle_certainty} margin={_margin}")

    # ─────────────────────────────────────────────────────────────────────────
    # ROUTING DECISION
    # ─────────────────────────────────────────────────────────────────────────
    _probe_confident_upright = (base_angle == 0)
    # Route A/B: upright — cheap fixes before expensive rotation
    # Route C:   rotated — rotation first
    _route = "UPRIGHT" if _probe_confident_upright else "ROTATED"
    log(f"[ROUTE] {_route} (base_angle={base_angle}°, image_only={_is_image_only})")

    # ─────────────────────────────────────────────────────────────────────────
    # FAST-LANE DEFER POINT (two-pass scheduling)
    # ─────────────────────────────────────────────────────────────────────────
    # For fast lane, run Stage 3.5 ROI then defer everything else
    # (probe has now been computed and stored in deferred metadata)

    def _run_roi_pass(src_img, stage_name):
        w, h = src_img.size
        y1 = max(0, int(h * 0.10))
        y2 = min(h, int(h * 0.62))
        if y2 <= y1 + 40:
            return False
        roi = src_img.crop((0, y1, w, y2))
        roi = _upscale(roi, 2)
        txt_roi = "\n".join([
            ocr_text_general(preprocess_for_text(roi, invert=False), psm=6),
            ocr_text_general(preprocess_for_text(roi, invert=False), psm=11),
            ocr_digits_only(preprocess(roi, thr=170, invert=False), psm=6),
        ])
        # Quick 400 check
        awb_400_roi = extract_awb_from_400_pattern(txt_roi)
        if awb_400_roi:
            complete_match(awb_400_roi, f"{stage_name}-400", "Matched by ROI 400 pattern")
            return True
        cr = run_clean_priority_gate(txt_roi, stage_name)
        if cr["status"] == "matched":
            complete_match(cr["awb"], f"{stage_name}-{cr['method']}", "Matched clean in ROI pass")
            return True
        if cr["status"] == "tie":
            send_review(f"Ambiguous ROI clean tie: {cr.get('ties', [])[:8]}",
                        f"{stage_name}-{cr['method']}")
            return True
        h_roi, s_roi = extract_tiered_candidates(txt_roi, awb_set)
        box_roi = extract_candidates_from_ocr_data(preprocess_for_text(roi, invert=False))
        if box_roi:
            s_roi.update(box_roi)
            s_roi.difference_update(h_roi)
        merge_stage_candidates(h_roi, s_roi, stage_name)
        snapshot(stage_name, h_roi | s_roi)
        res = run_exact_priority()
        if res["status"] == "matched":
            complete_match(res["awb"], f"{stage_name}-{res['method']}", "Matched exact in ROI pass")
            return True
        if res["status"] == "tie":
            send_review(f"Ambiguous ROI exact tie: {res.get('ties', [])[:8]}",
                        f"{stage_name}-{res['method']}")
            return True
        roi_unique = _unique_awb_candidate_count(h_roi | s_roi)
        if 0 < roi_unique <= 2:
            res = run_full_priority()
            if res["status"] == "matched":
                complete_match(res["awb"], f"{stage_name}-{res['method']}", "Matched in ROI full priority")
                return True
            if res["status"] == "tie":
                send_review(f"Ambiguous ROI priority tie: {res.get('ties', [])[:8]}",
                            f"{stage_name}-{res['method']}")
                return True
        return False

    # Stage 3.5 — ROI pass (both routes)
    roi_start = time.perf_counter()
    try:
        if _run_roi_pass(get_image(DPI_STRONG, base_angle), "OCR-ROI-ShipRow"):
            timings["ocr_context_ms"] += round((time.perf_counter() - roi_start) * 1000, 1)
            return "MATCHED"
        if _is_image_only:
            for _rot_roi in (90, 270):
                if _run_roi_pass(get_image(DPI_STRONG, _rot_roi), f"OCR-ROI-ShipRow-Rot{_rot_roi}"):
                    timings["ocr_context_ms"] += round((time.perf_counter() - roi_start) * 1000, 1)
                    return "MATCHED"
    except Exception as e:
        log(f"[ROI-PASS] Warning: {e}")
    timings["ocr_context_ms"] += round((time.perf_counter() - roi_start) * 1000, 1)

    # ─────────────────────────────────────────────────────────────────────────
    # ROUTE A/B — UPRIGHT: 5.5 upscale → 5 table → 4 rotation (last resort)
    # ROUTE C   — ROTATED: 4 rotation → 5 table → 5.5 upscale
    # ─────────────────────────────────────────────────────────────────────────

    def _run_upscale_rescue():
        # Only run if there is at most one HIGH-confidence persistent candidate.
        # Multiple persistent STANDARD noise numbers from bad rotation passes
        # must not block upscale rescue — they are noise, not evidence of ambiguity.
        rescue_trigger = [
            c for c in sorted(all_tried)
            if len(candidate_stage_hits.get(c, set())) >= 2
            and not _is_disqualified_candidate(c)
            and not _is_likely_date_reference(c)
            and candidate_confidence.get(c) == "HIGH"
        ]
        if len(rescue_trigger) > 1:
            return False  # multiple HIGH-confidence candidates — genuinely ambiguous
        rsc_start = time.perf_counter()
        try:
            base_src = table_clean_cache[0] if table_clean_cache[0] else get_image(DPI_STRONG, base_angle)
            upscaled = _upscale(base_src, 3)
            txt_rsc = "\n".join([
                ocr_text_general(preprocess_for_text(upscaled, invert=False), psm=6),
                ocr_text_general(preprocess_for_text(upscaled, invert=False), psm=11),
                ocr_digits_only(preprocess(upscaled, thr=170, invert=False), psm=6),
            ])
            awb_400_rsc = extract_awb_from_400_pattern(txt_rsc)
            if awb_400_rsc:
                timings["ocr_context_ms"] += round((time.perf_counter() - rsc_start) * 1000, 1)
                complete_match(awb_400_rsc, "OCR-Rescue-Upscaled-400", "Matched by upscale rescue 400")
                return True
            cr = run_clean_priority_gate(txt_rsc, "OCR-Rescue-Upscaled")
            if cr["status"] == "matched":
                timings["ocr_context_ms"] += round((time.perf_counter() - rsc_start) * 1000, 1)
                complete_match(cr["awb"], f"OCR-Rescue-Upscaled-{cr['method']}", "Matched in upscale rescue")
                return True
            if cr["status"] == "tie":
                timings["ocr_context_ms"] += round((time.perf_counter() - rsc_start) * 1000, 1)
                send_review(f"Upscale rescue tie: {cr.get('ties', [])[:8]}", f"OCR-Rescue-Upscaled-{cr['method']}")
                return True
            h_rsc, s_rsc = extract_tiered_candidates(txt_rsc, awb_set)
            try:
                box_rsc = extract_candidates_from_ocr_data(preprocess_for_text(upscaled, invert=False))
                box_rsc = {c for c in box_rsc if not _is_disqualified_candidate(c) and not _is_likely_date_reference(c)}
                if box_rsc:
                    if len(box_rsc) <= 2:
                        h_rsc.update(box_rsc)
                    else:
                        s_rsc.update(box_rsc)
                    s_rsc.difference_update(h_rsc)
            except Exception:
                pass
            merge_stage_candidates(h_rsc, s_rsc, "OCR-Rescue-Upscaled")
            snapshot("OCR-Rescue-Upscaled", h_rsc | s_rsc)
            res = run_full_priority()
            timings["ocr_context_ms"] += round((time.perf_counter() - rsc_start) * 1000, 1)
            if res["status"] == "matched":
                complete_match(res["awb"], f"OCR-Rescue-Upscaled-{res['method']}", "Matched by upscale rescue")
                return True
            if res["status"] == "tie":
                send_review(f"Upscale rescue priority tie: {res.get('ties', [])[:8]}",
                            f"OCR-Rescue-Upscaled-{res['method']}")
                return True
        except Exception as e:
            log(f"[RESCUE-UPSCALED] Warning: {e}")
        return False

    def _run_table_pass():
        tbl_start = time.perf_counter()
        tbl_img = remove_table_lines_image(get_image(DPI_STRONG, base_angle))
        if tbl_img is None:
            log("[TABLE-PASS] cv2 unavailable — skipping table pass.")
            return False
        table_clean_cache[0] = tbl_img  # store for upscale rescue reuse
        txt_t = ocr_text_general(preprocess_for_text(tbl_img, invert=False), psm=3)
        awb_400_t = extract_awb_from_400_pattern(txt_t)
        if awb_400_t:
            timings["ocr_context_ms"] += round((time.perf_counter() - tbl_start) * 1000, 1)
            complete_match(awb_400_t, "OCR-Table-PSM3-400", "Matched by table pass 400 pattern")
            return True
        cr = run_clean_priority_gate(txt_t, "OCR-Table-PSM3")
        if cr["status"] == "matched":
            timings["ocr_context_ms"] += round((time.perf_counter() - tbl_start) * 1000, 1)
            complete_match(cr["awb"], f"OCR-Table-PSM3-{cr['method']}", "Matched clean in table pass")
            return True
        if cr["status"] == "tie":
            timings["ocr_context_ms"] += round((time.perf_counter() - tbl_start) * 1000, 1)
            send_review(f"Table pass clean tie: {cr.get('ties', [])[:8]}", f"OCR-Table-PSM3-{cr['method']}")
            return True
        ht, st = extract_tiered_candidates(txt_t, awb_set)
        try:
            box_t = {
                c for c in extract_candidates_from_ocr_data(preprocess_for_text(tbl_img, invert=False))
                if not _is_disqualified_candidate(c) and not _is_likely_date_reference(c)
            }
            if box_t:
                if len(box_t) <= 2:
                    ht.update(box_t)
                else:
                    st.update(box_t)
                st.difference_update(ht)
        except Exception:
            pass
        merge_stage_candidates(ht, st, "OCR-Table-PSM3")
        snapshot("OCR-Table-PSM3", ht | st)
        res = run_full_priority()
        timings["ocr_context_ms"] += round((time.perf_counter() - tbl_start) * 1000, 1)
        if res["status"] == "matched":
            complete_match(res["awb"], f"OCR-Table-{res['method']}", "Matched by table pass")
            return True
        if res["status"] == "tie":
            send_review(f"Ambiguous table-pass priority tie: {res.get('ties', [])[:8]}",
                        f"OCR-Table-{res['method']}")
            return True
        return False

    def _run_rotation_passes():
        if not ENABLE_ROTATION_LAST_RESORT:
            return False
        rot_start = time.perf_counter()

        # Build angle order using probe scores
        remaining = [r for r in [90, 180, 270, 0] if r != base_angle]
        if _angle_certainty == "CERTAIN":
            # Certain of base_angle — defer other angles to final fallback
            timings["rotation_ms"] += round((time.perf_counter() - rot_start) * 1000, 1)
            return False
        elif _angle_certainty == "LIKELY":
            # Only try probe angle first; others deferred
            rotation_order = [base_angle] if base_angle != 0 else []
        else:
            rotation_order = sorted(
                remaining,
                key=lambda r: probe_scores.get(r, 0),
                reverse=True,
            )
            if base_angle != 0:
                rotation_order = [base_angle] + [r for r in rotation_order if r != base_angle]

        # Pre-angle hint prioritisation: if the zero-cost checks strongly indicated
        # a specific rotation and the probe was uncertain, put that angle first.
        # This does not remove any angle — only reorders so the most likely angle
        # is tried first, reducing time-to-hit on clearly rotated documents.
        if (
            _rotation_hint in (90, 180, 270)
            and _rotation_hint in rotation_order
            and _angle_certainty == "UNCERTAIN"
        ):
            rotation_order = [_rotation_hint] + [
                a for a in rotation_order if a != _rotation_hint
            ]
            log(f"[ROTATION] Pre-angle hint {_rotation_hint}° moved to front of rotation order")

        for rot in rotation_order:
            rimg = get_image(DPI_STRONG, rot)
            rot_subpasses = [
                (f"OCR-Rotation-{rot}-PSM6",   170, False, 6),
                (f"OCR-Rotation-{rot}-PSM11",  170, False, 11),
            ]
            _run_rot_invert = not _has_quality_candidates()
            if _run_rot_invert:
                rot_subpasses += [
                    (f"OCR-Rotation-{rot}-Invert-PSM6",  200, True, 6),
                    (f"OCR-Rotation-{rot}-Invert-PSM11", 200, True, 11),
                ]

            for stage_nm, thr, inv, psm in rot_subpasses:
                txt_r = get_ocr_digits(DPI_STRONG, rot, thr, inv, psm)
                awb_400_r = extract_awb_from_400_pattern(txt_r)
                if awb_400_r:
                    timings["rotation_ms"] += round((time.perf_counter() - rot_start) * 1000, 1)
                    complete_match(awb_400_r, f"{stage_nm}-400", f"Matched by {stage_nm} 400 pattern")
                    return True
                cr = run_clean_priority_gate(txt_r, stage_nm)
                if cr["status"] == "matched":
                    timings["rotation_ms"] += round((time.perf_counter() - rot_start) * 1000, 1)
                    complete_match(cr["awb"], f"{stage_nm}-{cr['method']}", f"Matched clean in {stage_nm}")
                    return True
                if cr["status"] == "tie":
                    timings["rotation_ms"] += round((time.perf_counter() - rot_start) * 1000, 1)
                    send_review(f"Ambiguous {stage_nm} clean tie: {cr.get('ties', [])[:8]}",
                                f"{stage_nm}-{cr['method']}")
                    return True
                hr, sr = extract_tiered_candidates(txt_r, awb_set)
                merge_stage_candidates(hr, sr, stage_nm)
                snapshot(stage_nm, hr | sr)
                res = run_exact_priority()
                if res["status"] == "matched":
                    timings["rotation_ms"] += round((time.perf_counter() - rot_start) * 1000, 1)
                    complete_match(res["awb"], f"{stage_nm}-{res['method']}", f"Matched exact in {stage_nm}")
                    return True
                if res["status"] == "tie":
                    timings["rotation_ms"] += round((time.perf_counter() - rot_start) * 1000, 1)
                    send_review(f"Ambiguous {stage_nm} exact tie: {res.get('ties', [])[:8]}",
                                f"{stage_nm}-{res['method']}")
                    return True
                if psm == 6:
                    _run_rot_invert = not _has_quality_candidates()

            # Rotation soft pass
            txt_rs = get_ocr_text(DPI_STRONG, rot, False, 11)
            awb_400_rs = extract_awb_from_400_pattern(txt_rs)
            if awb_400_rs:
                timings["rotation_ms"] += round((time.perf_counter() - rot_start) * 1000, 1)
                complete_match(awb_400_rs, f"OCR-Rotation-{rot}-Soft-400", "Matched by rotation soft 400")
                return True
            hrs, srs = extract_tiered_candidates(txt_rs, awb_set)
            merge_stage_candidates(hrs, srs, f"OCR-Rotation-{rot}-Soft")
            snapshot(f"OCR-Rotation-{rot}-Soft", hrs | srs)
            if not (hrs or srs):
                box_r = extract_candidates_from_ocr_data(preprocess_for_text(rimg, invert=False))
                if box_r:
                    merge_stage_candidates(set(), box_r, f"OCR-Rotation-{rot}-Boxes")
            res = run_full_priority()
            if res["status"] == "matched":
                timings["rotation_ms"] += round((time.perf_counter() - rot_start) * 1000, 1)
                complete_match(res["awb"], f"OCR-Rotation-{rot}-{res['method']}", f"Matched after rotation {rot}°")
                return True
            if res["status"] == "tie":
                timings["rotation_ms"] += round((time.perf_counter() - rot_start) * 1000, 1)
                send_review(f"Ambiguous rotation {rot}° priority tie: {res.get('ties', [])[:8]}",
                            f"OCR-Rotation-{rot}-{res['method']}")
                return True

            # Angle complete — check budget before starting next angle
            _check_timeout()

        timings["rotation_ms"] += round((time.perf_counter() - rot_start) * 1000, 1)
        return False

    # ─────────────────────────────────────────────────────────────────────────
    # ROUTE EXECUTION — wrapped so timeout captures all accumulated state
    # ─────────────────────────────────────────────────────────────────────────
    try:
            # Execute routes
            if _route == "UPRIGHT":
                # Stage 5.5 → 5 → 4 (last resort)
                if ENABLE_UPSCALED_RESCUE_PASS and _run_upscale_rescue():
                    return "MATCHED"
                if _run_table_pass():
                    return "MATCHED"
                if _run_rotation_passes():
                    return "MATCHED"
            else:
                # ROTATED: Stage 4 → 5 → 5.5
                if _run_rotation_passes():
                    return "MATCHED"
                if _run_table_pass():
                    return "MATCHED"
                if ENABLE_UPSCALED_RESCUE_PASS and _run_upscale_rescue():
                    return "MATCHED"

            # Final angle fallback for CERTAIN/LIKELY — try deferred angles now
            if _angle_certainty in ("CERTAIN", "LIKELY") and ENABLE_ROTATION_LAST_RESORT:
                _deferred = sorted(
                    [r for r in [90, 180, 270, 0] if r != base_angle],
                    key=lambda r: probe_scores.get(r, 0),
                    reverse=True,
                )
                rot_fb_start = time.perf_counter()
                for rot in _deferred:
                    rimg = get_image(DPI_STRONG, rot)
                    for stage_nm, thr, inv, psm in [
                        (f"OCR-AngFallback-{rot}-PSM6",   170, False, 6),
                        (f"OCR-AngFallback-{rot}-PSM11",  170, False, 11),
                        (f"OCR-AngFallback-{rot}-Inv6",   200, True,  6),
                        (f"OCR-AngFallback-{rot}-Inv11",  200, True,  11),
                    ]:
                        txt_fb = get_ocr_digits(DPI_STRONG, rot, thr, inv, psm)
                        awb_400_fb = extract_awb_from_400_pattern(txt_fb)
                        if awb_400_fb:
                            timings["rotation_ms"] += round((time.perf_counter() - rot_fb_start) * 1000, 1)
                            complete_match(awb_400_fb, f"{stage_nm}-400", "Matched in angle fallback")
                            return "MATCHED"
                        cr = run_clean_priority_gate(txt_fb, stage_nm)
                        if cr["status"] == "matched":
                            timings["rotation_ms"] += round((time.perf_counter() - rot_fb_start) * 1000, 1)
                            complete_match(cr["awb"], f"{stage_nm}-{cr['method']}", "Matched clean in angle fallback")
                            return "MATCHED"
                        if cr["status"] == "tie":
                            timings["rotation_ms"] += round((time.perf_counter() - rot_fb_start) * 1000, 1)
                            send_review(f"Angle fallback tie: {cr.get('ties', [])[:8]}", f"{stage_nm}-{cr['method']}")
                            return "NEEDS_REVIEW"
                        hfb, sfb = extract_tiered_candidates(txt_fb, awb_set)
                        merge_stage_candidates(hfb, sfb, stage_nm)
                        res = run_exact_priority()
                        if res["status"] == "matched":
                            timings["rotation_ms"] += round((time.perf_counter() - rot_fb_start) * 1000, 1)
                            complete_match(res["awb"], f"{stage_nm}-{res['method']}", "Matched in angle fallback")
                            return "MATCHED"
                        if res["status"] == "tie":
                            timings["rotation_ms"] += round((time.perf_counter() - rot_fb_start) * 1000, 1)
                            send_review(f"Angle fallback exact tie: {res.get('ties', [])[:8]}",
                                        f"{stage_nm}-{res['method']}")
                            return "NEEDS_REVIEW"
                timings["rotation_ms"] += round((time.perf_counter() - rot_fb_start) * 1000, 1)

                # ─────────────────────────────────────────────────────────────────────────
            # STAGE 5.6 — AIRWAY LABEL RESCUE (image-only or no quality candidates yet)
            # ─────────────────────────────────────────────────────────────────────────
            _run_airway = ENABLE_AIRWAY_LABEL_RESCUE and (
            _is_image_only or (base_angle in (90, 270)) or not _has_quality_candidates()
            )
            if _run_airway:
                label_start = time.perf_counter()
                _MAX_LABEL_RESCUE_MS = int(getattr(config, "MAX_CONTEXT_RESCUE_MS", 60000))
                try:
                    rot_order = []
                    for r in (base_angle, (base_angle + 180) % 360, 0):
                        if r not in rot_order:
                            rot_order.append(r)

                    for rot in rot_order:
                        # Time budget guard — stop if we have already spent enough time
                        if (time.perf_counter() - label_start) * 1000 > _MAX_LABEL_RESCUE_MS:
                            log(f"[AIRWAY-LABEL] Time budget ({_MAX_LABEL_RESCUE_MS}ms) reached — stopping rescue")
                            break
                        src = get_image(DPI_STRONG, rot)
                        w_src, h_src = src.size
                        crops = [
                            ("RightMid",   (int(w_src * 0.50), int(h_src * 0.24), w_src, int(h_src * 0.62))),
                            ("UpperRight", (int(w_src * 0.40), int(h_src * 0.05), w_src, int(h_src * 0.45))),
                            ("RightWide",  (int(w_src * 0.32), int(h_src * 0.12), w_src, int(h_src * 0.70))),
                        ]
                        for crop_name, box in crops:
                            x1, y1, x2, y2 = box
                            if x2 <= x1 + 30 or y2 <= y1 + 30:
                                continue
                            crop = src.crop((x1, y1, x2, y2))
                            crop = _upscale(crop, 3)
                            txt_lbl = "\n".join([
                                ocr_text_general(preprocess_for_text(crop, invert=False), psm=6),
                                ocr_text_general(preprocess_for_text(crop, invert=False), psm=11),
                                ocr_text_general(preprocess_for_text(crop, invert=False), psm=7),
                                ocr_digits_only(preprocess(crop, thr=160, invert=False), psm=7),
                                ocr_digits_only(preprocess(crop, thr=170, invert=False), psm=6),
                            ])
                            cr = run_clean_priority_gate(txt_lbl, f"OCR-AirwayLabel-Rot{rot}-{crop_name}")
                            if cr["status"] == "matched":
                                timings["ocr_context_ms"] += round((time.perf_counter() - label_start) * 1000, 1)
                                complete_match(cr["awb"],
                                               f"OCR-AirwayLabel-Rot{rot}-{crop_name}-{cr['method']}",
                                               "Matched clean in airway-label rescue")
                                return "MATCHED"
                            if cr["status"] == "tie":
                                timings["ocr_context_ms"] += round((time.perf_counter() - label_start) * 1000, 1)
                                send_review(f"Airway-label rescue clean tie: {cr.get('ties', [])[:8]}",
                                            f"OCR-AirwayLabel-Rot{rot}-{crop_name}-{cr['method']}")
                                return "NEEDS_REVIEW"
                            h_l, s_l = extract_tiered_candidates(txt_lbl, awb_set)
                            merge_stage_candidates(h_l, s_l, f"OCR-AirwayLabel-Rot{rot}-{crop_name}")
                            snapshot(f"OCR-AirwayLabel-Rot{rot}-{crop_name}", h_l | s_l)
                            res = run_exact_priority()
                            if res["status"] == "matched":
                                timings["ocr_context_ms"] += round((time.perf_counter() - label_start) * 1000, 1)
                                complete_match(res["awb"],
                                               f"OCR-AirwayLabel-Rot{rot}-{crop_name}-{res['method']}",
                                               "Matched by airway-label rescue (exact)")
                                return "MATCHED"
                            if res["status"] == "tie":
                                timings["ocr_context_ms"] += round((time.perf_counter() - label_start) * 1000, 1)
                                send_review(f"Airway-label rescue exact tie: {res.get('ties', [])[:8]}",
                                            f"OCR-AirwayLabel-Rot{rot}-{crop_name}-{res['method']}")
                                return "NEEDS_REVIEW"

                    # Guarded full priority if tiny stable label candidate set
                    recent_lbl = {
                        c for c in all_tried
                        if any("OCR-AirwayLabel-" in s for s in candidate_stage_hits.get(c, set()))
                    }
                    if 0 < len(recent_lbl) <= 2:
                        res = run_full_priority()
                        if res["status"] == "matched":
                            timings["ocr_context_ms"] += round((time.perf_counter() - label_start) * 1000, 1)
                            complete_match(res["awb"], f"OCR-AirwayLabel-{res['method']}",
                                           "Matched by airway-label rescue (guarded full priority)")
                            return "MATCHED"
                        if res["status"] == "tie":
                            timings["ocr_context_ms"] += round((time.perf_counter() - label_start) * 1000, 1)
                            send_review(f"Airway-label rescue priority tie: {res.get('ties', [])[:8]}",
                                        f"OCR-AirwayLabel-{res['method']}")
                            return "NEEDS_REVIEW"
                except Exception as e:
                    log(f"[AIRWAY-LABEL-RESCUE] Warning: {e}")
                timings["ocr_context_ms"] += round((time.perf_counter() - label_start) * 1000, 1)

                # ─────────────────────────────────────────────────────────────────────────
            # STAGE 6 — EDM PERSISTENCE FALLBACK (HIGH confidence only)
            # ─────────────────────────────────────────────────────────────────────────
            persistent = [
                c for c in sorted(all_tried)
                if (
                    len(c) == AWB_LEN and c.isdigit()
                    and not _is_disqualified_candidate(c)
                    and not _is_likely_date_reference(c)
                    and len(candidate_stage_hits.get(c, set())) >= 2
                    and candidate_confidence.get(c) == "HIGH"   # HIGH only — guards against noise
                )
                ]

            if len(persistent) == 1:
                edm_candidate = persistent[0]
                edm_exists = edm_awb_exists_fallback(edm_candidate)
                if edm_exists:
                    complete_match(edm_candidate, "EDM-Exists-Persistent",
                                   "Single HIGH-confidence persistent candidate confirmed by EDM")
                    return "MATCHED"
                log(f"[EDM-AWB-FALLBACK] Persistent candidate {edm_candidate} not confirmed by EDM.")
            elif len(persistent) > 1:
                send_review(f"EDM fallback tie across persistent candidates: {persistent[:8]}",
                            "EDM-Persistent-Tie")
                return "NEEDS_REVIEW"

            # ─────────────────────────────────────────────────────────────────────────
            # STAGE 7 — NEEDS REVIEW
            # ─────────────────────────────────────────────────────────────────────────
            send_review("No AWB match after exhausting all stages", "No-Match")
            return "NEEDS_REVIEW"

    except _TimeoutDeferred:
        # Capture all accumulated state so the third-pass can resume without
        # re-running any stage that already completed.
        _captured = {
            "probe_scores":           dict(probe_scores),
            "probe_texts":            {k: (v[0], v[1]) for k, v in probe_texts.items()},
            "base_angle":             base_angle,
            "_angle_certainty":       _angle_certainty,
            "_rotation_hint":         _rotation_hint,
            "_is_image_only":         _is_image_only,
            "running_high":           list(running_high),
            "running_standard":       list(running_standard),
            "candidate_stage_hits":   {k: list(v) for k, v in candidate_stage_hits.items()},
            "candidate_confidence":   dict(candidate_confidence),
            "all_tried":              list(all_tried),
            "quarantine":             {k: list(v) for k, v in quarantine.items()},
            # ocr_cache: serialise only the string values — PIL images are not picklable
            # but we store text results which are plain strings, safe to keep
            "ocr_cache":              {
                repr(k): v for k, v in ocr_cache.items()
                if isinstance(v, str)
            },
            "timings":                dict(timings),
        }
        if _state_out is not None:
            _state_out.update(_captured)
        close_pdf()
        return "TIMEOUT_DEFERRED"


# =============================================================================
# WATCHDOG HANDLER
# =============================================================================
class InboxPDFHandler(FileSystemEventHandler):
    def __init__(self, q):
        self.q = q
        self._last_seen = {}

    def _enqueue(self, path):
        p = str(path)
        if not p.lower().endswith(".pdf"):
            return
        now = time.time()
        if now - self._last_seen.get(p, 0) < 0.8:
            return
        self._last_seen[p] = now
        self.q.put(p)

    def on_created(self, event):
        if not event.is_directory:
            self._enqueue(event.src_path)

    def on_moved(self, event):
        if not event.is_directory:
            self._enqueue(event.dest_path)

    def on_modified(self, event):
        if not event.is_directory:
            self._enqueue(event.src_path)


# =============================================================================
# MAIN LOOP — TWO-PASS SCHEDULING
# =============================================================================
def main():
    config.ensure_dirs()
    require_tesseract()
    _reset_edm_exists_cache()

    startup_token = _get_edm_token()
    if not startup_token:
        log("[WARNING] No EDM token found; EDM fallback stage will be skipped.")
    else:
        log("[EDM] Token present at startup (expiry checked on first fallback call).")

    awb_set, by_prefix, by_suffix = set(), {}, {}
    last_excel_mtime = 0
    last_excel_load  = 0
    last_heartbeat   = 0
    last_rescan      = 0

    # Two-pass state
    deferred_long_pass    = []   # list of pdf paths deferred by fast lane
    timeout_deferred_state = {}  # path -> captured state dict (third-pass tier)

    file_queue = Queue()
    handler    = InboxPDFHandler(file_queue)
    observer   = Observer()
    observer.schedule(handler, str(INBOX_DIR), recursive=False)
    observer.start()

    log("=== AWB Hot Folder Pipeline started ===")
    log(f"INBOX:  {INBOX_DIR}")
    log(f"EXCEL:  {AWB_EXCEL_PATH}")
    log(f"LOGS:   {AWB_LOGS_PATH}")
    log("Scheduling: two-pass (fast lane = Stages 0-3 only, defer after Stage 3 fail; "
        "long-pass = full pipeline on deferred docs when fast queue empty)"
        if ENABLE_INBOX_TWO_PASS else "Scheduling: single-pass full pipeline")
    log("Hard-doc mode: OFF (fast primary path)")
    log("Mode: watchdog event-driven with periodic safety rescan")

    try:
        for fn in INBOX_DIR.iterdir():
            if fn.suffix.lower() == ".pdf":
                handler._enqueue(str(fn))
    except Exception as e:
        log(f"Startup scan warning: {e}")

    try:
        while True:
            loop_sleep = POLL_SECONDS
            try:
                now = time.time()

                # Refresh Excel
                if now - last_excel_load >= EXCEL_REFRESH_SECONDS:
                    try:
                        mtime = AWB_EXCEL_PATH.stat().st_mtime
                    except Exception:
                        mtime = 0
                    if mtime != last_excel_mtime:
                        awb_set = load_awb_set_from_excel(AWB_EXCEL_PATH)
                        by_prefix, by_suffix = build_buckets(awb_set)
                        last_excel_mtime = mtime
                        log(f"Loaded AWBs: {len(awb_set)} (Excel refreshed)")
                    last_excel_load = now

                # Heartbeat
                if now - last_heartbeat >= HEARTBEAT_SECONDS:
                    try:
                        fc = len([x for x in INBOX_DIR.iterdir() if x.suffix.lower() == ".pdf"])
                    except Exception:
                        fc = -1
                    log(f"Watching INBOX | PDF Files: {fc} | AWBs loaded: {len(awb_set)} | "
                        f"deferred-long-pass: {len(deferred_long_pass)} | "
                        f"timeout-deferred: {len(timeout_deferred_state)}")
                    last_heartbeat = now

                # Safety rescan
                if now - last_rescan >= max(POLL_SECONDS, 3):
                    try:
                        for fn in INBOX_DIR.iterdir():
                            if fn.suffix.lower() == ".pdf":
                                handler._enqueue(str(fn))
                    except Exception as e:
                        log(f"Rescan warning: {e}")
                    last_rescan = now

                processed_any = False

                # Drain fast-lane queue
                while True:
                    try:
                        path = file_queue.get_nowait()
                    except Empty:
                        break
                    if not os.path.exists(path) or not path.lower().endswith(".pdf"):
                        continue
                    if ENABLE_INBOX_TWO_PASS:
                        result = process_pdf(str(path), awb_set, by_prefix, by_suffix,
                                             allow_long_pass=False)
                        if result == "DEFERRED":
                            deferred_long_pass.append(str(path))
                    else:
                        process_pdf(str(path), awb_set, by_prefix, by_suffix,
                                    allow_long_pass=True)
                    processed_any = True

                # Process deferred long-pass when fast queue is empty
                if ENABLE_INBOX_TWO_PASS and file_queue.empty() and deferred_long_pass:
                    # Process one deferred file per cycle to stay responsive to new arrivals
                    path = deferred_long_pass.pop(0)
                    if os.path.exists(path):
                        log(f"[LONG-PASS] Processing deferred: {os.path.basename(path)}")
                        state_out = {}
                        result = process_pdf(
                            str(path), awb_set, by_prefix, by_suffix,
                            allow_long_pass=True,
                            timeout_seconds=LONG_PASS_TIMEOUT_SECONDS,
                            _state_out=state_out,
                        )
                        if result == "TIMEOUT_DEFERRED":
                            timeout_deferred_state[path] = state_out
                            log(f"[TIMEOUT-DEFERRED] {os.path.basename(path)} "
                                f"queued for third-pass "
                                f"(total queued: {len(timeout_deferred_state)})")
                        processed_any = True
                    # Immediately check if new fast files arrived
                    for fn in INBOX_DIR.iterdir():
                        if fn.suffix.lower() == ".pdf":
                            handler._enqueue(str(fn))

                # Third-pass: resume timeout-deferred files only when both queues
                # are empty — these are the genuinely hard documents that exceeded
                # the 75s budget in long-pass. Resume uses cached OCR/state so
                # no already-completed stage reruns.
                if (
                    ENABLE_INBOX_TWO_PASS
                    and file_queue.empty()
                    and not deferred_long_pass
                    and timeout_deferred_state
                ):
                    path, saved_state = next(iter(timeout_deferred_state.items()))
                    del timeout_deferred_state[path]
                    if os.path.exists(path):
                        log(f"[THIRD-PASS] Resuming: {os.path.basename(path)} "
                            f"(remaining in third-pass queue: {len(timeout_deferred_state)})")
                        # No timeout on third-pass — let it run to completion
                        process_pdf(
                            str(path), awb_set, by_prefix, by_suffix,
                            allow_long_pass=True,
                            timeout_seconds=None,
                            resume_state=saved_state,
                        )
                        processed_any = True
                    # Rescan inbox after each third-pass file
                    for fn in INBOX_DIR.iterdir():
                        if fn.suffix.lower() == ".pdf":
                            handler._enqueue(str(fn))

                if processed_any:
                    loop_sleep = 0.2

            except Exception as e:
                log(f"LOOP ERROR: {e}")

            time.sleep(loop_sleep)

    except KeyboardInterrupt:
        log("Shutting down hotfolder watcher...")
    finally:
        observer.stop()
        observer.join()


if __name__ == "__main__":
    main()
